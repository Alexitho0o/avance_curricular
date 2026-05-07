"""
kpi_dashboard.py
================
Construye / actualiza la hoja KPI_Dashboard en el XLSX de Matrícula Unificada.

Reglas que cumple este módulo
------------------------------
1.  Ningún KPI contiene valores estáticos: todos son fórmulas Excel recalculables.
2.  Los KPIs referencian `tbl_MU2026`, tabla nombrada definida sobre la hoja
    MATRICULA_UNIFICADA_32.
3.  Si la hoja KPI_Dashboard ya existe se respalda con timestamp y se reconstruye.
4.  Si la tabla tbl_MU2026 ya existe se elimina y se redefine sobre el rango actual.
5.  La función es autocontenida: no depende de openpyxl versiones muy antiguas.
6.  Ante cualquier excepción retorna un dict con status='error' sin propagar.
7.  Si `guardar_copia_backup=True` (por defecto False) escribe además un .xlsx de
    respaldo físico en el mismo directorio.

Uso desde el pipeline
---------------------
    from src.kpi_dashboard import construir_kpi_dashboard
    resultado = construir_kpi_dashboard(out_path)
    # resultado: {'status': 'creado'|'actualizado'|'respaldado_y_reemplazado'|'error',
    #             'message': str, 'backup_sheet': str|None, 'backup_file': str|None}
"""

from __future__ import annotations

import shutil
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any


# ────────────────────────────────────────────────────────────────────────────
# Constantes
# ────────────────────────────────────────────────────────────────────────────

SHEET_BASE = "MATRICULA_UNIFICADA_32"
TABLE_NAME = "tbl_MU2026"
SHEET_KPI = "KPI_Dashboard"

# Descripción de FOR_ING_ACT según catálogo SIES
FOR_ING_DESC: dict[int, str] = {
    1: "Enseñanza Media Nacional",
    2: "PSU / PAES",
    3: "Cambio Interno",
    4: "Cambio Externo",
    5: "Reconocimiento Aprendizajes Previos",
    6: "Estudiante Extranjero",
    7: "Convalidación",
    8: "Reincorporación",
    9: "Traslado",
    10: "Otras Vías",
    11: "Otra Forma Especial",
}

MODALIDAD_DESC: dict[int, str] = {
    1: "Presencial",
    2: "Semi-presencial",
    3: "A distancia",
}

JOR_DESC: dict[int, str] = {
    1: "Diurna",
    2: "Vespertina",
    3: "Nocturna",
    4: "Sin jornada definida",
}

# Colores en hex
COLOR_TITLE_BG = "1F3864"   # azul oscuro
COLOR_TITLE_FG = "FFFFFF"
COLOR_SECT_BG = "2E75B6"    # azul medio
COLOR_HDR_BG = "D6E4F7"     # azul claro
COLOR_ALT_ROW = "F0F7FF"    # azul muy claro fila par
COLOR_WHITE = "FFFFFF"
COLOR_LABEL_FG = "1F2D3A"
COLOR_TOTAL_BG = "FFF2CC"   # amarillo pálido para totales


# ────────────────────────────────────────────────────────────────────────────
# Función pública
# ────────────────────────────────────────────────────────────────────────────

def construir_kpi_dashboard(
    excel_path: Path,
    guardar_copia_backup: bool = False,
) -> dict[str, Any]:
    """Construye o reconstruye la hoja KPI_Dashboard en *excel_path*.

    Parámetros
    ----------
    excel_path : Path
        Ruta al archivo XLSX generado por el pipeline.
    guardar_copia_backup : bool
        Si True, escribe una copia física del XLSX antes de modificarlo.

    Retorna
    -------
    dict con claves:
        status       : 'creado' | 'actualizado' | 'respaldado_y_reemplazado' | 'error'
        message      : descripción del resultado
        backup_sheet : nombre de hoja de respaldo (str) o None
        backup_file  : ruta del archivo respaldo (str) o None
    """
    result: dict[str, Any] = {
        "status": "error",
        "message": "",
        "backup_sheet": None,
        "backup_file": None,
    }

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (
            Alignment,
            Font,
            PatternFill,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        result["message"] = f"openpyxl no disponible: {exc}"
        return result

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ts_display = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]} {ts[9:11]}:{ts[11:13]}:{ts[13:15]}"

    if not excel_path.exists():
        result["message"] = f"Archivo no encontrado: {excel_path}"
        return result

    # ── Copia de respaldo física (opcional) ─────────────────────────────────
    if guardar_copia_backup:
        bkp_file = excel_path.parent / f"{excel_path.stem}_bkp_{ts}{excel_path.suffix}"
        shutil.copy2(excel_path, bkp_file)
        result["backup_file"] = str(bkp_file)

    try:
        wb = load_workbook(excel_path)
    except Exception as exc:
        result["message"] = f"No se pudo abrir el workbook: {exc}"
        return result

    try:
        # ── 1. Verificar hoja base ───────────────────────────────────────────
        if SHEET_BASE not in wb.sheetnames:
            result["message"] = f"Hoja '{SHEET_BASE}' no encontrada en el workbook."
            wb.close()
            return result

        ws_base = wb[SHEET_BASE]
        n_data_rows = ws_base.max_row - 1   # excluyendo encabezado
        n_cols = ws_base.max_column         # 32
        col_last_letter = get_column_letter(n_cols)
        table_ref = f"A1:{col_last_letter}{ws_base.max_row}"

        # ── 2. Definir / actualizar tabla tbl_MU2026 ───────────────────────
        existing_tables = list(ws_base.tables.keys())
        for t_name in existing_tables:
            del ws_base.tables[t_name]

        tab = Table(displayName=TABLE_NAME, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws_base.add_table(tab)

        # ── 3. Leer valores únicos de columnas categóricas ──────────────────
        headers = [c.value for c in next(ws_base.iter_rows(min_row=1, max_row=1))]

        def col_index(name: str) -> int | None:
            try:
                return headers.index(name)
            except ValueError:
                return None

        def unique_sorted_counter(col_name: str) -> Counter:
            idx = col_index(col_name)
            if idx is None:
                return Counter()
            ctr: Counter = Counter()
            for row in ws_base.iter_rows(min_row=2, values_only=True):
                v = row[idx]
                if v is not None:
                    ctr[v] += 1
            return ctr

        cod_sed_ctr = unique_sorted_counter("COD_SED")
        cod_car_ctr = unique_sorted_counter("COD_CAR")
        for_ing_ctr = unique_sorted_counter("FOR_ING_ACT")
        modalidad_ctr = unique_sorted_counter("MODALIDAD")
        jor_ctr = unique_sorted_counter("JOR")

        # ── 4. Respaldo de la hoja KPI_Dashboard si existe ─────────────────
        kpi_existia = SHEET_KPI in wb.sheetnames
        backup_sheet_name: str | None = None
        if kpi_existia:
            backup_sheet_name = f"KPI_bkp_{ts}"[:31]
            src_ws = wb[SHEET_KPI]
            tgt_ws = wb.copy_worksheet(src_ws)
            tgt_ws.title = backup_sheet_name
            del wb[SHEET_KPI]
            result["backup_sheet"] = backup_sheet_name

        # ── 5. Crear hoja KPI_Dashboard (primera posición) ─────────────────
        ws = wb.create_sheet(SHEET_KPI, 0)

        # ── helpers de estilo ───────────────────────────────────────────────
        def _s_fill(hex_color: str) -> PatternFill:
            return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)

        def _thin_border():
            from openpyxl.styles import Border, Side
            thin = Side(border_style="thin", color="BFBFBF")
            return Border(bottom=thin)

        def _set_title(row: int, col: int, value: str, ncols: int = 7) -> None:
            ws.row_dimensions[row].height = 36
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + ncols - 1)
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=True, color=COLOR_TITLE_FG, size=16)
            c.fill = _s_fill(COLOR_TITLE_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")

        def _set_meta(row: int, value: str, ncols: int = 7) -> None:
            ws.row_dimensions[row].height = 18
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=value)
            c.font = Font(italic=True, size=9, color="595959")
            c.alignment = Alignment(horizontal="center", vertical="center")

        def _set_section(row: int, col: int, value: str, ncols: int = 3) -> None:
            ws.row_dimensions[row].height = 22
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + ncols - 1)
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=True, color=COLOR_TITLE_FG, size=11)
            c.fill = _s_fill(COLOR_SECT_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        def _set_col_hdr(row: int, col: int, value: str) -> None:
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(bold=True, size=10, color=COLOR_LABEL_FG)
            c.fill = _s_fill(COLOR_HDR_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")

        def _set_label(row: int, col: int, value: str, indent: bool = False) -> None:
            disp = ("    " if indent else "") + str(value)
            c = ws.cell(row=row, column=col, value=disp)
            c.font = Font(size=10, color=COLOR_LABEL_FG)
            bg = COLOR_ALT_ROW if row % 2 == 0 else COLOR_WHITE
            c.fill = _s_fill(bg)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=(1 if indent else 0))

        def _set_formula(row: int, col: int, formula: str, fmt: str = "#,##0", bold: bool = False) -> None:
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(bold=bold, size=10, color=COLOR_LABEL_FG)
            bg = COLOR_ALT_ROW if row % 2 == 0 else COLOR_WHITE
            c.fill = _s_fill(bg)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right", vertical="center")

        # ── anchos y alturas ────────────────────────────────────────────────
        col_widths = {
            "A": 38, "B": 16, "C": 10, "D": 2,
            "E": 30, "F": 16, "G": 10,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A4"

        # ═══════════════════════════════════════════════════════════════════
        # Fila 1 — Título
        # ═══════════════════════════════════════════════════════════════════
        _set_title(1, 1, "KPI Dashboard · Matrícula Unificada 2026")

        # Fila 2 — Metadatos
        _set_meta(
            2,
            f"Generado: {ts_display}  ·  Fuente: {TABLE_NAME} ({SHEET_BASE})  "
            f"·  {n_data_rows:,} registros",
        )

        # Fila 3 — espacio
        ws.row_dimensions[3].height = 6

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE IZQUIERDO: cols A-C  (KPI | Valor | %)
        # BLOQUE DERECHO:  cols E-G  (KPI | Valor | %)
        # ═══════════════════════════════════════════════════════════════════

        r = 4          # fila actual izquierda
        r_r = 4        # fila actual derecha

        # ── RESUMEN GENERAL ─────────────────────────────────────────────────
        _set_section(r, 1, "RESUMEN GENERAL", ncols=3)
        r += 1
        _set_col_hdr(r, 1, "KPI")
        _set_col_hdr(r, 2, "Valor")
        _set_col_hdr(r, 3, "%")
        r += 1
        total_row = r

        _set_label(r, 1, "Total registros (tbl_MU2026)")
        _set_formula(r, 2, f"=COUNTA({TABLE_NAME}[N_DOC])", bold=True)
        r += 1

        _set_label(r, 1, "Vigentes (VIG = 1)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[VIG],1)')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "Egresados (VIG = 2)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[VIG],2)')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "No vigentes (VIG = 0)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[VIG],0)')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "Hombres (SEXO = H)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[SEXO],"H")')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "Mujeres (SEXO = M)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[SEXO],"M")')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "No binario (SEXO = NB)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[SEXO],"NB")')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "Nacionales (NAC = 38)", indent=True)
        _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[NAC],38)')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        _set_label(r, 1, "Extranjeros (NAC ≠ 38)", indent=True)
        _set_formula(r, 2, f'=B{total_row}-COUNTIF({TABLE_NAME}[NAC],38)')
        _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
        r += 1

        # Resaltar fila total
        for col in (1, 2, 3):
            ws.cell(row=total_row, column=col).fill = _s_fill(COLOR_TOTAL_BG)
            ws.cell(row=total_row, column=col).font = Font(bold=True, size=10)

        r += 1  # separador

        # ── FORMA DE INGRESO (FOR_ING_ACT) ──────────────────────────────────
        _set_section(r, 1, "FORMA DE INGRESO ACTUAL (FOR_ING_ACT)", ncols=3)
        r += 1
        _set_col_hdr(r, 1, "Código")
        _set_col_hdr(r, 2, "Descripción")
        _set_col_hdr(r, 3, "n")
        r += 1

        # Mostrar solo los valores presentes más todos los catalogados
        all_for_ing = set(FOR_ING_DESC.keys()) | {int(k) for k in for_ing_ctr if str(k).isdigit()}
        for code in sorted(all_for_ing):
            desc = FOR_ING_DESC.get(code, f"Código {code}")
            _set_label(r, 1, str(code))
            _set_label(r, 2, desc)
            _set_formula(r, 3, f'=COUNTIF({TABLE_NAME}[FOR_ING_ACT],{code})')
            r += 1

        r += 1

        # ── SEDE (COD_SED) ───────────────────────────────────────────────────
        _set_section(r, 1, "SEDE (COD_SED)", ncols=3)
        r += 1
        _set_col_hdr(r, 1, "COD_SED")
        _set_col_hdr(r, 2, "n")
        _set_col_hdr(r, 3, "%")
        r += 1
        for sed_val in sorted(cod_sed_ctr.keys()):
            _set_label(r, 1, str(sed_val))
            _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[COD_SED],{sed_val})')
            _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
            r += 1

        r += 1

        # ── CARRERA (COD_CAR) ────────────────────────────────────────────────
        _set_section(r, 1, "CARRERA (COD_CAR) — ordenado por n desc", ncols=3)
        r += 1
        _set_col_hdr(r, 1, "COD_CAR")
        _set_col_hdr(r, 2, "n")
        _set_col_hdr(r, 3, "%")
        r += 1
        for car_val, _ in cod_car_ctr.most_common():
            _set_label(r, 1, str(car_val))
            _set_formula(r, 2, f'=COUNTIF({TABLE_NAME}[COD_CAR],{car_val})')
            _set_formula(r, 3, f"=IFERROR(B{r}/B{total_row},\"\")", fmt="0.0%")
            r += 1

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE DERECHO (cols E, F, G)
        # ═══════════════════════════════════════════════════════════════════

        # ── MODALIDAD ────────────────────────────────────────────────────────
        _set_section(r_r, 5, "MODALIDAD", ncols=3)
        r_r += 1
        _set_col_hdr(r_r, 5, "Código")
        _set_col_hdr(r_r, 6, "Descripción")
        _set_col_hdr(r_r, 7, "n")
        r_r += 1
        all_mod = set(MODALIDAD_DESC.keys()) | {int(k) for k in modalidad_ctr if str(k).isdigit()}
        for code in sorted(all_mod):
            desc = MODALIDAD_DESC.get(code, f"Código {code}")
            _set_label(r_r, 5, str(code))
            _set_label(r_r, 6, desc)
            _set_formula(r_r, 7, f'=COUNTIF({TABLE_NAME}[MODALIDAD],{code})')
            r_r += 1

        r_r += 1

        # ── JORNADA (JOR) ─────────────────────────────────────────────────────
        _set_section(r_r, 5, "JORNADA (JOR)", ncols=3)
        r_r += 1
        _set_col_hdr(r_r, 5, "Código")
        _set_col_hdr(r_r, 6, "Descripción")
        _set_col_hdr(r_r, 7, "n")
        r_r += 1
        all_jor = set(JOR_DESC.keys()) | {int(k) for k in jor_ctr if str(k).isdigit()}
        for code in sorted(all_jor):
            desc = JOR_DESC.get(code, f"Código {code}")
            _set_label(r_r, 5, str(code))
            _set_label(r_r, 6, desc)
            _set_formula(r_r, 7, f'=COUNTIF({TABLE_NAME}[JOR],{code})')
            r_r += 1

        r_r += 1

        # ── ANIO_ING_ACT ──────────────────────────────────────────────────────
        anio_ctr = unique_sorted_counter("ANIO_ING_ACT")
        _set_section(r_r, 5, "AÑO INGRESO ACTUAL (ANIO_ING_ACT)", ncols=3)
        r_r += 1
        _set_col_hdr(r_r, 5, "Año")
        _set_col_hdr(r_r, 6, "n")
        _set_col_hdr(r_r, 7, "%")
        r_r += 1
        for anio_val in sorted(anio_ctr.keys()):
            _set_label(r_r, 5, str(anio_val))
            _set_formula(r_r, 6, f'=COUNTIF({TABLE_NAME}[ANIO_ING_ACT],{anio_val})')
            _set_formula(r_r, 7, f"=IFERROR(F{r_r}/B{total_row},\"\")", fmt="0.0%")
            r_r += 1

        r_r += 1

        # ── NIV_ACA ───────────────────────────────────────────────────────────
        niv_ctr = unique_sorted_counter("NIV_ACA")
        _set_section(r_r, 5, "NIVEL ACADÉMICO (NIV_ACA)", ncols=3)
        r_r += 1
        _set_col_hdr(r_r, 5, "NIV_ACA")
        _set_col_hdr(r_r, 6, "n")
        _set_col_hdr(r_r, 7, "%")
        r_r += 1
        for niv_val in sorted(niv_ctr.keys()):
            _set_label(r_r, 5, str(niv_val))
            _set_formula(r_r, 6, f'=COUNTIF({TABLE_NAME}[NIV_ACA],{niv_val})')
            _set_formula(r_r, 7, f"=IFERROR(F{r_r}/B{total_row},\"\")", fmt="0.0%")
            r_r += 1

        # ══════════════════════════════════════════════════════════════════════
        # 6. Guardar (atómico vía tempfile)
        # ══════════════════════════════════════════════════════════════════════
        with tempfile.TemporaryDirectory(prefix="kpi_dash_") as tmpdir:
            tmp_path = Path(tmpdir) / excel_path.name
            wb.save(tmp_path)
            wb.close()
            if excel_path.exists():
                excel_path.unlink()
            shutil.copy2(tmp_path, excel_path)

        if kpi_existia:
            result["status"] = "respaldado_y_reemplazado"
            result["message"] = (
                f"KPI_Dashboard reemplazado. Respaldo en hoja '{backup_sheet_name}'."
            )
        else:
            result["status"] = "creado"
            result["message"] = f"KPI_Dashboard creado con {n_data_rows:,} registros."

    except Exception as exc:  # pylint: disable=broad-except
        result["status"] = "error"
        result["message"] = f"Error al construir KPI_Dashboard: {exc}"
        try:
            wb.close()
        except Exception:
            pass

    return result
