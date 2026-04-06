from __future__ import annotations

import argparse
import json
import re
import shutil
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd

# ==============================
# FUENTE ÚNICA GOBERNANZA SIES: DURACION_ESTUDIOS.tsv
# ==============================
def _parse_sies_codigo_unico(codigo_unico: object) -> tuple[str | None, str | None]:
    text = str(codigo_unico or "").strip().upper()
    m = re.match(r"^I\d+S\d+C\d+J(?P<jor>\d+)V(?P<ver>\d+)$", text)
    if not m:
        return (None, None)
    j = m.group("jor")
    jornada = {"1": "D", "2": "V", "3": "O", "4": "O"}.get(j)
    version = f"V{m.group('ver')}"
    return (jornada, version)


def _split_codcarpr_candidates(canon: object, aliases: object) -> list[str]:
    out: set[str] = set()
    canon_txt = str(canon or "").strip().upper()
    if canon_txt:
        out.add(canon_txt)
    alias_txt = str(aliases or "").strip().upper()
    if alias_txt:
        for val in alias_txt.split("|"):
            val = val.strip()
            if val:
                out.add(val)
    return sorted(out)


def _load_duracion_as_governance_df() -> pd.DataFrame:
    candidates = [
        Path(__file__).with_name("DURACION_ESTUDIOS.tsv"),
        Path.cwd() / "DURACION_ESTUDIOS.tsv",
    ]
    for path in candidates:
        if path.exists():
            try:
                df = pd.read_csv(path, sep="\t", dtype=str, keep_default_na=False)
                if "CODIGO_UNICO" in df.columns:
                    return df
            except Exception:
                continue
    return pd.DataFrame()


def _cargar_matriz_desambiguacion_desde_duracion() -> dict:
    """Construye matriz (CODCARPR, JORNADA, VERSION)->SIES desde DURACION_ESTUDIOS.tsv."""
    dur = _load_duracion_as_governance_df()
    matriz_dict: dict[tuple[str, str, str], tuple[str, str, str]] = {}
    if dur.empty:
        print("⚠️  DURACION_ESTUDIOS.tsv no disponible para construir matriz SIES")
        return matriz_dict

    conflictos = 0
    for row in dur.itertuples(index=False):
        jornada, version = _parse_sies_codigo_unico(getattr(row, "CODIGO_UNICO", ""))
        if not jornada or not version:
            continue
        codcarprs = _split_codcarpr_candidates(
            getattr(row, "CODCARPR_CANONICO", ""),
            getattr(row, "CODCARPR_ALIAS_LIST", ""),
        )
        for codcarpr in codcarprs:
            key = (codcarpr, jornada, version)
            val = (getattr(row, "CODIGO_UNICO", ""), "100%", "AUTO_DURACION_ESTUDIOS")
            if key in matriz_dict and matriz_dict[key][0] != val[0]:
                conflictos += 1
                continue
            matriz_dict[key] = val

    if conflictos:
        print(f"⚠️  Conflictos en matriz auto desde DURACION_ESTUDIOS: {conflictos} (se conserva primera ocurrencia)")
    print(f"✅ Matriz SIES auto construida desde DURACION_ESTUDIOS: {len(matriz_dict)} llaves")
    return matriz_dict


# Cargar matriz al iniciar desde DURACION_ESTUDIOS.tsv
MATRIZ_DESAMBIGUACION = _cargar_matriz_desambiguacion_desde_duracion()

# ==============================
# Contratos oficiales (Capa C)
# ==============================
MATRICULA_UNIFICADA_COLUMNS = [
    "TIPO_DOC", "N_DOC", "DV", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "NOMBRE", "SEXO", "FECH_NAC",
    "NAC", "PAIS_EST_SEC", "COD_SED", "COD_CAR", "MODALIDAD", "JOR", "VERSION", "FOR_ING_ACT",
    "ANIO_ING_ACT", "SEM_ING_ACT", "ANIO_ING_ORI", "SEM_ING_ORI", "ASI_INS_ANT", "ASI_APR_ANT",
    "PROM_PRI_SEM", "PROM_SEG_SEM", "ASI_INS_HIS", "ASI_APR_HIS", "NIV_ACA", "SIT_FON_SOL", "SUS_PRE",
    "FECHA_MATRICULA", "REINCORPORACION", "VIG",
]

CARRERAS_AC_COLUMNS = [
    "CODIGO_IES_NUM", "CODIGO_UNICO", "PLAN_ESTUDIOS", "NOMBRE_SEDE", "NOMBRE_CARRERA", "JORNADA", "VERSION",
    "DURACION_ESTUDIOS", "DURACION_TITULACION", "DURACION_TOTAL", "NIVEL_CARRERA", "TIPO_UNIDAD_MEDIDA",
    "OTRA_UNIDAD_MEDIDA", "TOTAL_UNIDADES_MEDIDA", "UNIDADES_1ER_ANIO", "UNIDADES_2DO_ANIO", "UNIDADES_3ER_ANIO",
    "UNIDADES_4TO_ANIO", "UNIDADES_5TO_ANIO", "UNIDADES_6TO_ANIO", "UNIDADES_7MO_ANIO", "VIGENCIA",
]

MATRICULA_AC_COLUMNS = [
    "CODIGO_IES_NUM", "TIPO_DOCUMENTO", "NUM_DOCUMENTO", "DV", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "NOMBRES",
    "SEXO", "FECHA_NACIMIENTO", "CODIGO_UNICO", "PLAN_ESTUDIOS", "ANIO_INGRESO_CARRERA_ACTUAL",
    "SEM_INGRESO_CARRERA_ACTUAL", "ANIO_INGRESO_CARRERA_ORIGEN", "SEM_INGRESO_CARRERA_ORIGEN",
    "CURSO_1ER_SEM", "CURSO_2DO_SEM", "UNIDADES_CURSADAS", "UNIDADES_APROBADAS", "UNID_CURSADAS_TOTAL",
    "UNID_APROBADAS_TOTAL", "VIGENCIA",
]

ANUAL_COLS = [
    "UNIDADES_1ER_ANIO", "UNIDADES_2DO_ANIO", "UNIDADES_3ER_ANIO", "UNIDADES_4TO_ANIO",
    "UNIDADES_5TO_ANIO", "UNIDADES_6TO_ANIO", "UNIDADES_7MO_ANIO",
]

MU_FUSION_OUTPUT_FILENAME = "archivo_listo_para_sies.xlsx"
MU_PREGRADO_CSV_FILENAME = "matricula_unificada_2026_pregrado.csv"
FINAL_SIES_CODE_COL = "CODIGO_CARRERA_SIES_FINAL"
MAX_SIES_CODES_PER_KEY = 5
DEFAULT_EXCLUIR_DIPLOMADOS = True

DEFAULT_INPUT_CANDIDATES = [
    Path.home() / "Downloads" / "PROMEDIOSDEALUMNOS_7804.xlsx",
]
DEFAULT_CATALOGO_MANUAL_CANDIDATES = [
    Path(__file__).with_name("catalogo_manual.tsv"),
    Path.cwd() / "catalogo_manual.tsv",
    Path.home() / "Downloads" / "catalogo_manual.tsv",
]
DEFAULT_PUENTE_SIES_CANDIDATES = [
    Path(__file__).with_name("puente_sies.tsv"),
    Path.cwd() / "puente_sies.tsv",
    Path.home() / "Downloads" / "puente_sies.tsv",
]
DEFAULT_PUENTE_SIES_COMPILADO_PATH = Path(__file__).with_name("control") / "catalogos" / "PUENTE_SIES_COMPILADO.tsv"
DEFAULT_GOB_NAC_CANDIDATES = [
    Path(__file__).with_name("gobernanza_nac.tsv"),
    Path.cwd() / "gobernanza_nac.tsv",
    Path.home() / "Downloads" / "gobernanza_nac.tsv",
]
DEFAULT_GOB_PAIS_EST_SEC_CANDIDATES = [
    Path(__file__).with_name("gobernanza_pais_est_sec.tsv"),
    Path.cwd() / "gobernanza_pais_est_sec.tsv",
    Path.home() / "Downloads" / "gobernanza_pais_est_sec.tsv",
]
DEFAULT_GOB_SEDE_CANDIDATES = [
    Path(__file__).with_name("gobernanza_sede.tsv"),
    Path.cwd() / "gobernanza_sede.tsv",
    Path.home() / "Downloads" / "gobernanza_sede.tsv",
]
DEFAULT_GOB_FOR_ING_ACT_CANDIDATES = [
    Path(__file__).with_name("gobernanza_for_ing_act.tsv"),
    Path.cwd() / "gobernanza_for_ing_act.tsv",
    Path.home() / "Downloads" / "gobernanza_for_ing_act.tsv",
]
DEFAULT_GOB_HOJA1_ESTADO_DESC_CANDIDATES = [
    Path(__file__).with_name("gobernanza_catalogos") / "gob_promedios_hoja1_estado_academico_descripcion.tsv",
    Path.cwd() / "gobernanza_catalogos" / "gob_promedios_hoja1_estado_academico_descripcion.tsv",
]
DEFAULT_GOB_DA_ESTADO_SITUACION_CANDIDATES = [
    Path(__file__).with_name("gobernanza_catalogos") / "gob_datosalumnos_estadoacademico_situacion.tsv",
    Path.cwd() / "gobernanza_catalogos" / "gob_datosalumnos_estadoacademico_situacion.tsv",
]
DEFAULT_OFERTA_ACADEMICA_XLSX_CANDIDATES = [
    Path(__file__).with_name("oferta_academica_2026.xlsx"),
    Path.cwd() / "oferta_academica_2026.xlsx",
    Path.home() / "Downloads" / "1668382 Captura IES Incremental5320 Oferta Académica Titulados5319 Titulados Histórico.xlsx",
]

# Puedes pegar aquí el TSV completo del código legacy si quieres dejarlo embebido.
CATALOGO_MANUAL_TSV = ""
PUENTE_SIES_TSV = ""


@dataclass
class Issue:
    severity: str
    area: str
    message: str
    count: int = 0


def _first_existing_path(candidates: list[Path]) -> Path | None:
    for candidate in candidates:
        candidate = candidate.expanduser()
        if candidate.exists():
            return candidate
    return None


def _default_input_arg() -> str:
    existing = _first_existing_path(DEFAULT_INPUT_CANDIDATES)
    if existing is not None:
        return str(existing)
    return str(DEFAULT_INPUT_CANDIDATES[-1])


def _resolve_optional_path(path_value: str | None, candidates: list[Path]) -> str | None:
    if path_value:
        return str(Path(path_value).expanduser())
    existing = _first_existing_path(candidates)
    return str(existing) if existing else None


def _normalize_doc(num: object, dv: object) -> str:
    return "".join(ch for ch in str(num) if ch.isdigit()) + str(dv).strip().upper()


def _pick_sheet(book: dict[str, pd.DataFrame], required: Iterable[str]) -> pd.DataFrame:
    req = set(required)
    for df in book.values():
        if req.issubset(df.columns):
            return df.copy()
    raise ValueError(f"No se encontró hoja con columnas: {sorted(req)}")


def _series_or_default(df: pd.DataFrame, col: str, default: str = "") -> pd.Series:
    if col in df.columns:
        return df[col].astype(str)
    return pd.Series([default] * len(df), index=df.index, dtype=str)


def _normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def _nonempty_mask(series: pd.Series) -> pd.Series:
    return series.notna() & series.astype(str).str.strip().ne("")


def _to_int_year(value: object) -> float:
    if pd.isna(value):
        return float("nan")
    match = re.search(r"\b(19\d{2}|20\d{2}|21\d{2})\b", str(value))
    return float(match.group(1)) if match else float("nan")


def _infer_year_from_codcli(value: object) -> float:
    if pd.isna(value):
        return float("nan")
    text = re.sub(r"\D", "", str(value))
    if len(text) >= 4:
        year = int(text[:4])
        if 1900 <= year <= 2100:
            return float(year)
    return float("nan")


def _pick_first_column(df: pd.DataFrame, options: list[str]) -> str | None:
    upper_map = {c.upper(): c for c in df.columns}
    for opt in options:
        if opt.upper() in upper_map:
            return upper_map[opt.upper()]
    return None


def _require_column(col: str | None, label: str) -> str:
    if col is None:
        raise ValueError(f"No se encontró columna obligatoria: {label}")
    return col


def _load_for_ing_act_catalog() -> tuple[set[int], str]:
    path = _first_existing_path(DEFAULT_GOB_FOR_ING_ACT_CANDIDATES)
    if path is not None:
        df = _load_governance_tsv(str(path), ["FOR_ING_ACT", "DESCRIPCION_MANUAL"])
        if not df.empty:
            codes = set(pd.to_numeric(df["FOR_ING_ACT"], errors="coerce").dropna().astype(int).tolist())
            if codes:
                return codes, str(path)
    return set(range(1, 12)), "fallback:hardcoded_1_11"


def _normalize_period_to_semester(values: pd.Series) -> pd.Series:
    periodo = pd.to_numeric(values, errors="coerce")
    sem = pd.Series(pd.NA, index=values.index, dtype="Int64")
    sem.loc[periodo == 1] = 1
    # El histórico Hoja1 usa PERIODO=3 como equivalente operacional de segundo semestre.
    sem.loc[periodo.isin([2, 3])] = 2
    return sem


def _normalize_grade_to_mu_scale(values: pd.Series) -> pd.Series:
    # Escala fuente → escala MU (100-700), documentada en gobernanza_escala_notas.tsv
    # Fuente 1.0-7.0  → multiplicar por 100 (ej. 5.8 → 580)
    # Fuente 10-70    → multiplicar por 10  (ej. 58 → 580)
    # Fuente 100-700  → ya en escala, usar directamente
    # Fuente 0 o fuera de rango → NA (sin calificacion; no contamina promedios)
    nota = pd.to_numeric(values, errors="coerce")
    out = pd.Series(pd.NA, index=values.index, dtype="Float64")
    out.loc[(nota >= 1) & (nota <= 7)] = (nota.loc[(nota >= 1) & (nota <= 7)] * 100).round()
    out.loc[(nota > 7) & (nota <= 70)] = (nota.loc[(nota > 7) & (nota <= 70)] * 10).round()
    out.loc[(nota >= 100) & (nota <= 700)] = nota.loc[(nota >= 100) & (nota <= 700)].round()
    return out.astype("Int64")


def _coerce_mu_average(series: pd.Series) -> int:
    nota = pd.to_numeric(series, errors="coerce").dropna()
    if nota.empty:
        return 0
    avg = int(round(float(nota.mean())))
    if avg == 0:
        return 0
    return min(max(avg, 100), 700)


def _build_mu_historico_summary(
    src: pd.DataFrame,
    rut_col: str,
    dv_col: str,
    codcarr_col: str,
    anio_ref_override: int | None = None,
) -> tuple[pd.DataFrame, int | None]:
    required = {"ANO", "PERIODO", "CODRAMO", rut_col, dv_col, codcarr_col}
    if not required.issubset(src.columns):
        return pd.DataFrame(), None

    hist_cols = ["ANO", "PERIODO", "CODRAMO", rut_col, dv_col, codcarr_col]
    extra_cols = [c for c in ["DESCRIPCION_ESTADO", "ESTADO", "CONVALIDADO", "NOTA_FINAL"] if c in src.columns]
    hist = src[hist_cols + extra_cols].copy().rename(
        columns={
            rut_col: "RUT",
            dv_col: "DIG",
            codcarr_col: "CODCARR",
        }
    )
    hist["RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(hist["RUT"], hist["DIG"])]
    hist["CODCARPR_NORM"] = hist["CODCARR"].map(_normalize_text)
    hist["ANO_NUM"] = pd.to_numeric(hist["ANO"], errors="coerce")
    hist["SEMESTRE_HIST"] = _normalize_period_to_semester(hist["PERIODO"])
    hist["NOTA_MU"] = _normalize_grade_to_mu_scale(hist["NOTA_FINAL"]) if "NOTA_FINAL" in hist.columns else pd.Series(pd.NA, index=hist.index, dtype="Int64")
    hist["ESTADO_HIST_NORM"] = _series_or_default(hist, "DESCRIPCION_ESTADO").map(_normalize_text)
    hist["CONVALIDADO_NORM"] = _series_or_default(hist, "CONVALIDADO").map(_normalize_text)

    anio_vals = hist["ANO_NUM"].dropna()
    if anio_vals.empty:
        return pd.DataFrame(), None
    anio_ref_data = int(anio_vals.max())
    # Si se proporciona override (anio_anterior_prom = ANIO_ING_ACT - 1), usarlo para PROM;
    # de lo contrario, caer al max del histórico (comportamiento legacy).
    anio_ref = anio_ref_override if anio_ref_override is not None else anio_ref_data
    print(f"  [HIST] Año referencia PROM: {anio_ref} (max datos: {anio_ref_data}, override: {anio_ref_override})")

    rows: list[dict[str, object]] = []
    for (rut_norm, codcarpr_norm), sub in hist.groupby(["RUT_NORM", "CODCARPR_NORM"], dropna=False):
        sub_ref = sub[sub["ANO_NUM"] == anio_ref].copy()
        estado_ref = sub_ref["ESTADO_HIST_NORM"]
        estado_hist = sub["ESTADO_HIST_NORM"]
        transfer_ref = estado_ref.str.contains(r"CONVALID|HOMOLOG|RECONOC|EQUIV", regex=True, na=False) | sub_ref["CONVALIDADO_NORM"].eq("S")
        graded_ref = sub_ref["NOTA_MU"].notna() & ~transfer_ref
        sem_ref = sub_ref["SEMESTRE_HIST"]
        aprob_ref = estado_ref.str.contains("APROB", na=False) & ~transfer_ref
        aprob_hist = estado_hist.str.contains(r"APROB|CONVALID|RECONOC|EQUIV|HOMOLOG", regex=True, na=False)
        anios_grupo = sorted({int(v) for v in sub["ANO_NUM"].dropna().astype(int).tolist()})
        anio_min = anios_grupo[0] if anios_grupo else anio_ref
        anio_max = anios_grupo[-1] if anios_grupo else anio_ref
        anios_disponibles = len(anios_grupo)
        hist_scope_status = "ALCANCE_MULTIANUAL" if anios_disponibles > 1 else "ALCANCE_LIMITADO_ANIO_UNICO"

        codramo_ref = sub_ref["CODRAMO"] if "CODRAMO" in sub_ref.columns else pd.Series(index=sub_ref.index, dtype=object)
        codramo_hist = sub["CODRAMO"] if "CODRAMO" in sub.columns else pd.Series(index=sub.index, dtype=object)

        rows.append(
            {
                "RUT_NORM": rut_norm,
                "CODCARPR_NORM": codcarpr_norm,
                "UZ_HIST_KEY": f"{rut_norm}|{codcarpr_norm}",
                "ANIO_REFERENCIA_HIST_UZ": anio_ref,
                "UZ_HIST_ANIO_MIN": anio_min,
                "UZ_HIST_ANIO_MAX": anio_max,
                "UZ_HIST_ANIOS_DISPONIBLES": anios_disponibles,
                "UZ_HIST_SCOPE_STATUS": hist_scope_status,
                "UZ_HIST_FILAS_TOTAL": int(len(sub)),
                "UZ_HIST_FILAS_ANIO_REFERENCIA": int(len(sub_ref)),
                "UZ_HIST_FILAS_REF_APROB": int(aprob_ref.sum()),
                "UZ_HIST_FILAS_REF_REPROB": int(estado_ref.str.contains("REPROB", na=False).sum()),
                "UZ_HIST_FILAS_REF_TRANSFER": int(transfer_ref.sum()),
                "UZ_HIST_FILAS_REF_SEM1_CALIFICADAS": int((graded_ref & sem_ref.eq(1)).sum()),
                "UZ_HIST_FILAS_REF_SEM2_CALIFICADAS": int((graded_ref & sem_ref.eq(2)).sum()),
                "ASI_INS_ANT_HIST": int(codramo_ref[~transfer_ref].nunique()),
                "ASI_APR_ANT_HIST": int(codramo_ref[aprob_ref].nunique()),
                "PROM_PRI_SEM_HIST": _coerce_mu_average(sub_ref.loc[graded_ref & sem_ref.eq(1), "NOTA_MU"]),
                "PROM_SEG_SEM_HIST": _coerce_mu_average(sub_ref.loc[graded_ref & sem_ref.eq(2), "NOTA_MU"]),
                "ASI_INS_HIS_HIST": int(codramo_hist.dropna().count()),
                "ASI_APR_HIS_HIST": int(codramo_hist[aprob_hist].nunique()),
                "UZ_FUENTE_HIST": f"HISTORICO_HOJA1_ANIO_{anio_ref}",
            }
        )

    if not rows:
        return pd.DataFrame(), anio_ref_data
    return pd.DataFrame(rows), anio_ref_data


def _resolve_for_ing_act_row(
    raw_input_value: object,
    vias_admision_value: object,
    carrera_value: object,
    codcarpr_value: object,
    valid_codes: set[int],
) -> dict[str, object]:
    def _numeric_code(value: object) -> int | None:
        num = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        if pd.isna(num):
            return None
        num_int = int(num)
        return num_int if num_int in valid_codes else None

    input_code = _numeric_code(raw_input_value)
    if input_code is not None:
        return {
            "FOR_ING_ACT": input_code,
            "FOR_ING_ACT_FUENTE_VALOR": raw_input_value,
            "FOR_ING_ACT_FUENTE_CAMPO": "INPUT_FOR_ING_ACT_EQUIVALENTE",
            "FOR_ING_ACT_FUENTE_NORM": _normalize_text(raw_input_value),
            "FOR_ING_ACT_METODO": "NUMERICO_EXACTO_INPUT",
            "FOR_ING_ACT_IMPUTADO": "NO",
            "FOR_ING_ACT_REQUIERE_REVISION": "NO",
        }

    source_field = "SIN_FUENTE"
    source_value = pd.NA
    if not pd.isna(vias_admision_value) and str(vias_admision_value).strip():
        source_field = "DA_VIASDEADMISION"
        source_value = vias_admision_value
    elif not pd.isna(raw_input_value) and str(raw_input_value).strip():
        source_field = "INPUT_FOR_ING_ACT_EQUIVALENTE"
        source_value = raw_input_value

    source_norm = _normalize_text(source_value)
    carrera_norm = _normalize_text(carrera_value)
    codcarpr_norm = _normalize_text(codcarpr_value)

    exact_map = {
        "ENSENANZA MEDIA NACIONAL": 1,
        "EXTRANJERO": 6,
    }
    if source_norm in exact_map and exact_map[source_norm] in valid_codes:
        return {
            "FOR_ING_ACT": exact_map[source_norm],
            "FOR_ING_ACT_FUENTE_VALOR": source_value,
            "FOR_ING_ACT_FUENTE_CAMPO": source_field,
            "FOR_ING_ACT_FUENTE_NORM": source_norm,
            "FOR_ING_ACT_METODO": "CATALOGO_EXACTO_VIAS_ADMISION",
            "FOR_ING_ACT_IMPUTADO": "NO",
            "FOR_ING_ACT_REQUIERE_REVISION": "NO",
        }

    token_rules = [
        ("CAMBIO EXTERNO", 4, "TOKEN_CAMBIO_EXTERNO"),
        ("CAMBIO INTERNO", 3, "TOKEN_CAMBIO_INTERNO"),
        ("RECONOCIMIENTO DE APRENDIZAJES PREVIOS", 5, "TOKEN_RAP"),
        ("RAP", 5, "TOKEN_RAP"),
        ("PACE", 7, "TOKEN_PACE"),
        ("INCLUSION", 8, "TOKEN_INCLUSION"),
        ("PLAN COMUN", 2, "TOKEN_PLAN_COMUN"),
        ("BACHILLER", 2, "TOKEN_BACHILLERATO"),
        ("ARTICUL", 11, "TOKEN_ARTICULACION"),
        ("EXTRANJER", 6, "TOKEN_EXTRANJERO"),
        ("ENSENANZA MEDIA", 1, "TOKEN_INGRESO_DIRECTO"),
    ]
    for token, code, method in token_rules:
        if token in source_norm and code in valid_codes:
            return {
                "FOR_ING_ACT": code,
                "FOR_ING_ACT_FUENTE_VALOR": source_value,
                "FOR_ING_ACT_FUENTE_CAMPO": source_field,
                "FOR_ING_ACT_FUENTE_NORM": source_norm,
                "FOR_ING_ACT_METODO": method,
                "FOR_ING_ACT_IMPUTADO": "NO",
                "FOR_ING_ACT_REQUIERE_REVISION": "NO",
            }

    if source_norm == "PROGRAMA DE EDUCACION CONTINUA" and 11 in valid_codes:
        if "CONTINUIDAD" in carrera_norm or "ARTICUL" in carrera_norm or codcarpr_norm.startswith("CI"):
            return {
                "FOR_ING_ACT": 11,
                "FOR_ING_ACT_FUENTE_VALOR": source_value,
                "FOR_ING_ACT_FUENTE_CAMPO": source_field,
                "FOR_ING_ACT_FUENTE_NORM": source_norm,
                "FOR_ING_ACT_METODO": "REGLA_CONTINUIDAD_ARTICULACION_CARRERA",
                "FOR_ING_ACT_IMPUTADO": "NO",
                "FOR_ING_ACT_REQUIERE_REVISION": "NO",
            }
        if 10 in valid_codes:
            return {
                "FOR_ING_ACT": 10,
                "FOR_ING_ACT_FUENTE_VALOR": source_value,
                "FOR_ING_ACT_FUENTE_CAMPO": source_field,
                "FOR_ING_ACT_FUENTE_NORM": source_norm,
                "FOR_ING_ACT_METODO": "FALLBACK_CONTROLADO_PROGRAMA_EDUCACION_CONTINUA",
                "FOR_ING_ACT_IMPUTADO": "SI",
                "FOR_ING_ACT_REQUIERE_REVISION": "SI",
            }

    if source_norm == "MNP AA" and 10 in valid_codes:
        return {
            "FOR_ING_ACT": 10,
            "FOR_ING_ACT_FUENTE_VALOR": source_value,
            "FOR_ING_ACT_FUENTE_CAMPO": source_field,
            "FOR_ING_ACT_FUENTE_NORM": source_norm,
            "FOR_ING_ACT_METODO": "FALLBACK_CONTROLADO_MNP_AA",
            "FOR_ING_ACT_IMPUTADO": "SI",
            "FOR_ING_ACT_REQUIERE_REVISION": "SI",
        }

    if source_field == "SIN_FUENTE":
        return {
            "FOR_ING_ACT": pd.NA,
            "FOR_ING_ACT_FUENTE_VALOR": pd.NA,
            "FOR_ING_ACT_FUENTE_CAMPO": source_field,
            "FOR_ING_ACT_FUENTE_NORM": "",
            "FOR_ING_ACT_METODO": "SIN_FUENTE_FOR_ING_ACT",
            "FOR_ING_ACT_IMPUTADO": "NO",
            "FOR_ING_ACT_REQUIERE_REVISION": "SI",
        }

    if 10 in valid_codes:
        return {
            "FOR_ING_ACT": 10,
            "FOR_ING_ACT_FUENTE_VALOR": source_value,
            "FOR_ING_ACT_FUENTE_CAMPO": source_field,
            "FOR_ING_ACT_FUENTE_NORM": source_norm,
            "FOR_ING_ACT_METODO": "FALLBACK_CONTROLADO_OTRAS_FORMAS",
            "FOR_ING_ACT_IMPUTADO": "SI",
            "FOR_ING_ACT_REQUIERE_REVISION": "SI",
        }

    return {
        "FOR_ING_ACT": pd.NA,
        "FOR_ING_ACT_FUENTE_VALOR": source_value,
        "FOR_ING_ACT_FUENTE_CAMPO": source_field,
        "FOR_ING_ACT_FUENTE_NORM": source_norm,
        "FOR_ING_ACT_METODO": "SIN_CATALOGO_VALIDO_FOR_ING_ACT",
        "FOR_ING_ACT_IMPUTADO": "NO",
        "FOR_ING_ACT_REQUIERE_REVISION": "SI",
    }


def _map_jornada_to_mod_jor(series: pd.Series) -> tuple[pd.Series, pd.Series]:
    s_norm = series.fillna("").map(_normalize_text)
    s_low = s_norm.str.lower()
    modalidad = pd.Series(pd.NA, index=series.index, dtype="object")
    jor = pd.Series(pd.NA, index=series.index, dtype="object")

    diurna = s_norm.isin({"D", "1", "DIURNA"}) | s_low.str.contains("diurn", na=False)
    vespertina = s_norm.isin({"V", "2", "VESPERTINA"}) | s_low.str.contains("vespert", na=False)
    semi = s_norm.isin({"3", "SEMIPRESENCIAL"}) | s_low.str.contains("semi", na=False)
    distancia = (
        s_norm.isin({"O", "4", "A DISTANCIA", "DISTANCIA", "ONLINE"})
        | s_low.str.contains("dist", na=False)
        | s_low.str.contains("online", na=False)
    )

    modalidad.loc[diurna | vespertina] = "1"  # Presencial
    modalidad.loc[semi] = "2"  # Semipresencial
    modalidad.loc[distancia] = "3"  # No presencial

    jor.loc[diurna] = "1"
    jor.loc[vespertina] = "2"
    jor.loc[semi] = "3"
    jor.loc[distancia] = "4"

    return modalidad, jor


_SIES_CODE_RE = re.compile(r"^I\d+S(?P<cod_sed>\d+)C(?P<cod_car>\d+)J(?P<jor>\d+)V(?P<version>\d+)$", re.IGNORECASE)


def _extract_shared_cod_car_from_potenciales(potenciales: object) -> object:
    """Extrae COD_CAR si TODOS los códigos SIES potenciales comparten el mismo componente C."""
    if pd.isna(potenciales) or not str(potenciales).strip():
        return pd.NA
    cod_cars = set()
    for m in re.finditer(r"C(\d+)J", str(potenciales)):
        cod_cars.add(int(m.group(1)))
    if len(cod_cars) == 1:
        return cod_cars.pop()
    return pd.NA


def _build_nombre_carrera_to_cod_car(oferta_dim: pd.DataFrame) -> dict[str, int]:
    """Construye mapeo NOMBRE_CARRERA normalizado → CODIGO_CARRERA desde oferta/DURACION."""
    mapping: dict[str, int] = {}
    if oferta_dim.empty:
        return mapping
    for _, row in oferta_dim.iterrows():
        nombre = str(row.get("NOMBRE_CARRERA", "")).strip().upper()
        nombre = re.sub(r"\s+", " ", nombre)
        cod_car = row.get("CODIGO_CARRERA")
        if nombre and pd.notna(cod_car):
            cod_car_int = int(cod_car)
            if nombre not in mapping:
                mapping[nombre] = cod_car_int
    # Alias de normalización conocidos (singular/plural, variantes)
    _aliases: list[tuple[str, str]] = [
        ("TECNICO EN ANALISIS DE SISTEMA", "TECNICO EN ANALISIS DE SISTEMAS"),
        ("DIPLOMADO EN CIBERSEGURIDAD", "DIPLOMADO EN CIBERSEGURIDAD APLICADA"),
    ]
    for alias, canonical in _aliases:
        if alias not in mapping and canonical in mapping:
            mapping[alias] = mapping[canonical]
    return mapping


def _normalize_nombre_carrera_for_lookup(nombre: str) -> str:
    """Normaliza un nombre de carrera para buscar en el mapeo."""
    n = re.sub(r"\s+", " ", nombre.strip().upper())
    # Quitar prefijo "CONTINUIDAD " para matchear la carrera base
    if n.startswith("CONTINUIDAD "):
        n = n[len("CONTINUIDAD "):]
    return n


def _extract_sies_components(value: object) -> tuple[str, str, str, str] | None:
    if pd.isna(value):
        return None
    text = str(value).strip().upper()
    m = _SIES_CODE_RE.match(text)
    if not m:
        return None
    cod_sed = str(int(m.group("cod_sed")))
    cod_car = str(int(m.group("cod_car")))
    jor = str(int(m.group("jor")))
    version = str(int(m.group("version")))
    return cod_sed, cod_car, jor, version


def _modalidad_from_jor(jor_series: pd.Series) -> pd.Series:
    j = pd.to_numeric(jor_series, errors="coerce")
    mod = pd.Series(pd.NA, index=jor_series.index, dtype="object")
    mod.loc[j.isin([1, 2])] = "1"  # Presencial
    mod.loc[j == 3] = "2"  # Semipresencial
    mod.loc[j.isin([4, 5])] = "3"  # No presencial
    return mod


def _normalize_sexo_mu(value: object) -> object:
    if pd.isna(value):
        return pd.NA
    raw = str(value).strip().upper()
    # DatosAlumnos usa códigos institucionales M/F/S; el manual MU exige H/M/NB.
    if raw == "M":
        return "H"
    if raw == "F":
        return "M"
    if raw == "S":
        return "NB"
    s = _normalize_text(value)
    if s in {"H", "HOMBRE", "MASCULINO", "MALE"}:
        return "H"
    if s in {"M", "MUJER", "FEMENINO", "FEMALE"}:
        return "M"
    if s in {"NB", "NO BINARIO", "NO-BINARIO", "X"}:
        return "NB"
    return pd.NA


def _to_ddmmyyyy(series: pd.Series, fallback: str | None = None) -> pd.Series:
    dt = pd.to_datetime(series, errors="coerce", dayfirst=True)
    out = dt.dt.strftime("%d/%m/%Y")
    if fallback is not None:
        out = out.fillna(fallback)
    return out


def _load_datos_alumnos_lookup(input_file: Path) -> pd.DataFrame:
    """Carga lookup por CODCLI desde hoja DatosAlumnos para gobernanza v2.

    El lookup se usa solo cuando el flag --usar-gobernanza-v2 está activo.
    """
    try:
        xls = pd.ExcelFile(input_file)
        if "DatosAlumnos" not in xls.sheet_names:
            return pd.DataFrame()

        src = pd.read_excel(input_file, sheet_name="DatosAlumnos")
        if "CODCLI" not in src.columns:
            return pd.DataFrame()

        keep = [
            "CODCLI",
            "RUT",
            "DIG",
            "DV",
            "NACIONALIDAD",
            "SEDE",
            "NOMBRES",
            "APELLIDO PATERNO",
            "APELLIDO MATERNO",
            "SEXO",
            "FECHANACIMIENTO",
            "ANOINGRESO",
            "PERIODOINGRESO",
            "ANOMATRICULA",
            "PERIODOMATRICULA",
            "FECHAMATRICULA",
            "NIVEL",
            "SITUACION",
            "ESTADOACADEMICO",
            "MATRICULA",
            "CON_FIRMA",
            "COMUNACOLEGIO",
            "CIUDADCOLEGIO",
            "VIASDEADMISION",
        ]
        cols = [c for c in keep if c in src.columns]
        out = src[cols].copy()
        out["CODCLI"] = out["CODCLI"].astype(str).str.strip()
        out = out[out["CODCLI"] != ""]

        rename_map = {
            "RUT": "DA_RUT",
            "DIG": "DA_DIG",
            "DV": "DA_DV",
            "NACIONALIDAD": "DA_NACIONALIDAD",
            "SEDE": "DA_SEDE",
            "NOMBRES": "DA_NOMBRES",
            "APELLIDO PATERNO": "DA_APELLIDO_PATERNO",
            "APELLIDO MATERNO": "DA_APELLIDO_MATERNO",
            "SEXO": "DA_SEXO",
            "FECHANACIMIENTO": "DA_FECHANACIMIENTO",
            "ANOINGRESO": "DA_ANOINGRESO",
            "PERIODOINGRESO": "DA_PERIODOINGRESO",
            "ANOMATRICULA": "DA_ANOMATRICULA",
            "PERIODOMATRICULA": "DA_PERIODOMATRICULA",
            "FECHAMATRICULA": "DA_FECHAMATRICULA",
            "NIVEL": "DA_NIVEL",
            "SITUACION": "DA_SITUACION",
            "ESTADOACADEMICO": "DA_ESTADOACADEMICO",
            "MATRICULA": "DA_MATRICULA",
            "CON_FIRMA": "DA_CON_FIRMA",
            "COMUNACOLEGIO": "DA_COMUNACOLEGIO",
            "CIUDADCOLEGIO": "DA_CIUDADCOLEGIO",
            "VIASDEADMISION": "DA_VIASDEADMISION",
        }
        out = out.rename(columns=rename_map)

        if "DA_DV" not in out.columns and "DA_DIG" in out.columns:
            out["DA_DV"] = out["DA_DIG"]
        if "DA_DIG" not in out.columns and "DA_DV" in out.columns:
            out["DA_DIG"] = out["DA_DV"]

        # Si no existe DA_DV ni DA_DIG, extraer del RUT (formato "12345678-K")
        if "DA_DV" not in out.columns and "DA_RUT" in out.columns:
            rut_str = out["DA_RUT"].astype(str).str.strip()
            out["DA_DV"] = rut_str.str.split("-").str[-1].str.strip()
            out["DA_RUT_NUM"] = rut_str.str.split("-").str[0].str.strip()

        if "DA_RUT_NORM" not in out.columns:
            if "DA_RUT_NUM" in out.columns:
                out["DA_RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(out["DA_RUT_NUM"], out["DA_DV"])]
            elif "DA_DV" in out.columns:
                out["DA_RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(out["DA_RUT"], out["DA_DV"])]
            else:
                out["DA_RUT_NORM"] = ""

        out["DA_MATCH_FLAG"] = "1"

        # DatosAlumnos debería tener 1 fila por CODCLI; ante conflicto, conservar primera.
        out = out.drop_duplicates(subset=["CODCLI"], keep="first").reset_index(drop=True)
        return out
    except Exception:
        return pd.DataFrame()


def _status_from_vig(vig: object) -> str:
    try:
        return "Matrícula No Utilizada" if int(float(vig)) == 0 else "Matrícula OK"
    except Exception:
        return "Matrícula OK"


def _build_bridge_codcarpr_to_codcar(df_bridge: pd.DataFrame) -> dict[str, int]:
    """Construye mapa CODCARPR → CODIGO_CARRERA desde el puente SIES."""
    if df_bridge.empty or "CODIGO_CARRERA" not in df_bridge.columns:
        return {}
    mapping: dict[str, int] = {}
    work = df_bridge.copy()
    work["CODIGO_CARRERA_NUM"] = pd.to_numeric(work["CODIGO_CARRERA"], errors="coerce")
    for _, row in work.dropna(subset=["CODIGO_CARRERA_NUM"]).iterrows():
        codcarpr = str(row.get("CODCARPR", "")).strip()
        if codcarpr:
            mapping[codcarpr] = int(row["CODIGO_CARRERA_NUM"])
    return mapping


def _build_revision_manual(
    archivo_subida: pd.DataFrame,
    sin_match: pd.DataFrame,
    sin_match_datos_alumnos_df: pd.DataFrame,
) -> pd.DataFrame:
    """Construye hoja REVISION_MANUAL con resumen accionable de pendientes."""
    rows: list[dict] = []
    # --- Sección 1: PENDIENTE_GOBERNANZA (ambiguos sin resolver) ---
    pend = archivo_subida[archivo_subida.get("SIES_RESOLUCION_HEURISTICA", pd.Series()) == "PENDIENTE_GOBERNANZA"]
    if not pend.empty:
        for combo, grp in pend.groupby("CODIGOS_SIES_POTENCIALES", dropna=False):
            codes = [c.strip() for c in str(combo).split("|")] if pd.notna(combo) else []
            cod_cars = set()
            for c in codes:
                m = _SIES_CODE_RE.match(c)
                if m:
                    cod_cars.add(int(m.group("cod_car")))
            mismo_cod = len(cod_cars) <= 1
            rows.append({
                "SECCION": "PENDIENTE_GOBERNANZA",
                "PRIORIDAD": "BAJA (misma carrera, distinta versión)" if mismo_cod else "ALTA (carreras distintas)",
                "CODIGOS_SIES_POTENCIALES": combo,
                "CODIGO_CARRERA_OPCIONES": " | ".join(str(c) for c in sorted(cod_cars)) if cod_cars else "",
                "N_ALUMNOS_AFECTADOS": len(grp),
                "CODCARPR_EJEMPLO": grp.iloc[0].get("CODCARPR_NORM", grp.iloc[0].get("COD_CAR", "")),
                "NOMBRE_CARRERA_EJEMPLO": grp.iloc[0].get("NOMBRE_CARRERA_FUENTE", ""),
                "ACCION_REQUERIDA": "Definir versión correcta en gobernanza" if mismo_cod else "REVISAR: definir carrera correcta",
            })
    # --- Sección 2: SIN_MATCH_SIES ---
    if not sin_match.empty:
        codcarpr_col = "CODCARPR_NORM" if "CODCARPR_NORM" in sin_match.columns else "COD_CAR"
        for codcarpr, grp in sin_match.groupby(codcarpr_col, dropna=False):
            rows.append({
                "SECCION": "SIN_MATCH_SIES",
                "PRIORIDAD": "ALTA (sin código SIES)",
                "CODIGOS_SIES_POTENCIALES": "",
                "CODIGO_CARRERA_OPCIONES": "",
                "N_ALUMNOS_AFECTADOS": len(grp),
                "CODCARPR_EJEMPLO": codcarpr,
                "NOMBRE_CARRERA_EJEMPLO": grp.iloc[0].get("NOMBRE_CARRERA_FUENTE", ""),
                "ACCION_REQUERIDA": "Agregar mapeo SIES para este CODCARPR",
            })
    # --- Sección 3: SIN_MATCH_DATOS_ALUMNOS (resumen) ---
    if not sin_match_datos_alumnos_df.empty:
        for codcarpr, grp in sin_match_datos_alumnos_df.groupby("COD_CAR_FUENTE", dropna=False):
            rows.append({
                "SECCION": "SIN_MATCH_DATOS_ALUMNOS",
                "PRIORIDAD": "MEDIA (sin match en DatosAlumnos)",
                "CODIGOS_SIES_POTENCIALES": "",
                "CODIGO_CARRERA_OPCIONES": "",
                "N_ALUMNOS_AFECTADOS": len(grp),
                "CODCARPR_EJEMPLO": codcarpr,
                "NOMBRE_CARRERA_EJEMPLO": grp.iloc[0].get("NOMBRE_CARRERA_FUENTE", ""),
                "ACCION_REQUERIDA": "Verificar si alumnos existen en hoja DatosAlumnos",
            })
    if not rows:
        return pd.DataFrame(columns=[
            "SECCION", "PRIORIDAD", "CODIGOS_SIES_POTENCIALES",
            "CODIGO_CARRERA_OPCIONES", "N_ALUMNOS_AFECTADOS",
            "CODCARPR_EJEMPLO", "NOMBRE_CARRERA_EJEMPLO", "ACCION_REQUERIDA",
        ])
    result = pd.DataFrame(rows)
    result = result.sort_values(["SECCION", "N_ALUMNOS_AFECTADOS"], ascending=[True, False]).reset_index(drop=True)
    return result


def _write_excel_atomic(
    sheets: dict[str, pd.DataFrame],
    final_path: Path,
    red_rows_sheet: str | None = None,
    red_rows_mask: pd.Series | None = None,
) -> None:
    from openpyxl.styles import PatternFill, Font
    final_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="mu_export_") as tmpdir:
        tmp_path = Path(tmpdir) / final_path.name
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            # ── Formato rojo en CODCLI para filas que requieren revisión manual ──
            if red_rows_sheet and red_rows_mask is not None and red_rows_mask.any():
                ws = writer.sheets[red_rows_sheet[:31]]
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                red_font = Font(color="9C0006")
                df_target = sheets[red_rows_sheet]
                codcli_col_idx = list(df_target.columns).index("CODCLI") + 1 if "CODCLI" in df_target.columns else 1
                # Columnas clave para marcar: CODCLI, N_CODES_SIES, SIES_RESOLUCION_HEURISTICA
                mark_cols = []
                for cname in ["CODCLI", "N_CODES_SIES", "CODIGOS_SIES_POTENCIALES", "SIES_RESOLUCION_HEURISTICA"]:
                    if cname in df_target.columns:
                        mark_cols.append(list(df_target.columns).index(cname) + 1)
                if not mark_cols:
                    mark_cols = [codcli_col_idx]
                pos_map = {v: i for i, v in enumerate(df_target.index)}
                for idx in red_rows_mask[red_rows_mask].index:
                    if idx in pos_map:
                        excel_row = pos_map[idx] + 2
                        for col in mark_cols:
                            cell = ws.cell(row=excel_row, column=col)
                            cell.fill = red_fill
                            cell.font = red_font
        if final_path.exists():
            final_path.unlink()
        shutil.copy2(tmp_path, final_path)


def _write_mu_csv_atomic(df: pd.DataFrame, final_path: Path) -> None:
    final_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="mu_export_csv_") as tmpdir:
        tmp_path = Path(tmpdir) / final_path.name
        df.to_csv(
            tmp_path,
            sep=";",
            header=False,
            index=False,
            encoding="utf-8",
            lineterminator="\n",
        )
        if final_path.exists():
            final_path.unlink()
        shutil.copy2(tmp_path, final_path)


_AUDIT_CONSOL_COLUMNS = [
    "CODCLI", "TIPO_DOC", "N_DOC", "DV", "COD_CAR",
    "NOMBRE_CARRERA", "VIG", "ANIO_ING_ACT",
    "CASO", "CLASIFICACION", "ACCION",
]


def _consolidar_candidatos_por_codcli(
    candidatos: pd.DataFrame,
    estado_carga: pd.Series,
) -> tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    """Consolidación de candidatos usando CODCLI como clave interna primaria.

    Paso 1 – Intra-CODCLI (Caso A): si un mismo CODCLI tiene >1 fila,
             se queda la de FECHA_MATRICULA más reciente.
    Paso 2 – Clave compuesta 8-col (red de seguridad legacy):
             dedup por [TIPO_DOC, N_DOC, DV, COD_SED, COD_CAR, MODALIDAD, JOR, VERSION].
    Paso 3 – Multi-identidad (Caso C): clasifica sin eliminar.
             Multi-carrera legítima se MANTIENE per manual.

    Retorna (candidatos_consolidados, estado_carga, auditoría).
    """
    audit_rows: list[dict] = []
    identity_keys = ["TIPO_DOC", "N_DOC", "DV"]

    # ── Paso 1: Intra-CODCLI ──────────────────────────────────────────
    # Tie-breaker: cuando _FECHA_MAT_TMP empata (ej. sentinel 01/01/1900),
    # preferir fila con NIV_ACA más alto (nivel más avanzado).
    _niv_sort = pd.to_numeric(candidatos.get("NIV_ACA"), errors="coerce").fillna(0)
    candidatos["_NIV_SORT"] = _niv_sort
    candidatos = candidatos.sort_values(
        ["CODCLI", "_FECHA_MAT_TMP", "_NIV_SORT"], ascending=[True, False, False],
    )
    intra_dup = candidatos.duplicated(subset=["CODCLI"], keep="first")
    for idx in candidatos.index[intra_dup]:
        r = candidatos.loc[idx]
        audit_rows.append({
            "CODCLI": r["CODCLI"], "TIPO_DOC": r["TIPO_DOC"],
            "N_DOC": r["N_DOC"], "DV": r["DV"],
            "COD_CAR": r["COD_CAR"],
            "NOMBRE_CARRERA": r.get("NOMBRE_CARRERA_FUENTE", ""),
            "VIG": r["VIG"], "ANIO_ING_ACT": r.get("ANIO_ING_ACT", ""),
            "CASO": "A_INTRA_CODCLI",
            "CLASIFICACION": "DUPLICADO_TECNICO",
            "ACCION": "EXCLUIDO",
        })
    estado_carga.loc[candidatos.index[intra_dup]] = "EXCLUIDO_DUPLICADO_INTRA_CODCLI"
    candidatos = candidatos.loc[~intra_dup].copy()
    candidatos.drop(columns=["_NIV_SORT"], inplace=True, errors="ignore")

    # ── Paso 2: Dedup clave compuesta 8-col (legacy safety-net) ───────
    dedupe_keys = [
        "TIPO_DOC", "N_DOC", "DV", "COD_SED",
        "COD_CAR", "MODALIDAD", "JOR", "VERSION",
    ]
    candidatos = candidatos.sort_values(
        ["TIPO_DOC", "N_DOC", "DV", "COD_CAR", "_FECHA_MAT_TMP"],
        ascending=[True, True, True, True, False],
    )
    dup_8col = candidatos.duplicated(subset=dedupe_keys, keep="first")
    for idx in candidatos.index[dup_8col]:
        r = candidatos.loc[idx]
        audit_rows.append({
            "CODCLI": r["CODCLI"], "TIPO_DOC": r["TIPO_DOC"],
            "N_DOC": r["N_DOC"], "DV": r["DV"],
            "COD_CAR": r["COD_CAR"],
            "NOMBRE_CARRERA": r.get("NOMBRE_CARRERA_FUENTE", ""),
            "VIG": r["VIG"], "ANIO_ING_ACT": r.get("ANIO_ING_ACT", ""),
            "CASO": "B_DUPLICADO_CLAVE_CARGA",
            "CLASIFICACION": "DUPLICADO_CLAVE_8_COL",
            "ACCION": "EXCLUIDO",
        })
    estado_carga.loc[candidatos.index[dup_8col]] = "EXCLUIDO_DUPLICADO_CLAVE_CARGA"
    candidatos = candidatos.loc[~dup_8col].copy()

    # ── Paso 3: Clasificar multi-identidad (Caso C) ──────────────────
    id_counts = candidatos.groupby(identity_keys, sort=False).size()
    multi_ids = id_counts[id_counts > 1]

    for keys in multi_ids.index:
        tipo, ndoc, dv = keys
        mask = (
            (candidatos["TIPO_DOC"] == tipo)
            & (candidatos["N_DOC"] == ndoc)
            & (candidatos["DV"] == dv)
        )
        group = candidatos.loc[mask]
        vigs = set(group["VIG"].astype(int).unique())

        if vigs == {1}:
            clasificacion = "MULTI_CARRERA_AMBAS_VIGENTES"
        elif vigs == {2}:
            clasificacion = "MULTI_CARRERA_AMBAS_EGRESADAS"
        elif vigs.issuperset({1, 2}):
            clasificacion = "MULTI_CARRERA_MIXTA_VIG1_VIG2"
        else:
            clasificacion = "MULTI_CARRERA_OTRO"

        for _, r in group.iterrows():
            audit_rows.append({
                "CODCLI": r["CODCLI"], "TIPO_DOC": r["TIPO_DOC"],
                "N_DOC": r["N_DOC"], "DV": r["DV"],
                "COD_CAR": r["COD_CAR"],
                "NOMBRE_CARRERA": r.get("NOMBRE_CARRERA_FUENTE", ""),
                "VIG": r["VIG"], "ANIO_ING_ACT": r.get("ANIO_ING_ACT", ""),
                "CASO": "C_MULTI_CODCLI_MISMA_IDENTIDAD",
                "CLASIFICACION": clasificacion,
                "ACCION": "MANTENER_MULTI_CARRERA",
            })

    auditoria = (
        pd.DataFrame(audit_rows, columns=_AUDIT_CONSOL_COLUMNS)
        if audit_rows
        else pd.DataFrame(columns=_AUDIT_CONSOL_COLUMNS)
    )
    return candidatos, estado_carga, auditoria


def _read_tsv_text(tsv_text: str) -> pd.DataFrame:
    lines = [line.strip() for line in tsv_text.splitlines() if line.strip()]
    if not lines:
        return pd.DataFrame()
    header = lines[0].split("\t")
    rows = [line.split("\t") for line in lines[1:]]
    return pd.DataFrame(rows, columns=header)


def _load_tsv_table(path: str | None, embedded: str) -> pd.DataFrame:
    if path:
        p = Path(path).expanduser().resolve()
        text = p.read_text(encoding="utf-8")
        return _read_tsv_text(text)
    return _read_tsv_text(embedded)


def _load_governance_tsv(path: str | None, required_columns: list[str]) -> pd.DataFrame:
    """Carga TSV de gobernanza desde archivo externo.

    Si no existe archivo o faltan columnas obligatorias, retorna DataFrame vacío.
    """
    if not path:
        return pd.DataFrame()

    p = Path(path).expanduser().resolve()
    if not p.exists():
        return pd.DataFrame()

    try:
        df = pd.read_csv(p, sep="\t", dtype=str, keep_default_na=False)
    except Exception:
        return pd.DataFrame()

    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        return pd.DataFrame()

    return df.copy()


def _normalize_code_or_na(value: object) -> object:
    if pd.isna(value):
        return pd.NA
    text = str(value).strip()
    return text if text else pd.NA


def _extract_alpha_prefix(value: object) -> str:
    text = _normalize_text(value)
    m = re.match(r"^([A-Z]+)", text)
    return m.group(1) if m else ""


def _build_key_3(jornada: object, codcarpr: object, nombre: object) -> str:
    return f"{_normalize_text(jornada)}|{_normalize_text(codcarpr)}|{_normalize_text(nombre)}"


def _build_key_no_jornada(codcarpr: object, nombre: object) -> str:
    return f"|{_normalize_text(codcarpr)}|{_normalize_text(nombre)}"


def _is_diplomado_name(name: object) -> bool:
    text = _normalize_text(name)
    return text.startswith("DIPLOMADO") or text.startswith("DIPLOMADOS") or "EC CURSOS Y DIPLOMADOS" in text


def _prepare_catalog_manual(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    required = {"GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Catálogo manual inválido: faltan columnas {sorted(missing)}")

    out = df.copy()
    out["GRUPO_TRAZA"] = out["GRUPO_TRAZA"].astype(str).str.strip()
    out["JORNADA"] = out["JORNADA"].map(_normalize_text)
    out["CODCARPR"] = out["CODCARPR"].map(_normalize_text)
    out["NOMBRE_L"] = out["NOMBRE_L"].map(_normalize_text)
    out["FAMILIA_TRAZA"] = out["GRUPO_TRAZA"].map(_extract_alpha_prefix)
    out["FAMILIA_CODCARPR"] = out["CODCARPR"].map(_extract_alpha_prefix)
    out["MANUAL_KEY_3"] = out.apply(lambda r: _build_key_3(r["JORNADA"], r["CODCARPR"], r["NOMBRE_L"]), axis=1)
    out["MANUAL_KEY_NO_JORNADA"] = out.apply(lambda r: _build_key_no_jornada(r["CODCARPR"], r["NOMBRE_L"]), axis=1)
    return out.drop_duplicates().reset_index(drop=True)


def _load_oferta_academica_dim(input_file: Path, explicit_path: str | None = None) -> pd.DataFrame:
    """Carga dimensión de oferta académica para validaciones de carga.

    Busca un archivo XLSX con columnas:
    CODIGO_UNICO, MODALIDAD, JORNADA, DURACION_ESTUDIOS, TIPO_PLAN_CARRERA, NIVEL_CARRERA.
    Si no encuentra, retorna DataFrame vacío (no bloqueante).
    """
    required_cols = {"CODIGO_UNICO", "MODALIDAD", "JORNADA", "DURACION_ESTUDIOS"}
    candidates: list[Path] = []

    if explicit_path:
        candidates.append(Path(explicit_path).expanduser().resolve())
    candidates.append(input_file)
    candidates.extend(DEFAULT_OFERTA_ACADEMICA_XLSX_CANDIDATES)
    downloads = Path.home() / "Downloads"
    if downloads.exists():
        candidates.extend(sorted(downloads.glob("*Oferta*Acad*mica*.xlsx")))

    seen: set[Path] = set()
    for p in candidates:
        try:
            rp = p.resolve()
        except Exception:
            continue
        if rp in seen:
            continue
        seen.add(rp)
        if not rp.exists() or rp.suffix.lower() != ".xlsx":
            continue
        try:
            xls = pd.ExcelFile(rp)
        except Exception:
            continue

        target_sheet: str | None = None
        for sheet in xls.sheet_names:
            try:
                cols = set(pd.read_excel(rp, sheet_name=sheet, nrows=0).columns)
            except Exception:
                continue
            if required_cols.issubset(cols):
                target_sheet = sheet
                break

        if target_sheet is None:
            continue

        try:
            # Leer columnas base + columnas extendidas si existen
            base_usecols = ["CODIGO_UNICO", "MODALIDAD", "JORNADA", "DURACION_ESTUDIOS", "VIGENCIA"]
            extended_cols = ["TIPO_PLAN_CARRERA", "NIVEL_CARRERA", "CODIGO_CARRERA"]
            available_cols = set(pd.read_excel(rp, sheet_name=target_sheet, nrows=0).columns)
            usecols = base_usecols + [c for c in extended_cols if c in available_cols]
            dim = pd.read_excel(
                rp,
                sheet_name=target_sheet,
                usecols=usecols,
            )
            dim = dim.dropna(subset=["CODIGO_UNICO"]).copy()
            dim["CODIGO_UNICO"] = dim["CODIGO_UNICO"].astype(str).str.strip().str.upper()
            dim = dim.drop_duplicates(subset=["CODIGO_UNICO"], keep="first").reset_index(drop=True)
            dim["OFERTA_SOURCE_PATH"] = str(rp)
            dim["OFERTA_SOURCE_SHEET"] = target_sheet
            return dim
        except Exception:
            continue

    # ── Fallback: DURACION_ESTUDIOS.tsv como dimensión oferta ────────────────
    tsv_candidates = [
        Path(__file__).with_name("DURACION_ESTUDIOS.tsv"),
        Path.cwd() / "DURACION_ESTUDIOS.tsv",
    ]
    for tsv_path in tsv_candidates:
        if tsv_path.exists():
            try:
                dim = pd.read_csv(tsv_path, sep="\t", dtype=str)
                if "CODIGO_UNICO" in dim.columns:
                    dim["CODIGO_UNICO"] = dim["CODIGO_UNICO"].astype(str).str.strip().str.upper()
                    for nc in ["MODALIDAD", "JORNADA", "DURACION_ESTUDIOS", "VIGENCIA",
                               "TIPO_PLAN_CARRERA", "NIVEL_CARRERA", "CODIGO_CARRERA"]:
                        if nc in dim.columns:
                            dim[nc] = pd.to_numeric(dim[nc], errors="coerce").astype("Int64")
                    dim = dim.drop_duplicates(subset=["CODIGO_UNICO"], keep="first").reset_index(drop=True)
                    dim["OFERTA_SOURCE_PATH"] = str(tsv_path)
                    dim["OFERTA_SOURCE_SHEET"] = "DURACION_ESTUDIOS_TSV"
                    return dim
            except Exception:
                continue

    return pd.DataFrame(columns=["CODIGO_UNICO", "MODALIDAD", "JORNADA", "DURACION_ESTUDIOS", "VIGENCIA"])


def _prepare_puente_sies(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    required = {"GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L", "CODIGO_CARRERA_SIES"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Puente SIES inválido: faltan columnas {sorted(missing)}")

    base = df.copy()
    base["GRUPO_TRAZA"] = base["GRUPO_TRAZA"].astype(str).str.strip()
    base["JORNADA"] = base["JORNADA"].map(_normalize_text)
    base["CODCARPR"] = base["CODCARPR"].map(_normalize_text)
    base["NOMBRE_L"] = base["NOMBRE_L"].map(_normalize_text)
    base["CODIGO_CARRERA_SIES"] = base["CODIGO_CARRERA_SIES"].astype(str).str.strip()
    base["FAMILIA_TRAZA"] = base["GRUPO_TRAZA"].map(_extract_alpha_prefix)
    base["FAMILIA_CODCARPR"] = base["CODCARPR"].map(_extract_alpha_prefix)
    base["BRIDGE_KEY_3"] = base.apply(lambda r: _build_key_3(r["JORNADA"], r["CODCARPR"], r["NOMBRE_L"]), axis=1)
    base["BRIDGE_KEY_NO_JORNADA"] = base.apply(lambda r: _build_key_no_jornada(r["CODCARPR"], r["NOMBRE_L"]), axis=1)

    rows = []
    for key, sub in base.groupby("BRIDGE_KEY_3", dropna=False):
        codes = (
            sub["CODIGO_CARRERA_SIES"]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .drop_duplicates()
            .tolist()
        )
        row: dict[str, object] = {
            "BRIDGE_KEY_3": key,
            "BRIDGE_KEY_NO_JORNADA": sub["BRIDGE_KEY_NO_JORNADA"].iloc[0],
            "GRUPO_TRAZA": " | ".join(sub["GRUPO_TRAZA"].dropna().astype(str).drop_duplicates().tolist()),
            "FAMILIA_TRAZA": sub["FAMILIA_TRAZA"].iloc[0],
            "FAMILIA_CODCARPR": sub["FAMILIA_CODCARPR"].iloc[0],
            "JORNADA": sub["JORNADA"].iloc[0],
            "CODCARPR": sub["CODCARPR"].iloc[0],
            "NOMBRE_L": sub["NOMBRE_L"].iloc[0],
            "N_CODES_SIES": len(codes),
            "CODIGOS_SIES_POTENCIALES": " | ".join(codes),
        }
        for idx in range(MAX_SIES_CODES_PER_KEY):
            row[f"CODIGO_CARRERA_SIES_{idx + 1}"] = codes[idx] if idx < len(codes) else pd.NA
        rows.append(row)

    result = pd.DataFrame(rows).sort_values(["GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L"]).reset_index(drop=True)
    result["CODIGO_CARRERA"] = result["CODIGO_CARRERA_SIES_1"].apply(
        lambda x: int(m.group("cod_car")) if pd.notna(x) and (m := _SIES_CODE_RE.match(str(x).strip())) else pd.NA
    )
    return result


def _build_catalog_and_bridge_from_duracion() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Genera catálogo manual y puente SIES desde DURACION_ESTUDIOS.tsv."""
    dur = _load_duracion_as_governance_df()
    if dur.empty:
        return pd.DataFrame(), pd.DataFrame()

    required = {"CODIGO_UNICO", "NOMBRE_CARRERA", "JORNADA"}
    if not required.issubset(set(dur.columns)):
        print("⚠️  DURACION_ESTUDIOS sin columnas mínimas para construir puente/catálogo")
        return pd.DataFrame(), pd.DataFrame()

    rows: list[dict[str, str]] = []
    for row in dur.itertuples(index=False):
        codigo_unico = str(getattr(row, "CODIGO_UNICO", "")).strip().upper()
        nombre = str(getattr(row, "NOMBRE_CARRERA", "")).strip()
        jornada_num = str(getattr(row, "JORNADA", "")).strip()
        jornada = {"1": "D", "2": "V", "3": "O", "4": "O"}.get(jornada_num, jornada_num.upper())
        codcarprs = _split_codcarpr_candidates(
            getattr(row, "CODCARPR_CANONICO", ""),
            getattr(row, "CODCARPR_ALIAS_LIST", ""),
        )
        if not codigo_unico or not nombre or not jornada or not codcarprs:
            continue
        for codcarpr in codcarprs:
            pref = _extract_alpha_prefix(codcarpr) or "X"
            rows.append(
                {
                    "GRUPO_TRAZA": f"DUR_{pref}",
                    "JORNADA": jornada,
                    "CODCARPR": codcarpr,
                    "NOMBRE_L": nombre,
                    "CODIGO_CARRERA_SIES": codigo_unico,
                }
            )

    if not rows:
        return pd.DataFrame(), pd.DataFrame()

    base = pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)
    cat_manual = _prepare_catalog_manual(base[["GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L"]])
    puente = _prepare_puente_sies(base)
    print(f"✅ Catálogo/puente auto desde DURACION_ESTUDIOS: catalogo={len(cat_manual)} puente={len(puente)}")
    return cat_manual, puente


# ==============================
# CAPA A: Ingesta flexible
# ==============================
def cargar_fuentes(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    book = pd.read_excel(path, sheet_name=None)
    req_hist = {"ANO", "PERIODO", "RUT", "DIG", "CODCARR", "CODRAMO"}
    try:
        carreras = _pick_sheet(book, {"CODIGO_UNICO", "PLAN_ESTUDIOS"})
        matricula = _pick_sheet(book, {"NUM_DOCUMENTO", "CODIGO_UNICO", "PLAN_ESTUDIOS", "DV"})
        hist = pd.concat([df for df in book.values() if req_hist.issubset(df.columns)], ignore_index=True)
        if hist.empty:
            raise ValueError("No se encontró histórico con ANO/PERIODO/RUT/DIG/CODCARR/CODRAMO")

        equiv = book.get("Equivalencia")
        if equiv is None:
            raise ValueError("Falta hoja Equivalencia")
        return carreras, matricula, hist.copy(), equiv.copy()
    except ValueError as err:
        # Modo compatible para archivos legacy tipo "Hoja1" con CODCARR/PLAN_DE_ESTUDIO.
        hoja1 = book.get("Hoja1")
        req_hoja1 = {"RUT", "DIG", "CODCARR", "PLAN_DE_ESTUDIO", "ANO", "PERIODO", "CODRAMO"}
        if hoja1 is None or not req_hoja1.issubset(hoja1.columns):
            raise err

        src = hoja1.copy()
        carreras = (
            src[["CODCARR", "PLAN_DE_ESTUDIO"]]
            .dropna(subset=["CODCARR", "PLAN_DE_ESTUDIO"])
            .drop_duplicates()
            .rename(columns={"CODCARR": "CODIGO_UNICO", "PLAN_DE_ESTUDIO": "PLAN_ESTUDIOS"})
        )

        mat_cols = ["RUT", "DIG", "CODCARR", "PLAN_DE_ESTUDIO"]
        if "ESTADO_ACADEMICO" in src.columns:
            mat_cols.append("ESTADO_ACADEMICO")
        matricula = (
            src[mat_cols]
            .dropna(subset=["RUT", "DIG", "CODCARR", "PLAN_DE_ESTUDIO"])
            .drop_duplicates()
            .rename(
                columns={
                    "RUT": "NUM_DOCUMENTO",
                    "DIG": "DV",
                    "CODCARR": "CODIGO_UNICO",
                    "PLAN_DE_ESTUDIO": "PLAN_ESTUDIOS",
                }
            )
        )
        matricula["TIPO_DOCUMENTO"] = "R"

        hist_cols = ["ANO", "PERIODO", "RUT", "DIG", "CODCARR", "CODRAMO"]
        extra_hist = [c for c in ["DESCRIPCION_ESTADO", "ESTADO_ACADEMICO", "JORNADA"] if c in src.columns]
        hist = src[hist_cols + extra_hist].copy()
        if "DESCRIPCION_ESTADO" not in hist.columns and "ESTADO" in src.columns:
            hist["DESCRIPCION_ESTADO"] = src["ESTADO"]

        equiv_cols = ["CODCARR"] + (["JORNADA"] if "JORNADA" in src.columns else [])
        equiv = src[equiv_cols].dropna(subset=["CODCARR"]).drop_duplicates().copy()
        equiv["CODIGO_UNICO"] = equiv["CODCARR"]

        print("Modo compatible activado: se derivaron Carreras/Matrícula/Equivalencia desde Hoja1.")
        return carreras, matricula, hist, equiv


def preparar_matricula_intermedia(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(out["NUM_DOCUMENTO"], out["DV"])]
    # Regla: no colapsar por RUT, solo deduplicación exacta de clave de negocio
    key = ["NUM_DOCUMENTO", "DV", "CODIGO_UNICO", "PLAN_ESTUDIOS"]
    return out.drop_duplicates(key)


def construir_puente_equiv(df_equiv: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    cols = [c for c in ["CODCARR", "CODIGO_UNICO", "JORNADA", "VERSION"] if c in df_equiv.columns]
    bridge = df_equiv[cols].dropna(subset=["CODCARR", "CODIGO_UNICO"]).copy()
    bridge["CODCARR"] = bridge["CODCARR"].astype(str)

    g = (
        bridge.groupby("CODCARR")
        .agg(
            codigos_unicos=("CODIGO_UNICO", "nunique"),
            jornadas=("JORNADA", "nunique") if "JORNADA" in bridge.columns else ("CODIGO_UNICO", "size"),
            versiones=("VERSION", "nunique") if "VERSION" in bridge.columns else ("CODIGO_UNICO", "size"),
        )
        .reset_index()
    )
    g["es_ambiguo"] = g["codigos_unicos"] > 1
    return bridge, g


def mapear_historico_con_equiv(df_hist: pd.DataFrame, bridge: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    hist = df_hist.copy()
    hist["RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(hist["RUT"], hist["DIG"])]
    hist["CODCARR"] = hist["CODCARR"].astype(str)
    if "CODIGO_UNICO" not in hist.columns:
        hist["CODIGO_UNICO"] = pd.NA

    key_jornada = {"CODCARR", "JORNADA"}.issubset(hist.columns) and {"CODCARR", "JORNADA"}.issubset(bridge.columns)

    if key_jornada:
        hist["JORNADA"] = hist["JORNADA"].astype(str)
        bridge = bridge.copy()
        bridge["JORNADA"] = bridge["JORNADA"].astype(str)
        pair = bridge.drop_duplicates(["CODCARR", "JORNADA", "CODIGO_UNICO"])
        counts = pair.groupby(["CODCARR", "JORNADA"])["CODIGO_UNICO"].nunique().reset_index(name="n")
        unique_pair = pair.merge(counts[counts["n"] == 1], on=["CODCARR", "JORNADA"], how="inner")
        unique_pair = unique_pair[["CODCARR", "JORNADA", "CODIGO_UNICO"]]
        hist = hist.merge(unique_pair, on=["CODCARR", "JORNADA"], how="left", suffixes=("", "_EQ"))
        if "CODIGO_UNICO_EQ" in hist.columns:
            hist["CODIGO_UNICO"] = hist["CODIGO_UNICO"].combine_first(hist["CODIGO_UNICO_EQ"])
            hist = hist.drop(columns=["CODIGO_UNICO_EQ"])

    unresolved = hist[hist["CODIGO_UNICO"].isna()].copy()
    if not unresolved.empty:
        single = bridge.groupby("CODCARR")["CODIGO_UNICO"].nunique().reset_index(name="n")
        single = single[single["n"] == 1].merge(
            bridge[["CODCARR", "CODIGO_UNICO"]].drop_duplicates("CODCARR"),
            on="CODCARR",
            how="left",
        )
        unresolved = unresolved.drop(columns=[c for c in ["CODIGO_UNICO"] if c in unresolved.columns]).merge(
            single[["CODCARR", "CODIGO_UNICO"]],
            on="CODCARR",
            how="left",
        )
        resolved = hist[hist["CODIGO_UNICO"].notna()].copy()
        hist = pd.concat([resolved, unresolved], ignore_index=True)

    review = hist[hist["CODIGO_UNICO"].isna()][["CODCARR"]].value_counts().reset_index(name="filas_sin_map")
    return hist, review


# ==============================
# CAPA B: Modelo intermedio
# ==============================
def construir_resumen_historico(hist_mapeado: pd.DataFrame) -> pd.DataFrame:
    valid = hist_mapeado[hist_mapeado["CODIGO_UNICO"].notna()].copy()
    if valid.empty:
        return pd.DataFrame(
            columns=[
                "RUT_NORM",
                "CODIGO_UNICO",
                "CURSO_1ER_SEM",
                "CURSO_2DO_SEM",
                "UNIDADES_CURSADAS",
                "UNIDADES_APROBADAS",
                "UNID_CURSADAS_TOTAL",
                "UNID_APROBADAS_TOTAL",
            ]
        )

    anio_vals = pd.to_numeric(valid["ANO"], errors="coerce").dropna()
    if anio_vals.empty:
        return pd.DataFrame(
            columns=[
                "RUT_NORM",
                "CODIGO_UNICO",
                "CURSO_1ER_SEM",
                "CURSO_2DO_SEM",
                "UNIDADES_CURSADAS",
                "UNIDADES_APROBADAS",
                "UNID_CURSADAS_TOTAL",
                "UNID_APROBADAS_TOTAL",
            ]
        )
    anio_ref = int(anio_vals.max())

    rows = []
    for (rut, cod), sub in valid.groupby(["RUT_NORM", "CODIGO_UNICO"]):
        s_ref = sub[sub["ANO"] == anio_ref]

        estado_ref = _series_or_default(s_ref, "DESCRIPCION_ESTADO").str.upper()
        estado_hist = _series_or_default(sub, "DESCRIPCION_ESTADO").str.upper()

        codramo_ref = s_ref["CODRAMO"] if "CODRAMO" in s_ref.columns else pd.Series(index=s_ref.index, dtype=object)
        codramo_hist = sub["CODRAMO"] if "CODRAMO" in sub.columns else pd.Series(index=sub.index, dtype=object)

        rows.append(
            {
                "RUT_NORM": rut,
                "CODIGO_UNICO": cod,
                "CURSO_1ER_SEM": "SI" if (s_ref["PERIODO"] == 1).any() else "NO",
                "CURSO_2DO_SEM": "SI" if (s_ref["PERIODO"] == 2).any() else "NO",
                "UNIDADES_CURSADAS": codramo_ref.nunique(),
                "UNIDADES_APROBADAS": codramo_ref[estado_ref.str.contains("APROB", na=False)].nunique(),
                "UNID_CURSADAS_TOTAL": codramo_hist.nunique(),
                "UNID_APROBADAS_TOTAL": codramo_hist[
                    estado_hist.str.contains(r"APROB|CONVALID|RECONOC|EQUIV|HOMOLOG", regex=True, na=False)
                ].nunique(),
            }
        )

    return pd.DataFrame(rows)


def construir_carreras_control(df_carreras: pd.DataFrame) -> pd.DataFrame:
    out = df_carreras.copy().drop_duplicates(["CODIGO_UNICO", "PLAN_ESTUDIOS"])
    for col in CARRERAS_AC_COLUMNS:
        if col not in out.columns:
            out[col] = pd.NA
    return out[CARRERAS_AC_COLUMNS].copy()


def construir_matricula_ac_control(df_matricula: pd.DataFrame, resumen: pd.DataFrame) -> pd.DataFrame:
    out = df_matricula.copy()
    if not resumen.empty:
        out = out.merge(resumen, on=["RUT_NORM", "CODIGO_UNICO"], how="left", suffixes=("", "_CALC"))

    calc_cols = [
        "CURSO_1ER_SEM",
        "CURSO_2DO_SEM",
        "UNIDADES_CURSADAS",
        "UNIDADES_APROBADAS",
        "UNID_CURSADAS_TOTAL",
        "UNID_APROBADAS_TOTAL",
    ]
    for col in calc_cols:
        if col not in out.columns and f"{col}_CALC" in out.columns:
            out[col] = out[f"{col}_CALC"]

    for col in MATRICULA_AC_COLUMNS:
        if col not in out.columns:
            out[col] = pd.NA
    return out[MATRICULA_AC_COLUMNS].copy()


def construir_matricula_unificada_control(mat_ac: pd.DataFrame, df_equiv: pd.DataFrame) -> pd.DataFrame:
    jmap: dict[object, tuple[object, object]] = {}
    if {"CODIGO_UNICO", "JORNADA"}.issubset(df_equiv.columns):
        for _, r in df_equiv[["CODIGO_UNICO", "JORNADA"]].dropna().drop_duplicates().iterrows():
            j = str(r["JORNADA"]).strip().lower()
            if "diurn" in j:
                jmap[r["CODIGO_UNICO"]] = ("PRESENCIAL", "1")
            elif "vespert" in j:
                jmap[r["CODIGO_UNICO"]] = ("PRESENCIAL", "2")
            elif "semi" in j:
                jmap[r["CODIGO_UNICO"]] = ("SEMIPRESENCIAL", "3")
            elif "dist" in j:
                jmap[r["CODIGO_UNICO"]] = ("DISTANCIA", "4")
            else:
                jmap[r["CODIGO_UNICO"]] = (pd.NA, pd.NA)

    out = pd.DataFrame(index=mat_ac.index)
    out["TIPO_DOC"] = mat_ac.get("TIPO_DOCUMENTO")
    out["N_DOC"] = mat_ac.get("NUM_DOCUMENTO")
    out["DV"] = mat_ac.get("DV")
    out["PRIMER_APELLIDO"] = mat_ac.get("PRIMER_APELLIDO")
    out["SEGUNDO_APELLIDO"] = mat_ac.get("SEGUNDO_APELLIDO")
    out["NOMBRE"] = mat_ac.get("NOMBRES")
    out["SEXO"] = mat_ac.get("SEXO")
    out["FECH_NAC"] = mat_ac.get("FECHA_NACIMIENTO")
    out["NAC"] = pd.NA
    if "PAIS_EST_SEC" in mat_ac.columns:
        out["PAIS_EST_SEC"] = mat_ac["PAIS_EST_SEC"]
    else:
        # Valor operativo por defecto cuando la fuente no incluye este campo obligatorio.
        out["PAIS_EST_SEC"] = "CL"
    out["COD_SED"] = pd.NA
    out["COD_CAR"] = mat_ac.get("CODIGO_UNICO")
    
    # Resolver MODALIDAD y JOR usando matriz de desambiguación
    out_sies, out_confianza, out_notas, is_ambiguo = [], [], [], []
    for idx, row in mat_ac.iterrows():
        codcarpr = row.get("CODIGO_UNICO", "")
        jornada_src = row.get("JORNADA", "")
        version_src = row.get("VERSION", "")
        
        # Extract VERSION from plan if not present - use plan's VERSION field
        if not version_src or pd.isna(version_src):
            plan_estudios = row.get("PLAN_ESTUDIOS", "")
            if plan_estudios and isinstance(plan_estudios, str):
                # Try to extract version from plan name (e.g., "Plan_V1_2024" -> "V1")
                match = re.search(r'(V\d+)', str(plan_estudios).upper())
                version_src = match.group(1) if match else "V1"
            else:
                version_src = "V1"
        
        codigo_sies, confianza, notas, ambiguo = resolver_ambiguedad_sies(codcarpr, jornada_src, version_src)
        out_sies.append(codigo_sies)
        out_confianza.append(confianza)
        out_notas.append(notas)
        is_ambiguo.append(ambiguo)

    out["MODALIDAD"] = pd.Series(out_sies, index=out.index)
    out["_SIES_CONFIANZA"] = pd.Series(out_confianza, index=out.index)
    out["_SIES_NOTAS"] = pd.Series(out_notas, index=out.index)
    out["_SIES_AMBIGUO"] = pd.Series(is_ambiguo, index=out.index)
    
    # Mapeo heredado de JORNADA (mantener compatibilidad con jmap si falla matriz)
    modality = mat_ac.get("CODIGO_UNICO").map(lambda c: jmap.get(c, (pd.NA, pd.NA)))
    legacy_mod, legacy_jor = zip(*modality.tolist()) if modality.tolist() else ([], [])
    
    # Usar matriz primero, fallback a legacy mapping
    out["JOR"] = pd.Series([legacy_jor[i] if pd.isna(out_sies[i]) else pd.NA for i in range(len(out_sies))], index=out.index)
    out["VERSION"] = pd.NA
    out["FOR_ING_ACT"] = pd.NA
    out["ANIO_ING_ACT"] = mat_ac.get("ANIO_INGRESO_CARRERA_ACTUAL")
    out["SEM_ING_ACT"] = mat_ac.get("SEM_INGRESO_CARRERA_ACTUAL")
    out["ANIO_ING_ORI"] = mat_ac.get("ANIO_INGRESO_CARRERA_ORIGEN")
    out["SEM_ING_ORI"] = mat_ac.get("SEM_INGRESO_CARRERA_ORIGEN")
    out["ASI_INS_ANT"] = mat_ac.get("UNIDADES_CURSADAS", 0)
    out["ASI_APR_ANT"] = mat_ac.get("UNIDADES_APROBADAS", 0)
    out["PROM_PRI_SEM"] = mat_ac.get("PROM_PRI_SEM", 0)
    out["PROM_SEG_SEM"] = mat_ac.get("PROM_SEG_SEM", 0)
    out["ASI_INS_HIS"] = mat_ac.get("UNID_CURSADAS_TOTAL", 0)
    out["ASI_APR_HIS"] = mat_ac.get("UNID_APROBADAS_TOTAL", 0)
    out["NIV_ACA"] = mat_ac.get("NIVEL", pd.NA)
    out["SIT_FON_SOL"] = mat_ac.get("SIT_FON_SOL", 0)
    out["SUS_PRE"] = mat_ac.get("SUS_PRE", 0)
    out["FECHA_MATRICULA"] = pd.NA
    if "REINCORPORACION" in mat_ac.columns:
        out["REINCORPORACION"] = mat_ac["REINCORPORACION"]
    else:
        out["REINCORPORACION"] = 0
    out["VIG"] = mat_ac.get("VIGENCIA")
    return out[MATRICULA_UNIFICADA_COLUMNS].copy()


def ejecutar_pipeline_matricula_unificada_legacy_like(
    input_file: Path,
    output_dir: Path,
    sheet_name: str | None = None,
    catalogo_manual_tsv_path: str | None = None,
    puente_sies_tsv_path: str | None = None,
    oferta_academica_xlsx_path: str | None = None,
    gob_nac_tsv_path: str | None = None,
    gob_pais_est_sec_tsv_path: str | None = None,
    gob_sede_tsv_path: str | None = None,
    excluir_diplomados: bool = DEFAULT_EXCLUIR_DIPLOMADOS,
    usar_gobernanza_v2: bool = False,
    filtro_base_datos_sheet: str | None = None,
) -> dict[str, object]:
    """
    Fase 1 de fusión con pipeline legacy:
    - Lee hoja fuente (primera por defecto).
    - Detecta columnas base CODCLI/CODCARR/JORNADA/CARRERA + RUT/DV.
    - Construye archivo tipo "ARCHIVO_LISTO_SUBIDA" con columnas de Matrícula Unificada
      y estados operativos del administrador de duplicados.
    """
    xls = pd.ExcelFile(input_file)
    selected_sheet = sheet_name or xls.sheet_names[0]
    src = pd.read_excel(input_file, sheet_name=selected_sheet)

    # --- Filtro por hoja base_datos: conservar solo filas cuyo RUT aparezca en la hoja ---
    # Auto-detectar: si no se pasó el flag pero existe "base_datos" en el Excel, usarla
    _filtro_bd_sheet = filtro_base_datos_sheet
    if _filtro_bd_sheet is None and "base_datos" in xls.sheet_names:
        _filtro_bd_sheet = "base_datos"
        print("  ℹ️ Auto-detectada hoja 'base_datos' → se aplicará filtro por RUT")

    _filtro_bd_stats: dict[str, object] = {}
    if _filtro_bd_sheet and _filtro_bd_sheet in xls.sheet_names:
        bd_df = pd.read_excel(input_file, sheet_name=_filtro_bd_sheet)
        bd_rut_col = None
        for _cand in ["N_DOC", "RUT", "NUM_DOCUMENTO"]:
            if _cand in bd_df.columns:
                bd_rut_col = _cand
                break
        if bd_rut_col is not None:
            bd_ruts = set(bd_df[bd_rut_col].dropna().astype(int))
            src_rut_col = None
            for _cand in ["RUT", "NUM_DOCUMENTO", "N_DOC"]:
                if _cand in src.columns:
                    src_rut_col = _cand
                    break
            if src_rut_col is not None:
                n_before = len(src)
                src = src[src[src_rut_col].astype(int).isin(bd_ruts)].reset_index(drop=True)
                n_after = len(src)
                _filtro_bd_stats = {
                    "hoja": _filtro_bd_sheet,
                    "ruts_en_hoja": len(bd_ruts),
                    "filas_antes": n_before,
                    "filas_despues": n_after,
                    "filas_descartadas": n_before - n_after,
                }
                print(f"  🔎 Filtro base_datos: {len(bd_ruts)} RUTs → {n_before} → {n_after} filas ({n_before - n_after} descartadas)")
            else:
                print(f"  ⚠️ Filtro base_datos: no se encontró columna RUT en hoja fuente")
        else:
            print(f"  ⚠️ Filtro base_datos: no se encontró columna N_DOC/RUT en hoja '{_filtro_bd_sheet}'")
    elif _filtro_bd_sheet:
        print(f"  ⚠️ Filtro base_datos: hoja '{_filtro_bd_sheet}' no existe en el Excel")

    # ── Depuración provisoria RUT ↔ CODCLI (pre-pipeline) ──────────────
    _pre_col_rut = _pick_first_column(src, ["RUT", "NUM_DOCUMENTO", "N_DOC"])
    _pre_col_codcli = _pick_first_column(src, ["CODCLI"])
    _stats_depur: dict[str, object] = {}
    if _pre_col_rut and _pre_col_codcli:
        from scripts.depurar_rut_multi_codcli import depurar_rut_multi_codcli
        src, _stats_depur = depurar_rut_multi_codcli(
            src, _pre_col_rut, _pre_col_codcli, output_dir,
        )
    # ────────────────────────────────────────────────────────────────────

    manual_source = catalogo_manual_tsv_path or "auto:DURACION_ESTUDIOS.tsv"
    puente_compilado_path = DEFAULT_PUENTE_SIES_COMPILADO_PATH.resolve()
    puente_source = str(puente_compilado_path)
    if puente_sies_tsv_path:
        print(
            "⚠️  --puente-sies-tsv recibido, pero no se consume directamente en el pipeline. "
            "Usa scripts/compile_puente_sies_compilado.py para materializar control/catalogos/PUENTE_SIES_COMPILADO.tsv."
        )
    oferta_dim = _load_oferta_academica_dim(input_file, oferta_academica_xlsx_path)
    oferta_source = (
        oferta_dim["OFERTA_SOURCE_PATH"].iloc[0]
        if not oferta_dim.empty and "OFERTA_SOURCE_PATH" in oferta_dim.columns
        else "not_found"
    )

    col_codcli = _pick_first_column(src, ["CODCLI"])
    col_codcarr = _pick_first_column(src, ["CODCARR", "CODCARPR"])
    col_nombre_carrera = _pick_first_column(src, ["CARRERA", "NOMBRE_L", "NOMBRE"])
    col_jornada = _pick_first_column(src, ["JORNADA"])
    col_rut = _pick_first_column(src, ["RUT", "NUM_DOCUMENTO", "N_DOC"])
    col_dv = _pick_first_column(src, ["DIG", "DV"])

    missing = [
        name
        for name, col in {
            "CODCLI": col_codcli,
            "CODCARR/CODCARPR": col_codcarr,
            "CARRERA/NOMBRE_L/NOMBRE": col_nombre_carrera,
            "JORNADA": col_jornada,
            "RUT/NUM_DOCUMENTO/N_DOC": col_rut,
            "DIG/DV": col_dv,
        }.items()
        if col is None
    ]
    if missing:
        raise ValueError(
            "No fue posible ejecutar la fase de Matrícula Unificada con la fuente entregada. "
            f"Faltan columnas base: {missing}"
        )

    req_codcli = _require_column(col_codcli, "CODCLI")
    req_codcarr = _require_column(col_codcarr, "CODCARR/CODCARPR")
    req_nombre_carrera = _require_column(col_nombre_carrera, "CARRERA/NOMBRE_L/NOMBRE")
    req_jornada = _require_column(col_jornada, "JORNADA")
    req_rut = _require_column(col_rut, "RUT/NUM_DOCUMENTO/N_DOC")
    req_dv = _require_column(col_dv, "DIG/DV")

    col_nombre = _pick_first_column(src, ["NOMBRE"]) or req_nombre_carrera
    col_pat = _pick_first_column(src, ["PATERNO", "PRIMER_APELLIDO"])
    col_mat = _pick_first_column(src, ["MATERNO", "SEGUNDO_APELLIDO"])
    col_sexo = _pick_first_column(src, ["SEXO"])
    col_fecha_nac = _pick_first_column(src, ["FECHANACIMIENTO", "FECHA_NACIMIENTO"])
    col_anio_ing = _pick_first_column(src, ["ANOINGRESO", "ANIO_INGRESO_CARRERA_ACTUAL"])
    col_sem_ing = _pick_first_column(src, ["PERIODOINGRESO", "SEM_INGRESO_CARRERA_ACTUAL"])
    col_fecha_matricula = _pick_first_column(src, ["FECHAMATRICULA", "FECHA_MATRICULA"])
    col_vig = _pick_first_column(src, ["VIGENCIA", "VIG"])
    col_nac = _pick_first_column(src, ["NACIONALIDAD", "NAC"])
    col_cod_sed = _pick_first_column(src, ["COD_SED"])
    col_plan = _pick_first_column(src, ["PLAN_DE_ESTUDIO", "PLAN_ESTUDIOS"])
    col_periodo = _pick_first_column(src, ["PERIODO"])
    col_asi_ins_ant = _pick_first_column(src, ["ASI_INS_ANT"])
    col_asi_apr_ant = _pick_first_column(src, ["ASI_APR_ANT"])
    col_prom_pri_sem = _pick_first_column(src, ["PROM_PRI_SEM"])
    col_prom_seg_sem = _pick_first_column(src, ["PROM_SEG_SEM"])
    col_asi_ins_his = _pick_first_column(src, ["ASI_INS_HIS"])
    col_asi_apr_his = _pick_first_column(src, ["ASI_APR_HIS"])
    col_niv_aca = _pick_first_column(src, ["NIV_ACA", "NIVEL"])
    col_sit_fon_sol = _pick_first_column(src, ["SIT_FON_SOL"])
    col_sus_pre = _pick_first_column(src, ["SUS_PRE"])
    col_for_ing_act = _pick_first_column(
        src,
        ["FOR_ING_ACT", "FORMA_INGRESO", "FORMAINGRESO", "FORMA_INGRESO_ACTUAL", "TIPO_INGRESO"],
    )

    src_work = src.copy()
    rows_enriquecidas_datos_alumnos = 0
    cod_sed_resueltos_regla = 0
    pais_est_sec_inferidos_localidad = 0
    sin_match_datos_alumnos_rows = 0
    da_match_modo = pd.Series("SIN_MATCH", index=src_work.index, dtype="object")

    gob_nac_df = _load_governance_tsv(
        gob_nac_tsv_path,
        ["NACIONALIDAD_NORM", "COD_NAC"],
    )
    gob_pais_est_sec_df = _load_governance_tsv(
        gob_pais_est_sec_tsv_path,
        ["COMUNACOLEGIO_NORM", "CIUDADCOLEGIO_NORM", "COD_PAIS_EST_SEC"],
    )
    gob_sede_df = _load_governance_tsv(
        gob_sede_tsv_path,
        ["SEDE_NORM", "COD_SED", "NOMBRE_SEDE"],
    )
    gob_hoja1_estado_desc_path = _first_existing_path(DEFAULT_GOB_HOJA1_ESTADO_DESC_CANDIDATES)
    gob_da_estado_situ_path = _first_existing_path(DEFAULT_GOB_DA_ESTADO_SITUACION_CANDIDATES)
    gob_hoja1_estado_desc_df = _load_governance_tsv(
        str(gob_hoja1_estado_desc_path) if gob_hoja1_estado_desc_path else None,
        ["ESTADO_ACADEMICO", "DESCRIPCION_ESTADO", "VIG_ESPERADO"],
    )
    gob_da_estado_situ_df = _load_governance_tsv(
        str(gob_da_estado_situ_path) if gob_da_estado_situ_path else None,
        ["ESTADOACADEMICO", "SITUACION", "VIG_ESPERADO"],
    )
    valid_for_ing_act_codes, gob_for_ing_act_source = _load_for_ing_act_catalog()

    # ── Período objetivo del run (para ANIO_ANTERIOR dinámico por período) ──
    if col_anio_ing and col_anio_ing in src.columns:
        _pf_anio = pd.to_numeric(src[col_anio_ing], errors="coerce").dropna()
        _pf_anio = _pf_anio[_pf_anio.between(1990, 2100)]
        periodo_filtro_anio = int(_pf_anio.mode().iloc[0]) if not _pf_anio.empty else 2026
    else:
        periodo_filtro_anio = 2026
    if col_sem_ing and col_sem_ing in src.columns:
        _pf_sem = pd.to_numeric(src[col_sem_ing], errors="coerce").dropna()
        _pf_sem = _pf_sem[_pf_sem.isin([1, 2, 3])].replace({3: 2})
        periodo_filtro_sem = int(_pf_sem.mode().iloc[0]) if not _pf_sem.empty else 1
    else:
        periodo_filtro_sem = 1
    anio_anterior_prom = periodo_filtro_anio - 1
    print(f"📌 Año anterior PROM por período: {anio_anterior_prom} (ANIO_ING_ACT={periodo_filtro_anio}, SEM={periodo_filtro_sem})")

    historico_mu_df, anio_ref_historico_mu = _build_mu_historico_summary(
        src, req_rut, req_dv, req_codcarr, anio_ref_override=anio_anterior_prom,
    )
    if not historico_mu_df.empty:
        src_work["_HIST_MU_KEY"] = [
            _normalize_doc(n, d) + "|" + _normalize_text(c)
            for n, d, c in zip(src_work[req_rut], src_work[req_dv], src_work[req_codcarr])
        ]
        historico_mu_join = historico_mu_df.copy()
        historico_mu_join["_HIST_MU_KEY"] = historico_mu_join["RUT_NORM"] + "|" + historico_mu_join["CODCARPR_NORM"]
        src_work = src_work.merge(
            historico_mu_join[
                [
                    "_HIST_MU_KEY",
                    "UZ_HIST_KEY",
                    "ANIO_REFERENCIA_HIST_UZ",
                    "UZ_HIST_ANIO_MIN",
                    "UZ_HIST_ANIO_MAX",
                    "UZ_HIST_ANIOS_DISPONIBLES",
                    "UZ_HIST_SCOPE_STATUS",
                    "UZ_HIST_FILAS_TOTAL",
                    "UZ_HIST_FILAS_ANIO_REFERENCIA",
                    "UZ_HIST_FILAS_REF_APROB",
                    "UZ_HIST_FILAS_REF_REPROB",
                    "UZ_HIST_FILAS_REF_TRANSFER",
                    "UZ_HIST_FILAS_REF_SEM1_CALIFICADAS",
                    "UZ_HIST_FILAS_REF_SEM2_CALIFICADAS",
                    "ASI_INS_ANT_HIST",
                    "ASI_APR_ANT_HIST",
                    "PROM_PRI_SEM_HIST",
                    "PROM_SEG_SEM_HIST",
                    "ASI_INS_HIS_HIST",
                    "ASI_APR_HIS_HIST",
                    "UZ_FUENTE_HIST",
                ]
            ],
            on="_HIST_MU_KEY",
            how="left",
        )

    # Nuevo flujo (v2) sólo con flag para facilitar rollback inmediato.
    if usar_gobernanza_v2:
        da_lookup = _load_datos_alumnos_lookup(input_file)
        if not da_lookup.empty:
            src_work[req_codcli] = src_work[req_codcli].astype(str).str.strip()
            src_work = src_work.merge(da_lookup, on=req_codcli, how="left")
            has_codcli_match = src_work["DA_MATCH_FLAG"].fillna("") == "1" if "DA_MATCH_FLAG" in src_work.columns else pd.Series(False, index=src_work.index)
            da_match_modo.loc[has_codcli_match] = "MATCH_CODCLI"

            # Fallback controlado: solo para no encontrados por CODCLI.
            if "DA_RUT_NORM" in da_lookup.columns:
                src_work["_SRC_RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(src_work[req_rut], src_work[req_dv])]
                da_by_rut = (
                    da_lookup[da_lookup["DA_RUT_NORM"].astype(str).str.strip() != ""]
                    .drop_duplicates(subset=["DA_RUT_NORM"], keep="first")
                    .copy()
                )
                if not da_by_rut.empty:
                    da_fill_cols = [
                        c
                        for c in da_by_rut.columns
                        if c.startswith("DA_") and c not in {"DA_MATCH_FLAG", "DA_RUT_NORM"}
                    ]
                    rut_right = ["DA_RUT_NORM"] + da_fill_cols
                    src_work = src_work.merge(
                        da_by_rut[rut_right].add_suffix("_BY_RUT"),
                        left_on="_SRC_RUT_NORM",
                        right_on="DA_RUT_NORM_BY_RUT",
                        how="left",
                    )

                    needs_rut_fill = da_match_modo == "SIN_MATCH"
                    rut_match_mask = needs_rut_fill & src_work["DA_RUT_NORM_BY_RUT"].notna()
                    for col in da_fill_cols:
                        col_rut = f"{col}_BY_RUT"
                        if col in src_work.columns and col_rut in src_work.columns:
                            src_work.loc[rut_match_mask, col] = src_work.loc[rut_match_mask, col].combine_first(
                                src_work.loc[rut_match_mask, col_rut]
                            )
                        elif col_rut in src_work.columns:
                            src_work[col] = pd.NA
                            src_work.loc[rut_match_mask, col] = src_work.loc[rut_match_mask, col_rut]
                    da_match_modo.loc[rut_match_mask] = "MATCH_RUT"

            rows_enriquecidas_datos_alumnos = int((da_match_modo != "SIN_MATCH").sum())
            sin_match_datos_alumnos_rows = int((da_match_modo == "SIN_MATCH").sum())

    def _na_series() -> pd.Series:
        return pd.Series(pd.NA, index=src_work.index, dtype="object")

    tipo_doc_status = pd.Series("REGLA_FIJA_R", index=src_work.index, dtype="object")

    n_doc_source = src_work[req_rut]
    n_doc_status = pd.Series("SOURCE_INPUT", index=src_work.index, dtype="object")
    n_doc_status.loc[~_nonempty_mask(n_doc_source)] = "SIN_FUENTE_INPUT"

    dv_source = src_work[req_dv]
    dv_status = pd.Series("SOURCE_INPUT", index=src_work.index, dtype="object")
    dv_status.loc[~_nonempty_mask(dv_source)] = "SIN_FUENTE_INPUT"

    out = pd.DataFrame(index=src_work.index)
    out["TIPO_DOC"] = "R"
    out["N_DOC"] = n_doc_source
    out["DV"] = dv_source

    out["PRIMER_APELLIDO"] = src_work[col_pat] if col_pat else _na_series()
    if usar_gobernanza_v2 and "DA_APELLIDO_PATERNO" in src_work.columns:
        out["PRIMER_APELLIDO"] = out["PRIMER_APELLIDO"].combine_first(src_work["DA_APELLIDO_PATERNO"])

    segundo_apellido_source = src_work[col_mat] if col_mat else _na_series()
    segundo_apellido_input_mask = _nonempty_mask(segundo_apellido_source)
    segundo_apellido_status = pd.Series("VACIO_SIN_FUENTE", index=src_work.index, dtype="object")
    segundo_apellido_status.loc[segundo_apellido_input_mask] = "SOURCE_INPUT"
    out["SEGUNDO_APELLIDO"] = segundo_apellido_source
    if usar_gobernanza_v2 and "DA_APELLIDO_MATERNO" in src_work.columns:
        segundo_apellido_da_mask = (~segundo_apellido_input_mask) & _nonempty_mask(src_work["DA_APELLIDO_MATERNO"])
        out["SEGUNDO_APELLIDO"] = out["SEGUNDO_APELLIDO"].combine_first(src_work["DA_APELLIDO_MATERNO"])
        segundo_apellido_status.loc[segundo_apellido_da_mask] = "FALLBACK_DATOS_ALUMNOS"

    out["NOMBRE"] = src_work[col_nombre] if col_nombre else _na_series()
    if usar_gobernanza_v2 and "DA_NOMBRES" in src_work.columns:
        out["NOMBRE"] = out["NOMBRE"].combine_first(src_work["DA_NOMBRES"])

    out["SEXO"] = src_work[col_sexo] if col_sexo else _na_series()
    if usar_gobernanza_v2 and "DA_SEXO" in src_work.columns:
        out["SEXO"] = out["SEXO"].combine_first(src_work["DA_SEXO"])

    fecha_nac_source = src_work[col_fecha_nac] if col_fecha_nac else _na_series()
    fecha_nac_input_mask = _nonempty_mask(fecha_nac_source)
    fech_nac_status = pd.Series("SIN_FUENTE", index=src_work.index, dtype="object")
    fech_nac_status.loc[fecha_nac_input_mask] = "SOURCE_INPUT"
    out["FECH_NAC"] = fecha_nac_source
    if usar_gobernanza_v2 and "DA_FECHANACIMIENTO" in src_work.columns:
        fecha_nac_da_mask = (~fecha_nac_input_mask) & _nonempty_mask(src_work["DA_FECHANACIMIENTO"])
        out["FECH_NAC"] = out["FECH_NAC"].combine_first(src_work["DA_FECHANACIMIENTO"])
        fech_nac_status.loc[fecha_nac_da_mask] = "FALLBACK_DATOS_ALUMNOS"

    out["NAC"] = src_work[col_nac] if col_nac else _na_series()
    if usar_gobernanza_v2 and "DA_NACIONALIDAD" in src_work.columns:
        out["NAC"] = out["NAC"].combine_first(src_work["DA_NACIONALIDAD"])
    nac_status = pd.Series("SIN_INSUMO", index=src_work.index, dtype="object")
    nac_status.loc[out["NAC"].notna()] = "SOURCE_TEXT"

    if usar_gobernanza_v2 and not gob_nac_df.empty:
        gob_nac = gob_nac_df.copy()
        gob_nac["NACIONALIDAD_NORM"] = gob_nac["NACIONALIDAD_NORM"].map(_normalize_text)
        nac_map = (
            gob_nac[gob_nac["NACIONALIDAD_NORM"] != ""]
            .drop_duplicates(subset=["NACIONALIDAD_NORM"], keep="first")
            .set_index("NACIONALIDAD_NORM")["COD_NAC"]
            .to_dict()
        )
        nac_state_map = (
            gob_nac[gob_nac["NACIONALIDAD_NORM"] != ""]
            .drop_duplicates(subset=["NACIONALIDAD_NORM"], keep="first")
            .set_index("NACIONALIDAD_NORM")
            .get("ESTADO_GOBERNANZA", pd.Series(dtype="object"))
            .to_dict()
        )
        nac_norm = out["NAC"].map(_normalize_text)
        nac_code = nac_norm.map(nac_map).map(_normalize_code_or_na)
        has_code = nac_code.notna()
        out.loc[has_code, "NAC"] = nac_code.loc[has_code]
        nac_status.loc[has_code] = "MAPEADO_GOB_NAC"
        nac_state = nac_norm.map(nac_state_map).fillna("").astype(str).str.upper()
        review_manual_mask = (~has_code) & (nac_norm != "") & (nac_state == "REVISION_MANUAL")
        out.loc[review_manual_mask, "NAC"] = pd.NA
        nac_status.loc[review_manual_mask] = "REVISION_MANUAL_GOB_NAC"
        nac_status.loc[(~has_code) & (nac_norm != "") & (~review_manual_mask)] = "SIN_MAPEO_GOB_NAC"

    out["PAIS_EST_SEC"] = src_work["PAIS_EST_SEC"] if "PAIS_EST_SEC" in src_work.columns else _na_series()
    pais_est_sec_status = pd.Series("SIN_INSUMO", index=src_work.index, dtype="object")
    if "PAIS_EST_SEC" in src_work.columns:
        pais_est_sec_status = pd.Series("SOURCE_EMPTY", index=src_work.index, dtype="object")
        pais_est_sec_status.loc[out["PAIS_EST_SEC"].notna()] = "SOURCE_EXACT"

    if usar_gobernanza_v2 and ("DA_COMUNACOLEGIO" in src_work.columns or "DA_CIUDADCOLEGIO" in src_work.columns):
        comuna_norm = (
            src_work["DA_COMUNACOLEGIO"].fillna("").map(_normalize_text)
            if "DA_COMUNACOLEGIO" in src_work.columns
            else pd.Series("", index=src_work.index, dtype="object")
        )
        ciudad_norm = (
            src_work["DA_CIUDADCOLEGIO"].fillna("").map(_normalize_text)
            if "DA_CIUDADCOLEGIO" in src_work.columns
            else pd.Series("", index=src_work.index, dtype="object")
        )
        if not gob_pais_est_sec_df.empty:
            gob_pais = gob_pais_est_sec_df.copy()
            gob_pais["COMUNACOLEGIO_NORM"] = gob_pais["COMUNACOLEGIO_NORM"].map(_normalize_text)
            gob_pais["CIUDADCOLEGIO_NORM"] = gob_pais["CIUDADCOLEGIO_NORM"].map(_normalize_text)
            gob_pais["KEY_BOTH"] = gob_pais["COMUNACOLEGIO_NORM"] + "|" + gob_pais["CIUDADCOLEGIO_NORM"]
            gob_pais["KEY_COMUNA"] = gob_pais["COMUNACOLEGIO_NORM"]
            gob_pais["KEY_CIUDAD"] = gob_pais["CIUDADCOLEGIO_NORM"]

            map_both = (
                gob_pais[gob_pais["KEY_BOTH"] != "|"]
                .drop_duplicates(subset=["KEY_BOTH"], keep="first")
                .set_index("KEY_BOTH")["COD_PAIS_EST_SEC"]
                .to_dict()
            )
            map_comuna = (
                gob_pais[gob_pais["KEY_COMUNA"] != ""]
                .drop_duplicates(subset=["KEY_COMUNA"], keep="first")
                .set_index("KEY_COMUNA")["COD_PAIS_EST_SEC"]
                .to_dict()
            )
            map_ciudad = (
                gob_pais[gob_pais["KEY_CIUDAD"] != ""]
                .drop_duplicates(subset=["KEY_CIUDAD"], keep="first")
                .set_index("KEY_CIUDAD")["COD_PAIS_EST_SEC"]
                .to_dict()
            )

            key_both = comuna_norm + "|" + ciudad_norm
            mapped_both = key_both.map(map_both).map(_normalize_code_or_na)
            mapped_comuna = comuna_norm.map(map_comuna).map(_normalize_code_or_na)
            mapped_ciudad = ciudad_norm.map(map_ciudad).map(_normalize_code_or_na)
            mapped_final = mapped_both.combine_first(mapped_comuna).combine_first(mapped_ciudad)
            default_rows = gob_pais[
                (gob_pais["COMUNACOLEGIO_NORM"] == "") & (gob_pais["CIUDADCOLEGIO_NORM"] == "")
            ]
            if not default_rows.empty:
                default_code = _normalize_code_or_na(default_rows.iloc[0]["COD_PAIS_EST_SEC"])
                has_localidad = (comuna_norm != "") | (ciudad_norm != "")
                if not pd.isna(default_code):
                    mapped_final = mapped_final.copy()
                    default_gob_mask = has_localidad & mapped_final.isna()
                    mapped_final.loc[has_localidad] = mapped_final.loc[has_localidad].combine_first(
                        pd.Series(default_code, index=mapped_final.loc[has_localidad].index)
                    )
                else:
                    default_gob_mask = pd.Series(False, index=src_work.index, dtype=bool)
            else:
                default_gob_mask = pd.Series(False, index=src_work.index, dtype=bool)

            before_na = out["PAIS_EST_SEC"].isna()
            out["PAIS_EST_SEC"] = out["PAIS_EST_SEC"].combine_first(mapped_final)
            mapped_mask = before_na & out["PAIS_EST_SEC"].notna()
            pais_est_sec_inferidos_localidad = int(mapped_mask.sum())
            pais_est_sec_status.loc[mapped_mask] = "MAPEADO_GOB_PAIS_EST_SEC"
            pais_est_sec_status.loc[before_na & default_gob_mask & out["PAIS_EST_SEC"].notna()] = "DEFAULT_GOB_PAIS_EST_SEC"
        else:
            # Fallback legacy mientras no exista tabla maestra persistida.
            has_localidad = (comuna_norm != "") | (ciudad_norm != "")
            inferido_chile = pd.Series(pd.NA, index=src_work.index, dtype="object")
            inferido_chile.loc[has_localidad] = "38"

            before_na = out["PAIS_EST_SEC"].isna()
            out["PAIS_EST_SEC"] = out["PAIS_EST_SEC"].combine_first(inferido_chile)
            inferidos_mask = before_na & out["PAIS_EST_SEC"].notna()
            pais_est_sec_inferidos_localidad = int(inferidos_mask.sum())
            pais_est_sec_status.loc[inferidos_mask] = "INFERIDO_LOCALIDAD_CHILE_38"

    if col_cod_sed:
        out["COD_SED"] = src_work[col_cod_sed]
        cod_sed_status = pd.Series("SOURCE_EXACT", index=src_work.index, dtype="object")
    else:
        out["COD_SED"] = _na_series()
        cod_sed_status = pd.Series("SIN_INSUMO", index=src_work.index, dtype="object")
        if usar_gobernanza_v2 and "DA_SEDE" in src_work.columns:
            sede_norm = src_work["DA_SEDE"].fillna("").map(_normalize_text)
            if not gob_sede_df.empty:
                gob_sede = gob_sede_df.copy()
                gob_sede["SEDE_NORM"] = gob_sede["SEDE_NORM"].map(_normalize_text)
                sede_map = (
                    gob_sede[gob_sede["SEDE_NORM"] != ""]
                    .drop_duplicates(subset=["SEDE_NORM"], keep="first")
                    .set_index("SEDE_NORM")["COD_SED"]
                    .to_dict()
                )
                cod_sed_map = sede_norm.map(sede_map)
            else:
                cod_sed_map = sede_norm.map({"RE": "2", "CO": "3"})
            out["COD_SED"] = out["COD_SED"].combine_first(cod_sed_map)
            cod_sed_resueltos_regla = int(cod_sed_map.notna().sum())
            cod_sed_status.loc[cod_sed_map.notna()] = "MAPEADO_GOB_SEDE"
            cod_sed_status.loc[(cod_sed_map.isna()) & (sede_norm != "")] = "SIN_MAPEO_GOB_SEDE"

    out["COD_CAR"] = src_work[req_codcarr]

    modalidad, jor = _map_jornada_to_mod_jor(src_work[req_jornada])
    out["MODALIDAD"] = modalidad
    out["JOR"] = jor

    out["VERSION"] = pd.NA
    for_ing_trace_df = src_work.apply(
        lambda row: pd.Series(
            _resolve_for_ing_act_row(
                row[col_for_ing_act] if col_for_ing_act else pd.NA,
                row["DA_VIASDEADMISION"] if "DA_VIASDEADMISION" in row.index else pd.NA,
                row[req_nombre_carrera],
                row[req_codcarr],
                valid_for_ing_act_codes,
            )
        ),
        axis=1,
    )
    src_work["FOR_ING_ACT_FUENTE_VALOR"] = for_ing_trace_df["FOR_ING_ACT_FUENTE_VALOR"]
    src_work["FOR_ING_ACT_FUENTE_CAMPO"] = for_ing_trace_df["FOR_ING_ACT_FUENTE_CAMPO"]
    src_work["FOR_ING_ACT_FUENTE_NORM"] = for_ing_trace_df["FOR_ING_ACT_FUENTE_NORM"]
    src_work["FOR_ING_ACT_METODO"] = for_ing_trace_df["FOR_ING_ACT_METODO"]
    src_work["FOR_ING_ACT_IMPUTADO"] = for_ing_trace_df["FOR_ING_ACT_IMPUTADO"]
    src_work["FOR_ING_ACT_REQUIERE_REVISION"] = for_ing_trace_df["FOR_ING_ACT_REQUIERE_REVISION"]
    src_work["FOR_ING_ACT_RESUELTO"] = for_ing_trace_df["FOR_ING_ACT"]
    out["FOR_ING_ACT"] = src_work["FOR_ING_ACT_RESUELTO"]

    anio_input_label = f"INPUT_{col_anio_ing.upper().replace(' ', '_')}" if col_anio_ing else "SIN_COLUMNA_INPUT"
    sem_input_label = f"INPUT_{col_sem_ing.upper().replace(' ', '_')}" if col_sem_ing else "SIN_COLUMNA_INPUT"
    niv_input_label = f"INPUT_{col_niv_aca.upper().replace(' ', '_')}" if col_niv_aca else "SIN_COLUMNA_INPUT"
    fecha_mat_input_label = f"INPUT_{col_fecha_matricula.upper().replace(' ', '_')}" if col_fecha_matricula else "SIN_COLUMNA_INPUT"
    sit_fon_input_label = f"INPUT_{col_sit_fon_sol.upper().replace(' ', '_')}" if col_sit_fon_sol else "SIN_COLUMNA_INPUT"
    sus_pre_input_label = f"INPUT_{col_sus_pre.upper().replace(' ', '_')}" if col_sus_pre else "SIN_COLUMNA_INPUT"
    vig_input_label = f"INPUT_{col_vig.upper().replace(' ', '_')}" if col_vig else "SIN_COLUMNA_INPUT"

    anio_input_raw = src_work[col_anio_ing] if col_anio_ing else _na_series()
    anio_input = anio_input_raw.map(_to_int_year)
    anio_da_raw = src_work["DA_ANOINGRESO"] if "DA_ANOINGRESO" in src_work.columns else _na_series()
    anio_da = anio_da_raw.map(_to_int_year)
    anio_codcli = src_work[req_codcli].map(_infer_year_from_codcli)
    anio_input_valid = anio_input.between(1990, 2026)
    anio_da_valid = anio_da.between(1990, 2026)
    anio_codcli_valid = anio_codcli.between(1990, 2026)
    anio_act_source = pd.Series("REGLA_DEFAULT_2026", index=src_work.index, dtype="object")
    anio_act_method = pd.Series("DEFAULT_2026", index=src_work.index, dtype="object")
    anio_act_audit = pd.Series("DEFAULT_2026_SIN_FUENTE", index=src_work.index, dtype="object")
    anio_act_source.loc[anio_input_valid] = anio_input_label
    anio_act_method.loc[anio_input_valid] = "SOURCE_INPUT"
    anio_act_audit.loc[anio_input_valid] = "SOURCE_INPUT_VALIDO"
    anio_da_mask = (~anio_input_valid) & anio_da_valid
    anio_act_source.loc[anio_da_mask] = "DA_ANOINGRESO"
    anio_act_method.loc[anio_da_mask] = "FALLBACK_DATOS_ALUMNOS"
    anio_act_audit.loc[anio_da_mask] = "FALLBACK_DA_VALIDO"
    anio_codcli_mask = (~anio_input_valid) & (~anio_da_valid) & anio_codcli_valid
    anio_act_source.loc[anio_codcli_mask] = "CODCLI"
    anio_act_method.loc[anio_codcli_mask] = "INFERENCIA_PREFIJO_CODCLI"
    anio_act_audit.loc[anio_codcli_mask] = "INFERIDO_CODCLI"
    anio_input_invalid_only = (~anio_input_valid) & (~anio_da_valid) & (~anio_codcli_valid) & _nonempty_mask(anio_input_raw) & ~_nonempty_mask(anio_da_raw)
    anio_da_invalid_only = (~anio_input_valid) & (~anio_da_valid) & (~_nonempty_mask(anio_input_raw)) & _nonempty_mask(anio_da_raw)
    anio_input_da_invalid = (~anio_input_valid) & (~anio_da_valid) & _nonempty_mask(anio_input_raw) & _nonempty_mask(anio_da_raw)
    anio_act_audit.loc[anio_input_invalid_only] = "DEFAULT_2026_INPUT_INVALIDO"
    anio_act_audit.loc[anio_da_invalid_only] = "DEFAULT_2026_DA_INVALIDO"
    anio_act_audit.loc[anio_input_da_invalid] = "DEFAULT_2026_INPUT_DA_INVALIDOS"

    sem_input_raw = src_work[col_sem_ing] if col_sem_ing else _na_series()
    sem_input = pd.to_numeric(sem_input_raw, errors="coerce")
    sem_da_raw = src_work["DA_PERIODOINGRESO"] if "DA_PERIODOINGRESO" in src_work.columns else _na_series()
    sem_da = pd.to_numeric(sem_da_raw, errors="coerce")
    sem_input_valid = sem_input.isin([1, 2, 3])
    sem_da_valid = sem_da.isin([1, 2, 3])
    sem_act_source = pd.Series("REGLA_DEFAULT_1", index=src_work.index, dtype="object")
    sem_act_method = pd.Series("DEFAULT_1", index=src_work.index, dtype="object")
    sem_act_audit = pd.Series("DEFAULT_1_SIN_FUENTE", index=src_work.index, dtype="object")
    sem_act_source.loc[sem_input_valid] = sem_input_label
    sem_act_method.loc[sem_input_valid] = "SOURCE_INPUT"
    sem_act_audit.loc[sem_input_valid & sem_input.ne(3)] = "SOURCE_INPUT_VALIDO"
    sem_act_audit.loc[sem_input_valid & sem_input.eq(3)] = "SOURCE_INPUT_NORMALIZADO_3_A_2"
    sem_da_mask = (~sem_input_valid) & sem_da_valid
    sem_act_source.loc[sem_da_mask] = "DA_PERIODOINGRESO"
    sem_act_method.loc[sem_da_mask] = "FALLBACK_DATOS_ALUMNOS"
    sem_act_audit.loc[sem_da_mask & sem_da.ne(3)] = "FALLBACK_DA_VALIDO"
    sem_act_audit.loc[sem_da_mask & sem_da.eq(3)] = "FALLBACK_DA_NORMALIZADO_3_A_2"
    sem_input_invalid_only = (~sem_input_valid) & (~sem_da_valid) & _nonempty_mask(sem_input_raw) & ~_nonempty_mask(sem_da_raw)
    sem_da_invalid_only = (~sem_input_valid) & (~sem_da_valid) & (~_nonempty_mask(sem_input_raw)) & _nonempty_mask(sem_da_raw)
    sem_input_da_invalid = (~sem_input_valid) & (~sem_da_valid) & _nonempty_mask(sem_input_raw) & _nonempty_mask(sem_da_raw)
    sem_act_audit.loc[sem_input_invalid_only] = "DEFAULT_1_INPUT_INVALIDO"
    sem_act_audit.loc[sem_da_invalid_only] = "DEFAULT_1_DA_INVALIDO"
    sem_act_audit.loc[sem_input_da_invalid] = "DEFAULT_1_INPUT_DA_INVALIDOS"

    niv_input_raw = src_work[col_niv_aca] if col_niv_aca else _na_series()
    niv_input = pd.to_numeric(niv_input_raw, errors="coerce")
    niv_da_raw = src_work["DA_NIVEL"] if "DA_NIVEL" in src_work.columns else _na_series()
    niv_da = pd.to_numeric(niv_da_raw, errors="coerce")
    niv_input_valid = niv_input.notna() & niv_input.between(1, 20)
    niv_da_valid = niv_da.notna() & niv_da.between(1, 20)
    niv_input_label = col_niv_aca if col_niv_aca else "SIN_COLUMNA_NIVEL"
    niv_aca_source = pd.Series("REGLA_DEFAULT_1", index=src_work.index, dtype="object")
    niv_aca_method = pd.Series("DEFAULT_1", index=src_work.index, dtype="object")
    niv_aca_audit = pd.Series("DEFAULT_1_SIN_FUENTE", index=src_work.index, dtype="object")
    # Prioridad: DA_NIVEL (nivel alumno en DatosAlumnos) > INPUT_NIVEL (nivel del ramo en malla Hoja1)
    niv_aca_source.loc[niv_da_valid] = "DA_NIVEL"
    niv_aca_method.loc[niv_da_valid] = "PRIMARY_DATOS_ALUMNOS"
    niv_aca_audit.loc[niv_da_valid] = "PRIMARY_DA_VALIDO"
    niv_input_fallback_mask = (~niv_da_valid) & niv_input_valid
    niv_aca_source.loc[niv_input_fallback_mask] = niv_input_label
    niv_aca_method.loc[niv_input_fallback_mask] = "FALLBACK_INPUT"
    niv_aca_audit.loc[niv_input_fallback_mask] = "FALLBACK_INPUT_VALIDO"
    niv_input_invalid_only = (~niv_da_valid) & (~niv_input_valid) & _nonempty_mask(niv_input_raw) & ~_nonempty_mask(niv_da_raw)
    niv_da_invalid_only = (~niv_da_valid) & (~niv_input_valid) & (~_nonempty_mask(niv_input_raw)) & _nonempty_mask(niv_da_raw)
    niv_input_da_invalid = (~niv_da_valid) & (~niv_input_valid) & _nonempty_mask(niv_input_raw) & _nonempty_mask(niv_da_raw)
    niv_aca_audit.loc[niv_input_invalid_only] = "DEFAULT_1_INPUT_INVALIDO"
    niv_aca_audit.loc[niv_da_invalid_only] = "DEFAULT_1_DA_INVALIDO"
    niv_aca_audit.loc[niv_input_da_invalid] = "DEFAULT_1_INPUT_DA_INVALIDOS"

    fecha_mat_input_raw = src_work[col_fecha_matricula] if col_fecha_matricula else _na_series()
    fecha_mat_input_dt = pd.to_datetime(fecha_mat_input_raw, errors="coerce", dayfirst=True)
    fecha_mat_da_raw = src_work["DA_FECHAMATRICULA"] if "DA_FECHAMATRICULA" in src_work.columns else _na_series()
    fecha_mat_da_dt = pd.to_datetime(fecha_mat_da_raw, errors="coerce", dayfirst=True)
    fecha_mat_input_valid = fecha_mat_input_dt.notna()
    fecha_mat_da_valid = fecha_mat_da_dt.notna()
    fecha_mat_source = pd.Series("REGLA_FALLBACK_1900", index=src_work.index, dtype="object")
    fecha_mat_method = pd.Series("DEFAULT_1900", index=src_work.index, dtype="object")
    fecha_mat_audit = pd.Series("DEFAULT_1900_SIN_FUENTE", index=src_work.index, dtype="object")
    fecha_mat_source.loc[fecha_mat_input_valid] = fecha_mat_input_label
    fecha_mat_method.loc[fecha_mat_input_valid] = "SOURCE_INPUT"
    fecha_mat_audit.loc[fecha_mat_input_valid] = "SOURCE_INPUT_VALIDA"
    fecha_mat_da_mask = (~fecha_mat_input_valid) & fecha_mat_da_valid
    fecha_mat_source.loc[fecha_mat_da_mask] = "DA_FECHAMATRICULA"
    fecha_mat_method.loc[fecha_mat_da_mask] = "FALLBACK_DATOS_ALUMNOS"
    fecha_mat_audit.loc[fecha_mat_da_mask] = "FALLBACK_DA_VALIDA"
    fecha_input_invalid_only = (~fecha_mat_input_valid) & (~fecha_mat_da_valid) & _nonempty_mask(fecha_mat_input_raw) & ~_nonempty_mask(fecha_mat_da_raw)
    fecha_da_invalid_only = (~fecha_mat_input_valid) & (~fecha_mat_da_valid) & (~_nonempty_mask(fecha_mat_input_raw)) & _nonempty_mask(fecha_mat_da_raw)
    fecha_input_da_invalid = (~fecha_mat_input_valid) & (~fecha_mat_da_valid) & _nonempty_mask(fecha_mat_input_raw) & _nonempty_mask(fecha_mat_da_raw)
    fecha_mat_audit.loc[fecha_input_invalid_only] = "DEFAULT_1900_INPUT_INVALIDA"
    fecha_mat_audit.loc[fecha_da_invalid_only] = "DEFAULT_1900_DA_INVALIDA"
    fecha_mat_audit.loc[fecha_input_da_invalid] = "DEFAULT_1900_INPUT_DA_INVALIDAS"

    sit_fon_source = pd.Series("POLITICA_LOCAL_FIJA_1", index=src_work.index, dtype="object")
    sit_fon_method = pd.Series("CONSTANTE_1_GLOBAL", index=src_work.index, dtype="object")
    sit_fon_audit = pd.Series("FIJADO_MANUALMENTE_A_1_EN_TODO_EL_PROYECTO", index=src_work.index, dtype="object")

    sus_pre_source = pd.Series("POLITICA_LOCAL_FIJA_0", index=src_work.index, dtype="object")
    sus_pre_method = pd.Series("CONSTANTE_0_GLOBAL", index=src_work.index, dtype="object")
    sus_pre_audit = pd.Series("FIJADO_MANUALMENTE_A_0_EN_TODO_EL_PROYECTO", index=src_work.index, dtype="object")

    reinc_source = pd.Series("POLITICA_LOCAL_FIJA_0", index=src_work.index, dtype="object")
    reinc_method = pd.Series("CONSTANTE_0_GLOBAL", index=src_work.index, dtype="object")
    reinc_audit = pd.Series("FIJADO_MANUALMENTE_A_0_EN_TODO_EL_PROYECTO", index=src_work.index, dtype="object")

    vig_input_raw = src_work[col_vig] if col_vig else _na_series()
    vig_input = pd.to_numeric(vig_input_raw, errors="coerce")
    vig_input_valid = vig_input.isin([0, 1, 2])
    vig_source = pd.Series("REGLA_DEFAULT_1_SIN_FUENTE", index=src_work.index, dtype="object")
    vig_method = pd.Series("DEFAULT_1", index=src_work.index, dtype="object")
    vig_audit = pd.Series("DEFAULT_1_SIN_FUENTE", index=src_work.index, dtype="object")
    vig_source.loc[vig_input_valid] = vig_input_label
    vig_method.loc[vig_input_valid] = "SOURCE_INPUT"
    vig_audit.loc[vig_input_valid] = "SOURCE_INPUT_VALIDO"
    vig_input_invalid_only = (~vig_input_valid) & _nonempty_mask(vig_input_raw)
    vig_source.loc[vig_input_invalid_only] = vig_input_label
    vig_method.loc[vig_input_invalid_only] = "DEFAULT_1_POR_INPUT_INVALIDO"
    vig_audit.loc[vig_input_invalid_only] = "DEFAULT_1_INPUT_INVALIDO"

    if col_anio_ing:
        out["ANIO_ING_ACT"] = src_work[col_anio_ing].map(_to_int_year)
        if usar_gobernanza_v2 and "DA_ANOINGRESO" in src_work.columns:
            out["ANIO_ING_ACT"] = out["ANIO_ING_ACT"].combine_first(src_work["DA_ANOINGRESO"].map(_to_int_year))
        out["ANIO_ING_ACT"] = out["ANIO_ING_ACT"].combine_first(src_work[req_codcli].map(_infer_year_from_codcli))
    elif usar_gobernanza_v2 and "DA_ANOINGRESO" in src_work.columns:
        out["ANIO_ING_ACT"] = src_work["DA_ANOINGRESO"].map(_to_int_year)
    else:
        out["ANIO_ING_ACT"] = src_work[req_codcli].map(_infer_year_from_codcli)

    if col_sem_ing:
        out["SEM_ING_ACT"] = src_work[col_sem_ing]
        if usar_gobernanza_v2 and "DA_PERIODOINGRESO" in src_work.columns:
            out["SEM_ING_ACT"] = out["SEM_ING_ACT"].combine_first(src_work["DA_PERIODOINGRESO"])
    elif usar_gobernanza_v2 and "DA_PERIODOINGRESO" in src_work.columns:
        out["SEM_ING_ACT"] = src_work["DA_PERIODOINGRESO"]
    else:
        out["SEM_ING_ACT"] = _na_series()

    out["ANIO_ING_ORI"] = out["ANIO_ING_ACT"]
    out["SEM_ING_ORI"] = out["SEM_ING_ACT"]
    out["ASI_INS_ANT"] = src_work[col_asi_ins_ant] if col_asi_ins_ant else _na_series()
    if "ASI_INS_ANT_HIST" in src_work.columns:
        out["ASI_INS_ANT"] = out["ASI_INS_ANT"].combine_first(src_work["ASI_INS_ANT_HIST"])
    out["ASI_APR_ANT"] = src_work[col_asi_apr_ant] if col_asi_apr_ant else _na_series()
    if "ASI_APR_ANT_HIST" in src_work.columns:
        out["ASI_APR_ANT"] = out["ASI_APR_ANT"].combine_first(src_work["ASI_APR_ANT_HIST"])
    out["PROM_PRI_SEM"] = src_work[col_prom_pri_sem] if col_prom_pri_sem else _na_series()
    if "PROM_PRI_SEM_HIST" in src_work.columns:
        out["PROM_PRI_SEM"] = out["PROM_PRI_SEM"].combine_first(src_work["PROM_PRI_SEM_HIST"])
    out["PROM_SEG_SEM"] = src_work[col_prom_seg_sem] if col_prom_seg_sem else _na_series()
    if "PROM_SEG_SEM_HIST" in src_work.columns:
        out["PROM_SEG_SEM"] = out["PROM_SEG_SEM"].combine_first(src_work["PROM_SEG_SEM_HIST"])
    out["ASI_INS_HIS"] = src_work[col_asi_ins_his] if col_asi_ins_his else _na_series()
    if "ASI_INS_HIS_HIST" in src_work.columns:
        out["ASI_INS_HIS"] = out["ASI_INS_HIS"].combine_first(src_work["ASI_INS_HIS_HIST"])
    out["ASI_APR_HIS"] = src_work[col_asi_apr_his] if col_asi_apr_his else _na_series()
    if "ASI_APR_HIS_HIST" in src_work.columns:
        out["ASI_APR_HIS"] = out["ASI_APR_HIS"].combine_first(src_work["ASI_APR_HIS_HIST"])
    # NIV_ACA: priorizar DA_NIVEL (nivel alumno) sobre Hoja1.NIVEL (nivel del ramo en malla)
    if usar_gobernanza_v2 and "DA_NIVEL" in src_work.columns:
        out["NIV_ACA"] = pd.to_numeric(src_work["DA_NIVEL"], errors="coerce")
        if col_niv_aca:
            out["NIV_ACA"] = out["NIV_ACA"].combine_first(pd.to_numeric(src_work[col_niv_aca], errors="coerce"))
    else:
        out["NIV_ACA"] = src_work[col_niv_aca] if col_niv_aca else _na_series()
    out["SIT_FON_SOL"] = 1
    out["SUS_PRE"] = 0

    if col_fecha_matricula:
        out["FECHA_MATRICULA"] = src_work[col_fecha_matricula]
    elif usar_gobernanza_v2 and "DA_FECHAMATRICULA" in src_work.columns:
        out["FECHA_MATRICULA"] = src_work["DA_FECHAMATRICULA"]
    else:
        out["FECHA_MATRICULA"] = _na_series()

    out["REINCORPORACION"] = 0

    out["VIG"] = src_work[col_vig] if col_vig else 1
    out = out[MATRICULA_UNIFICADA_COLUMNS].copy()

    def _build_fase4_trace(
        input_col: str | None,
        hist_col: str,
        input_method: str,
        hist_method: str,
    ) -> tuple[pd.Series, pd.Series, pd.Series]:
        input_values = src_work[input_col] if input_col and input_col in src_work.columns else _na_series()
        hist_values = src_work[hist_col] if hist_col in src_work.columns else _na_series()
        input_mask = input_values.notna()
        hist_mask = (~input_mask) & hist_values.notna()

        source = pd.Series("SIN_FUENTE_FINAL", index=src_work.index, dtype="object")
        method = pd.Series("SIN_REGLA_FINAL", index=src_work.index, dtype="object")
        audit = pd.Series("SIN_FUENTE_FINAL", index=src_work.index, dtype="object")

        if input_col:
            source.loc[input_mask] = f"INPUT_{input_col}"
            method.loc[input_mask] = input_method
            audit.loc[input_mask] = "FUENTE_INPUT_DIRECTA"

        if "UZ_FUENTE_HIST" in src_work.columns:
            source.loc[hist_mask] = src_work.loc[hist_mask, "UZ_FUENTE_HIST"].fillna("HISTORICO_HOJA1")
        else:
            source.loc[hist_mask] = "HISTORICO_HOJA1"
        method.loc[hist_mask] = hist_method
        audit.loc[hist_mask] = "CALCULADO_DESDE_HISTORICO_HOJA1"
        return source, method, audit

    asi_ins_ant_source, asi_ins_ant_method, asi_ins_ant_audit = _build_fase4_trace(
        col_asi_ins_ant,
        "ASI_INS_ANT_HIST",
        "INPUT_DIRECTO_ASI_INS_ANT",
        "COUNT_DISTINCT_CODRAMO_ANIO_REFERENCIA",
    )
    asi_apr_ant_source, asi_apr_ant_method, asi_apr_ant_audit = _build_fase4_trace(
        col_asi_apr_ant,
        "ASI_APR_ANT_HIST",
        "INPUT_DIRECTO_ASI_APR_ANT",
        "COUNT_DISTINCT_CODRAMO_APROB_ANIO_REFERENCIA_EXCL_EQUIV",
    )
    prom_pri_source, prom_pri_method, prom_pri_audit = _build_fase4_trace(
        col_prom_pri_sem,
        "PROM_PRI_SEM_HIST",
        "INPUT_DIRECTO_PROM_PRI_SEM",
        "AVG_NOTA_MU_SEM1_ANIO_REFERENCIA_EXCL_EQUIV",
    )
    prom_seg_source, prom_seg_method, prom_seg_audit = _build_fase4_trace(
        col_prom_seg_sem,
        "PROM_SEG_SEM_HIST",
        "INPUT_DIRECTO_PROM_SEG_SEM",
        "AVG_NOTA_MU_SEM2_ANIO_REFERENCIA_EXCL_EQUIV",
    )
    asi_ins_his_source, asi_ins_his_method, asi_ins_his_audit = _build_fase4_trace(
        col_asi_ins_his,
        "ASI_INS_HIS_HIST",
        "INPUT_DIRECTO_ASI_INS_HIS",
        "COUNT_DISTINCT_CODRAMO_ALCANCE_HIST_DISPONIBLE",
    )
    asi_apr_his_source, asi_apr_his_method, asi_apr_his_audit = _build_fase4_trace(
        col_asi_apr_his,
        "ASI_APR_HIS_HIST",
        "INPUT_DIRECTO_ASI_APR_HIS",
        "COUNT_DISTINCT_CODRAMO_APROB_O_EQUIV_ALCANCE_HIST_DISPONIBLE",
    )

    if "REINCORPORACION" not in src_work.columns and "DA_SITUACION" in src_work.columns:
        out["REINCORPORACION"] = 0

    rut_norm = out["N_DOC"].astype(str).str.replace(r"\D", "", regex=True) + out["DV"].astype(str).str.strip().str.upper()
    duplicated_vig = (
        rut_norm.map(rut_norm[out["VIG"].fillna(1).astype(str).isin(["1", "1.0"])].value_counts())
        .fillna(0)
        .astype(int)
        > 1
    )
    estado_inicial = out["VIG"].map(_status_from_vig)
    estado_inicial = estado_inicial.where(~duplicated_vig, "Matrícula Duplicada")
    estado_final = estado_inicial.copy()

    archivo_subida = out.copy()
    archivo_subida["CODCLI"] = src_work[req_codcli]
    archivo_subida["PLAN_DE_ESTUDIO"] = src_work[col_plan] if col_plan else pd.NA
    archivo_subida["PERIODO"] = src_work[col_periodo] if col_periodo else pd.NA
    archivo_subida["NOMBRE_CARRERA_FUENTE"] = src_work[req_nombre_carrera]
    archivo_subida["JORNADA_FUENTE"] = src_work[req_jornada]
    archivo_subida["ESTADO_INICIAL_REGISTRO"] = estado_inicial
    archivo_subida["RESOLUCION_DUPLICADO"] = "Mantener"
    archivo_subida["ACTIVAR_DESACTIVAR"] = "Registro Activo"
    archivo_subida["ESTADO_FINAL_REGISTRO"] = estado_final
    archivo_subida["SOURCE_KEY_3"] = (
        src_work[req_jornada].map(_normalize_text)
        + "|"
        + src_work[req_codcarr].map(_normalize_text)
        + "|"
        + src_work[req_nombre_carrera].map(_normalize_text)
    )
    archivo_subida["KEY_3_NO_JORNADA"] = "|" + src_work[req_codcarr].map(_normalize_text) + "|" + src_work[req_nombre_carrera].map(_normalize_text)
    archivo_subida["CODCARPR_NORM"] = src_work[req_codcarr].map(_normalize_text)
    archivo_subida["ES_DIPLOMADO"] = src_work[req_nombre_carrera].map(_is_diplomado_name)
    archivo_subida["MATCH_KEY_3"] = archivo_subida["SOURCE_KEY_3"]
    archivo_subida["FLAG_GOBERNANZA_V2"] = "SI" if usar_gobernanza_v2 else "NO"
    archivo_subida["DA_MATCH_MODO"] = da_match_modo
    archivo_subida["TIPO_DOC_STATUS"] = tipo_doc_status
    archivo_subida["N_DOC_STATUS"] = n_doc_status
    archivo_subida["DV_STATUS"] = dv_status
    archivo_subida["SEGUNDO_APELLIDO_STATUS"] = segundo_apellido_status
    archivo_subida["FECH_NAC_STATUS"] = fech_nac_status
    archivo_subida["DA_VIASDEADMISION"] = src_work["DA_VIASDEADMISION"] if "DA_VIASDEADMISION" in src_work.columns else pd.NA
    archivo_subida["FOR_ING_ACT_FUENTE_VALOR"] = src_work["FOR_ING_ACT_FUENTE_VALOR"] if "FOR_ING_ACT_FUENTE_VALOR" in src_work.columns else pd.NA
    archivo_subida["FOR_ING_ACT_FUENTE_CAMPO"] = src_work["FOR_ING_ACT_FUENTE_CAMPO"] if "FOR_ING_ACT_FUENTE_CAMPO" in src_work.columns else pd.NA
    archivo_subida["FOR_ING_ACT_FUENTE_NORM"] = src_work["FOR_ING_ACT_FUENTE_NORM"] if "FOR_ING_ACT_FUENTE_NORM" in src_work.columns else pd.NA
    archivo_subida["FOR_ING_ACT_METODO"] = src_work["FOR_ING_ACT_METODO"] if "FOR_ING_ACT_METODO" in src_work.columns else pd.NA
    archivo_subida["FOR_ING_ACT_IMPUTADO"] = src_work["FOR_ING_ACT_IMPUTADO"] if "FOR_ING_ACT_IMPUTADO" in src_work.columns else pd.NA
    archivo_subida["FOR_ING_ACT_REQUIERE_REVISION"] = src_work["FOR_ING_ACT_REQUIERE_REVISION"] if "FOR_ING_ACT_REQUIERE_REVISION" in src_work.columns else pd.NA
    archivo_subida["DA_ANOMATRICULA"] = src_work["DA_ANOMATRICULA"] if "DA_ANOMATRICULA" in src_work.columns else pd.NA
    archivo_subida["DA_PERIODOMATRICULA"] = src_work["DA_PERIODOMATRICULA"] if "DA_PERIODOMATRICULA" in src_work.columns else pd.NA
    archivo_subida["DA_ESTADOACADEMICO"] = src_work["DA_ESTADOACADEMICO"] if "DA_ESTADOACADEMICO" in src_work.columns else pd.NA
    archivo_subida["DA_SITUACION"] = src_work["DA_SITUACION"] if "DA_SITUACION" in src_work.columns else pd.NA
    archivo_subida["DA_MATRICULA"] = src_work["DA_MATRICULA"] if "DA_MATRICULA" in src_work.columns else pd.NA
    archivo_subida["DA_CON_FIRMA"] = src_work["DA_CON_FIRMA"] if "DA_CON_FIRMA" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_KEY"] = src_work["UZ_HIST_KEY"] if "UZ_HIST_KEY" in src_work.columns else pd.NA
    archivo_subida["ANIO_REFERENCIA_HIST_UZ"] = src_work["ANIO_REFERENCIA_HIST_UZ"] if "ANIO_REFERENCIA_HIST_UZ" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_ANIO_MIN"] = src_work["UZ_HIST_ANIO_MIN"] if "UZ_HIST_ANIO_MIN" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_ANIO_MAX"] = src_work["UZ_HIST_ANIO_MAX"] if "UZ_HIST_ANIO_MAX" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_ANIOS_DISPONIBLES"] = src_work["UZ_HIST_ANIOS_DISPONIBLES"] if "UZ_HIST_ANIOS_DISPONIBLES" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_SCOPE_STATUS"] = src_work["UZ_HIST_SCOPE_STATUS"] if "UZ_HIST_SCOPE_STATUS" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_TOTAL"] = src_work["UZ_HIST_FILAS_TOTAL"] if "UZ_HIST_FILAS_TOTAL" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_ANIO_REFERENCIA"] = src_work["UZ_HIST_FILAS_ANIO_REFERENCIA"] if "UZ_HIST_FILAS_ANIO_REFERENCIA" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_REF_APROB"] = src_work["UZ_HIST_FILAS_REF_APROB"] if "UZ_HIST_FILAS_REF_APROB" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_REF_REPROB"] = src_work["UZ_HIST_FILAS_REF_REPROB"] if "UZ_HIST_FILAS_REF_REPROB" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_REF_TRANSFER"] = src_work["UZ_HIST_FILAS_REF_TRANSFER"] if "UZ_HIST_FILAS_REF_TRANSFER" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_REF_SEM1_CALIFICADAS"] = src_work["UZ_HIST_FILAS_REF_SEM1_CALIFICADAS"] if "UZ_HIST_FILAS_REF_SEM1_CALIFICADAS" in src_work.columns else pd.NA
    archivo_subida["UZ_HIST_FILAS_REF_SEM2_CALIFICADAS"] = src_work["UZ_HIST_FILAS_REF_SEM2_CALIFICADAS"] if "UZ_HIST_FILAS_REF_SEM2_CALIFICADAS" in src_work.columns else pd.NA
    archivo_subida["UZ_FUENTE_HIST"] = src_work["UZ_FUENTE_HIST"] if "UZ_FUENTE_HIST" in src_work.columns else pd.NA
    archivo_subida["NAC_STATUS"] = nac_status
    archivo_subida["PAIS_EST_SEC_STATUS"] = pais_est_sec_status
    archivo_subida["COD_SED_STATUS"] = cod_sed_status
    archivo_subida["ANIO_ING_ACT_FUENTE_FINAL"] = anio_act_source
    archivo_subida["ANIO_ING_ACT_METODO_FINAL"] = anio_act_method
    archivo_subida["ANIO_ING_ACT_AUDIT_STATUS"] = anio_act_audit
    archivo_subida["SEM_ING_ACT_FUENTE_FINAL"] = sem_act_source
    archivo_subida["SEM_ING_ACT_METODO_FINAL"] = sem_act_method
    archivo_subida["SEM_ING_ACT_AUDIT_STATUS"] = sem_act_audit
    archivo_subida["NIV_ACA_FUENTE_FINAL"] = niv_aca_source
    archivo_subida["NIV_ACA_METODO_FINAL"] = niv_aca_method
    archivo_subida["NIV_ACA_AUDIT_STATUS"] = niv_aca_audit
    archivo_subida["FECHA_MATRICULA_FUENTE_FINAL"] = fecha_mat_source
    archivo_subida["FECHA_MATRICULA_METODO_FINAL"] = fecha_mat_method
    archivo_subida["FECHA_MATRICULA_AUDIT_STATUS"] = fecha_mat_audit
    archivo_subida["ASI_INS_ANT_FUENTE_FINAL"] = asi_ins_ant_source
    archivo_subida["ASI_INS_ANT_METODO_FINAL"] = asi_ins_ant_method
    archivo_subida["ASI_INS_ANT_AUDIT_STATUS"] = asi_ins_ant_audit
    archivo_subida["ASI_APR_ANT_FUENTE_FINAL"] = asi_apr_ant_source
    archivo_subida["ASI_APR_ANT_METODO_FINAL"] = asi_apr_ant_method
    archivo_subida["ASI_APR_ANT_AUDIT_STATUS"] = asi_apr_ant_audit
    archivo_subida["PROM_PRI_SEM_FUENTE_FINAL"] = prom_pri_source
    archivo_subida["PROM_PRI_SEM_METODO_FINAL"] = prom_pri_method
    archivo_subida["PROM_PRI_SEM_AUDIT_STATUS"] = prom_pri_audit
    archivo_subida["PROM_SEG_SEM_FUENTE_FINAL"] = prom_seg_source
    archivo_subida["PROM_SEG_SEM_METODO_FINAL"] = prom_seg_method
    archivo_subida["PROM_SEG_SEM_AUDIT_STATUS"] = prom_seg_audit
    archivo_subida["ASI_INS_HIS_FUENTE_FINAL"] = asi_ins_his_source
    archivo_subida["ASI_INS_HIS_METODO_FINAL"] = asi_ins_his_method
    archivo_subida["ASI_INS_HIS_AUDIT_STATUS"] = asi_ins_his_audit
    archivo_subida["ASI_APR_HIS_FUENTE_FINAL"] = asi_apr_his_source
    archivo_subida["ASI_APR_HIS_METODO_FINAL"] = asi_apr_his_method
    archivo_subida["ASI_APR_HIS_AUDIT_STATUS"] = asi_apr_his_audit
    archivo_subida["SIT_FON_SOL_FUENTE_FINAL"] = sit_fon_source
    archivo_subida["SIT_FON_SOL_METODO_FINAL"] = sit_fon_method
    archivo_subida["SIT_FON_SOL_AUDIT_STATUS"] = sit_fon_audit
    archivo_subida["SUS_PRE_FUENTE_FINAL"] = sus_pre_source
    archivo_subida["SUS_PRE_METODO_FINAL"] = sus_pre_method
    archivo_subida["SUS_PRE_AUDIT_STATUS"] = sus_pre_audit
    archivo_subida["REINCORPORACION_FUENTE_FINAL"] = reinc_source
    archivo_subida["REINCORPORACION_METODO_FINAL"] = reinc_method
    archivo_subida["REINCORPORACION_AUDIT_STATUS"] = reinc_audit
    archivo_subida["VIG_FUENTE_FINAL"] = vig_source
    archivo_subida["VIG_METODO_FINAL"] = vig_method
    archivo_subida["VIG_AUDIT_STATUS"] = vig_audit

    # Trazabilidad de consistencia VIG vs estados académicos institucionales.
    estado_da_norm = archivo_subida["DA_ESTADOACADEMICO"].fillna("").map(_normalize_text)
    situ_da_norm = archivo_subida["DA_SITUACION"].fillna("").map(_normalize_text)
    vig_esperado_da = pd.Series(pd.NA, index=archivo_subida.index, dtype="Int64")

    if not gob_da_estado_situ_df.empty:
        da_map_df = gob_da_estado_situ_df.copy()
        da_map_df["ESTADOACADEMICO_NORM"] = da_map_df["ESTADOACADEMICO"].map(_normalize_text)
        da_map_df["SITUACION_NORM"] = da_map_df["SITUACION"].map(_normalize_text)
        da_map_df["KEY_DA"] = da_map_df["ESTADOACADEMICO_NORM"] + "|" + da_map_df["SITUACION_NORM"]
        da_vig_map = (
            da_map_df.drop_duplicates(subset=["KEY_DA"], keep="first")
            .set_index("KEY_DA")["VIG_ESPERADO"]
            .to_dict()
        )
        key_da = estado_da_norm + "|" + situ_da_norm
        vig_esperado_da = pd.to_numeric(key_da.map(da_vig_map), errors="coerce").astype("Int64")

    # Fallback Hoja1 cuando no hay match completo en DatosAlumnos.
    if not gob_hoja1_estado_desc_df.empty and "DESCRIPCION_ESTADO" in src_work.columns:
        estado_h1_norm = src_work.get("ESTADO_ACADEMICO", pd.Series("", index=src_work.index)).fillna("").map(_normalize_text)
        desc_h1_norm = src_work["DESCRIPCION_ESTADO"].fillna("").map(_normalize_text)
        h1_map_df = gob_hoja1_estado_desc_df.copy()
        h1_map_df["ESTADO_ACADEMICO_NORM"] = h1_map_df["ESTADO_ACADEMICO"].map(_normalize_text)
        h1_map_df["DESCRIPCION_ESTADO_NORM"] = h1_map_df["DESCRIPCION_ESTADO"].fillna("").map(_normalize_text)
        h1_map_df["KEY_H1"] = h1_map_df["ESTADO_ACADEMICO_NORM"] + "|" + h1_map_df["DESCRIPCION_ESTADO_NORM"]
        h1_vig_map = (
            h1_map_df.drop_duplicates(subset=["KEY_H1"], keep="first")
            .set_index("KEY_H1")["VIG_ESPERADO"]
            .to_dict()
        )
        key_h1 = estado_h1_norm + "|" + desc_h1_norm
        vig_esperado_h1 = pd.to_numeric(key_h1.map(h1_vig_map), errors="coerce").astype("Int64")
        vig_esperado_da = vig_esperado_da.fillna(vig_esperado_h1)

    # Regla institucional explícita: estados sin matrícula siempre VIG=0.
    force_vig0_da = estado_da_norm.isin(["TITULADO", "ELIMINADO", "SUSPENDIDO"])
    vig_esperado_da = vig_esperado_da.where(~force_vig0_da, 0)
    archivo_subida["VIG_ESPERADO_DA"] = vig_esperado_da

    # Fuente base manual: se reconstruye desde DURACION_ESTUDIOS para trazabilidad
    # de GRUPO_TRAZA/FAMILIA. El cruce SIES central se consume EXCLUSIVAMENTE
    # desde el catálogo compilado control/catalogos/PUENTE_SIES_COMPILADO.tsv.
    df_manual, _ = _build_catalog_and_bridge_from_duracion()

    manual_override_raw = _load_governance_tsv(
        catalogo_manual_tsv_path,
        ["GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L"],
    )
    if not manual_override_raw.empty:
        try:
            manual_override = _prepare_catalog_manual(
                manual_override_raw[["GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L"]]
            )
        except Exception as exc:
            print(f"⚠️  No se aplicó override catalogo_manual.tsv: {exc}")
            manual_override = pd.DataFrame()
        if not manual_override.empty:
            if df_manual.empty:
                df_manual = manual_override.copy()
            else:
                override_keys = set(manual_override["MANUAL_KEY_3"])
                df_manual = (
                    pd.concat(
                        [
                            df_manual.loc[~df_manual["MANUAL_KEY_3"].isin(override_keys)],
                            manual_override,
                        ],
                        ignore_index=True,
                    )
                    .drop_duplicates(subset=["MANUAL_KEY_3"], keep="last")
                    .sort_values(["GRUPO_TRAZA", "JORNADA", "CODCARPR", "NOMBRE_L"])
                    .reset_index(drop=True)
                )
            print(
                "✅ Override catálogo manual aplicado: "
                f"{len(manual_override)} llaves desde {catalogo_manual_tsv_path}"
            )

    compiled_required = [
        "BRIDGE_KEY_3",
        "BRIDGE_KEY_NO_JORNADA",
        "GRUPO_TRAZA",
        "FAMILIA_TRAZA",
        "FAMILIA_CODCARPR",
        "JORNADA",
        "CODCARPR",
        "NOMBRE_L",
        "N_CODES_SIES",
        "CODIGOS_SIES_POTENCIALES",
        "CODIGO_CARRERA_SIES_1",
    ]
    df_bridge = _load_governance_tsv(str(puente_compilado_path), compiled_required)
    if df_bridge.empty:
        raise FileNotFoundError(
            "No se encontró catálogo compilado de cruce SIES en "
            f"{puente_compilado_path}. Ejecuta scripts/compile_puente_sies_compilado.py antes del run oficial."
        )
    for idx in range(1, MAX_SIES_CODES_PER_KEY + 1):
        col = f"CODIGO_CARRERA_SIES_{idx}"
        if col not in df_bridge.columns:
            df_bridge[col] = pd.NA
    if "CODIGO_CARRERA" not in df_bridge.columns:
        df_bridge["CODIGO_CARRERA"] = df_bridge["CODIGO_CARRERA_SIES_1"].apply(
            lambda x: int(m.group("cod_car")) if pd.notna(x) and (m := _SIES_CODE_RE.match(str(x).strip())) else pd.NA
        )
    df_bridge["N_CODES_SIES"] = pd.to_numeric(df_bridge["N_CODES_SIES"], errors="coerce").fillna(0).astype(int)
    df_bridge = df_bridge.drop_duplicates(subset=["BRIDGE_KEY_3"], keep="first").reset_index(drop=True)

    if not df_manual.empty:
        manual_exact = (
            df_manual[["MANUAL_KEY_3", "GRUPO_TRAZA", "FAMILIA_TRAZA", "FAMILIA_CODCARPR"]]
            .drop_duplicates(subset=["MANUAL_KEY_3"])
            .rename(
                columns={
                    "MANUAL_KEY_3": "SOURCE_KEY_3",
                    "GRUPO_TRAZA": "GRUPO_TRAZA_MANUAL",
                    "FAMILIA_TRAZA": "FAMILIA_TRAZA_MANUAL",
                    "FAMILIA_CODCARPR": "FAMILIA_CODCARPR_MANUAL",
                }
            )
        )
        archivo_subida = archivo_subida.merge(manual_exact, on="SOURCE_KEY_3", how="left")
        manual_key_set = set(df_manual["MANUAL_KEY_3"])
        archivo_subida["MANUAL_MATCH_STATUS"] = archivo_subida["SOURCE_KEY_3"].isin(manual_key_set).map(
            {True: "MATCH_MANUAL", False: "SIN_MATCH_MANUAL"}
        )
    else:
        archivo_subida["GRUPO_TRAZA_MANUAL"] = pd.NA
        archivo_subida["FAMILIA_TRAZA_MANUAL"] = pd.NA
        archivo_subida["FAMILIA_CODCARPR_MANUAL"] = pd.NA
        archivo_subida["MANUAL_MATCH_STATUS"] = "SIN_CATALOGO_MANUAL"

    for idx in range(1, MAX_SIES_CODES_PER_KEY + 1):
        archivo_subida[f"CODIGO_CARRERA_SIES_{idx}"] = pd.NA
    archivo_subida["N_CODES_SIES"] = pd.NA
    archivo_subida["CODIGOS_SIES_POTENCIALES"] = pd.NA
    archivo_subida[FINAL_SIES_CODE_COL] = pd.NA
    archivo_subida["SIES_RESOLUCION_HEURISTICA"] = pd.NA
    archivo_subida["SIES_CONFIANZA_POST"] = pd.NA

    if not df_bridge.empty:
        sies_cols = [
            "N_CODES_SIES",
            "CODIGOS_SIES_POTENCIALES",
        ] + [f"CODIGO_CARRERA_SIES_{idx}" for idx in range(1, MAX_SIES_CODES_PER_KEY + 1)]
        archivo_subida = archivo_subida.drop(columns=[c for c in sies_cols if c in archivo_subida.columns], errors="ignore")

        bridge_join_cols = [
            "BRIDGE_KEY_3",
            "GRUPO_TRAZA",
            "FAMILIA_TRAZA",
            "FAMILIA_CODCARPR",
            "N_CODES_SIES",
            "CODIGOS_SIES_POTENCIALES",
        ] + [f"CODIGO_CARRERA_SIES_{idx}" for idx in range(1, MAX_SIES_CODES_PER_KEY + 1)]
        bridge_exact = df_bridge[bridge_join_cols].rename(
            columns={
                "BRIDGE_KEY_3": "SOURCE_KEY_3",
                "GRUPO_TRAZA": "GRUPO_TRAZA_PUENTE",
                "FAMILIA_TRAZA": "FAMILIA_TRAZA_PUENTE",
                "FAMILIA_CODCARPR": "FAMILIA_CODCARPR_PUENTE",
            }
        )
        archivo_subida = archivo_subida.merge(bridge_exact, on="SOURCE_KEY_3", how="left")

        key_exact = set(df_bridge["BRIDGE_KEY_3"])
        key_no_jornada = set(df_bridge["BRIDGE_KEY_NO_JORNADA"])
        codcarpr_bridge = set(df_bridge["CODCARPR"])

        match_exact = archivo_subida["SOURCE_KEY_3"].isin(key_exact)
        exists_no_j = archivo_subida["KEY_3_NO_JORNADA"].isin(key_no_jornada)
        exists_cod = archivo_subida["CODCARPR_NORM"].isin(codcarpr_bridge)
        n_codes = pd.to_numeric(archivo_subida["N_CODES_SIES"], errors="coerce").fillna(0)

        archivo_subida["SIES_MATCH_STATUS"] = "SIN_MATCH_SIES"
        archivo_subida["SIES_MATCH_DIAG"] = "SIN_CODCARPR_EN_PUENTE_SIES"

        unique_mask = match_exact & (n_codes == 1)
        amb_mask = match_exact & (n_codes > 1)

        archivo_subida.loc[unique_mask, "SIES_MATCH_STATUS"] = "MATCH_SIES"
        archivo_subida.loc[unique_mask, "SIES_MATCH_DIAG"] = "MATCH_SIES_UNICO"
        archivo_subida.loc[unique_mask, FINAL_SIES_CODE_COL] = archivo_subida.loc[unique_mask, "CODIGO_CARRERA_SIES_1"]

        archivo_subida.loc[amb_mask, "SIES_MATCH_STATUS"] = "AMBIGUO_SIES"
        archivo_subida.loc[amb_mask, "SIES_MATCH_DIAG"] = "MATCH_SIES_AMBIGUO"

        archivo_subida.loc[(~match_exact) & exists_no_j, "SIES_MATCH_DIAG"] = "PROBABLE_PROBLEMA_JORNADA_SIES"
        archivo_subida.loc[(~match_exact) & (~exists_no_j) & exists_cod, "SIES_MATCH_DIAG"] = "PROBABLE_PROBLEMA_NOMBRE_SIES"
    else:
        archivo_subida["GRUPO_TRAZA_PUENTE"] = pd.NA
        archivo_subida["FAMILIA_TRAZA_PUENTE"] = pd.NA
        archivo_subida["FAMILIA_CODCARPR_PUENTE"] = pd.NA
        archivo_subida["SIES_MATCH_STATUS"] = "SIN_PUENTE_SIES"
        archivo_subida["SIES_MATCH_DIAG"] = "SIN_PUENTE_SIES"

    archivo_subida["GRUPO_TRAZA"] = archivo_subida["GRUPO_TRAZA_PUENTE"].combine_first(archivo_subida["GRUPO_TRAZA_MANUAL"])
    archivo_subida["FAMILIA_TRAZA"] = archivo_subida["GRUPO_TRAZA"].map(_extract_alpha_prefix)
    archivo_subida["FAMILIA_CODCARPR"] = archivo_subida["CODCARPR_NORM"].map(_extract_alpha_prefix)

    if excluir_diplomados:
        excl = archivo_subida["ES_DIPLOMADO"].fillna(False)
        archivo_subida.loc[excl, "SIES_MATCH_STATUS"] = "EXCLUIDO_DIPLOMADO"
        archivo_subida.loc[excl, "SIES_MATCH_DIAG"] = "EXCLUIDO_DIPLOMADO"
        archivo_subida.loc[excl, FINAL_SIES_CODE_COL] = pd.NA
        archivo_subida.loc[excl, "CODIGOS_SIES_POTENCIALES"] = pd.NA
        archivo_subida.loc[excl, "N_CODES_SIES"] = pd.NA
        for idx in range(1, MAX_SIES_CODES_PER_KEY + 1):
            archivo_subida.loc[excl, f"CODIGO_CARRERA_SIES_{idx}"] = pd.NA

    # Regla de gobernanza bloqueante: combinaciones SOURCE_KEY_3 no catalogadas en SIES.
    sin_match_bloqueante = archivo_subida["SIES_MATCH_STATUS"].eq("SIN_MATCH_SIES") & ~archivo_subida["ES_DIPLOMADO"].fillna(False)
    if sin_match_bloqueante.any():
        pendientes = (
            archivo_subida.loc[
                sin_match_bloqueante,
                ["SOURCE_KEY_3", "CODCARPR_NORM", "NOMBRE_CARRERA_FUENTE", "JORNADA_FUENTE", "SIES_MATCH_DIAG"],
            ]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        pendientes_path = output_dir / "sies_combinaciones_nuevas_bloqueantes.tsv"
        pendientes.to_csv(pendientes_path, sep="\t", index=False, encoding="utf-8")
        muestra = pendientes.head(5).to_dict(orient="records")
        raise RuntimeError(
            "BLOQUEANTE_SIES: se detectaron combinaciones SOURCE_KEY_3 no catalogadas "
            f"({len(pendientes)}). Revisa {pendientes_path} y actualiza catálogo."
            f" Muestra: {muestra}"
        )

    resumen = pd.DataFrame(
        [
            {"metrica": "filas_fuente", "valor": len(src)},
            {"metrica": "filas_archivo_subida", "valor": len(archivo_subida)},
            {"metrica": "usar_gobernanza_v2", "valor": int(usar_gobernanza_v2)},
            {"metrica": "filas_enriquecidas_datos_alumnos", "valor": rows_enriquecidas_datos_alumnos},
            {"metrica": "filas_sin_match_datos_alumnos", "valor": sin_match_datos_alumnos_rows},
            {"metrica": "cod_sed_resuelto_por_regla", "valor": cod_sed_resueltos_regla},
            {"metrica": "pais_est_sec_inferido_localidad", "valor": pais_est_sec_inferidos_localidad},
            {"metrica": "rut_duplicados_vigentes", "valor": int(duplicated_vig.sum())},
            {"metrica": "matricula_ok_inicial", "valor": int((estado_inicial == "Matrícula OK").sum())},
            {"metrica": "matricula_duplicada_inicial", "valor": int((estado_inicial == "Matrícula Duplicada").sum())},
            {"metrica": "matricula_no_utilizada_inicial", "valor": int((estado_inicial == "Matrícula No Utilizada").sum())},
            {"metrica": "catalogo_manual_rows", "valor": len(df_manual)},
            {"metrica": "puente_sies_rows", "valor": len(df_bridge)},
            {"metrica": "excluir_diplomados", "valor": int(excluir_diplomados)},
        ]
    )
    resumen_manual = (
        archivo_subida["MANUAL_MATCH_STATUS"].fillna("<NA>").value_counts(dropna=False).rename_axis("estado").reset_index(name="n")
    )
    resumen_sies = (
        archivo_subida["SIES_MATCH_DIAG"].fillna("<NA>").value_counts(dropna=False).rename_axis("estado").reset_index(name="n")
    )
    ambiguos_pre = archivo_subida[archivo_subida["SIES_MATCH_STATUS"] == "AMBIGUO_SIES"].copy()
    
    # FASE 3: Resolver ambigüedades SIES con heurística
    if not ambiguos_pre.empty:
        print(f"📋 Fase 3: Resolviendo {len(ambiguos_pre)} ambigüedades SIES...")
        # Construir índice de oferta y homologación para la cascada
        oferta_idx = _build_oferta_index(oferta_dim)
        homol_dict = _load_cuadro_homologacion(input_file)
        ambiguos_resueltos = _resolver_ambiguedades_sies_heuristica(ambiguos_pre, oferta_idx, homol_dict)
        # Actualizar archivo_subida con los ambiguos resueltos usando loc por índice
        for col in ["SIES_RESOLUCION_HEURISTICA", "SIES_CONFIANZA_POST", FINAL_SIES_CODE_COL, "SIES_MATCH_STATUS"]:
            if col in ambiguos_resueltos.columns:
                archivo_subida.loc[ambiguos_resueltos.index, col] = ambiguos_resueltos[col]
        ambiguos = archivo_subida[archivo_subida["SIES_MATCH_STATUS"] == "AMBIGUO_SIES"].copy()
    else:
        ambiguos = ambiguos_pre.copy()
    
    sin_match = archivo_subida[archivo_subida["SIES_MATCH_STATUS"] == "SIN_MATCH_SIES"].copy()

    sin_match_datos_alumnos_df = pd.DataFrame(
        columns=[
            "CODCLI",
            "N_DOC",
            "DV",
            "COD_CAR_FUENTE",
            "NOMBRE_CARRERA_FUENTE",
            "JORNADA_FUENTE",
            "MOTIVO_REVISION",
        ]
    )
    if usar_gobernanza_v2:
        mask_no_match_da = da_match_modo == "SIN_MATCH"
        if mask_no_match_da.any():
            sin_match_datos_alumnos_df = pd.DataFrame(
                {
                    "CODCLI": src_work[req_codcli],
                    "N_DOC": out["N_DOC"],
                    "DV": out["DV"],
                    "COD_CAR_FUENTE": src_work[req_codcarr],
                    "NOMBRE_CARRERA_FUENTE": src_work[req_nombre_carrera],
                    "JORNADA_FUENTE": src_work[req_jornada],
                    "MOTIVO_REVISION": "SIN_MATCH_POR_CODCLI_Y_RUT_EN_DATOS_ALUMNOS",
                }
            ).loc[mask_no_match_da].reset_index(drop=True)

    # Normalización final contra reglas del manual de carga pregrado.
    # Se conserva ARCHIVO_LISTO_SUBIDA completo, y se construye una hoja
    # MATRICULA_UNIFICADA_32 lista para carga (sin diplomados, sin duplicados).
    parsed_sies = archivo_subida[FINAL_SIES_CODE_COL].apply(_extract_sies_components)
    parsed_df = pd.DataFrame(
        {
            "_PARSED_COD_SED": parsed_sies.map(lambda x: x[0] if isinstance(x, tuple) else pd.NA),
            "_PARSED_COD_CAR": parsed_sies.map(lambda x: x[1] if isinstance(x, tuple) else pd.NA),
            "_PARSED_JOR": parsed_sies.map(lambda x: x[2] if isinstance(x, tuple) else pd.NA),
            "_PARSED_VERSION": parsed_sies.map(lambda x: x[3] if isinstance(x, tuple) else pd.NA),
        },
        index=archivo_subida.index,
    )
    archivo_subida = pd.concat([archivo_subida, parsed_df], axis=1)

    # Estandarización de identificadores de oferta (sede/carrera/jornada/version).
    parsed_cod_sed = pd.to_numeric(archivo_subida["_PARSED_COD_SED"], errors="coerce")
    parsed_cod_car = pd.to_numeric(archivo_subida["_PARSED_COD_CAR"], errors="coerce")
    parsed_jor = pd.to_numeric(archivo_subida["_PARSED_JOR"], errors="coerce")
    parsed_version = pd.to_numeric(archivo_subida["_PARSED_VERSION"], errors="coerce")

    oferta_mod = pd.Series(pd.NA, index=archivo_subida.index, dtype="object")
    oferta_jor = pd.Series(pd.NA, index=archivo_subida.index, dtype="object")
    oferta_duracion = pd.Series(pd.NA, index=archivo_subida.index, dtype="object")
    oferta_cod_car = pd.Series(pd.NA, index=archivo_subida.index, dtype="object")
    if not oferta_dim.empty:
        oferta_idx = oferta_dim.drop_duplicates(subset=["CODIGO_UNICO"], keep="first").set_index("CODIGO_UNICO")
        oferta_mod = archivo_subida[FINAL_SIES_CODE_COL].map(oferta_idx["MODALIDAD"].to_dict())
        oferta_jor = archivo_subida[FINAL_SIES_CODE_COL].map(oferta_idx["JORNADA"].to_dict())
        oferta_duracion = archivo_subida[FINAL_SIES_CODE_COL].map(oferta_idx["DURACION_ESTUDIOS"].to_dict())
        if "CODIGO_CARRERA" in oferta_idx.columns:
            oferta_cod_car = archivo_subida[FINAL_SIES_CODE_COL].map(oferta_idx["CODIGO_CARRERA"].to_dict())

    cod_sed_pre_final = pd.to_numeric(archivo_subida["COD_SED"], errors="coerce")
    cod_car_pre_final = pd.to_numeric(archivo_subida["COD_CAR"], errors="coerce")
    jor_pre_final = pd.to_numeric(archivo_subida["JOR"], errors="coerce")
    mod_pre_final = pd.to_numeric(archivo_subida["MODALIDAD"], errors="coerce")

    archivo_subida["COD_SED"] = (
        parsed_cod_sed.combine_first(cod_sed_pre_final).astype("Int64")
    )
    oferta_cod_car_num = pd.to_numeric(oferta_cod_car, errors="coerce")
    archivo_subida["COD_CAR"] = oferta_cod_car_num.combine_first(parsed_cod_car).combine_first(cod_car_pre_final).astype("Int64")

    # ── Fallback COD_CAR: extraer de CODIGOS_SIES_POTENCIALES (ambiguos con COD_CAR compartido) ──
    _cod_car_still_missing = archivo_subida["COD_CAR"].isna()
    if _cod_car_still_missing.any() and "CODIGOS_SIES_POTENCIALES" in archivo_subida.columns:
        _shared_cod_car = archivo_subida.loc[_cod_car_still_missing, "CODIGOS_SIES_POTENCIALES"].apply(
            _extract_shared_cod_car_from_potenciales
        )
        _shared_cod_car = pd.to_numeric(_shared_cod_car, errors="coerce").astype("Int64")
        archivo_subida.loc[_cod_car_still_missing, "COD_CAR"] = archivo_subida.loc[
            _cod_car_still_missing, "COD_CAR"
        ].fillna(_shared_cod_car)
        _filled_shared = _cod_car_still_missing & archivo_subida["COD_CAR"].notna()
        print(f"    ↳ COD_CAR fallback SIES_POTENCIALES_COMPARTIDO: {int(_filled_shared.sum())} filas")

    # ── Fallback COD_CAR: mapeo NOMBRE_CARRERA → CODIGO_CARRERA vía DURACION_ESTUDIOS ──
    _cod_car_still_missing2 = archivo_subida["COD_CAR"].isna()
    if _cod_car_still_missing2.any() and not oferta_dim.empty:
        _nombre_to_cod = _build_nombre_carrera_to_cod_car(oferta_dim)
        if _nombre_to_cod and "NOMBRE_CARRERA_FUENTE" in archivo_subida.columns:
            _nombre_norm = archivo_subida.loc[_cod_car_still_missing2, "NOMBRE_CARRERA_FUENTE"].astype(str).apply(
                _normalize_nombre_carrera_for_lookup
            )
            _cod_car_by_nombre = _nombre_norm.map(_nombre_to_cod)
            _cod_car_by_nombre = pd.to_numeric(_cod_car_by_nombre, errors="coerce").astype("Int64")
            archivo_subida.loc[_cod_car_still_missing2, "COD_CAR"] = archivo_subida.loc[
                _cod_car_still_missing2, "COD_CAR"
            ].fillna(_cod_car_by_nombre)
            _filled_nombre = _cod_car_still_missing2 & archivo_subida["COD_CAR"].notna()
            print(f"    ↳ COD_CAR fallback NOMBRE_CARRERA_DURACION: {int(_filled_nombre.sum())} filas")
    jor_from_offer = pd.to_numeric(oferta_jor, errors="coerce")
    archivo_subida["JOR"] = jor_from_offer.combine_first(parsed_jor).combine_first(jor_pre_final).astype("Int64")
    archivo_subida["VERSION"] = parsed_version.astype("Int64")

    # ── Fallback VERSION: cuando COD_CAR+JOR están, usar la versión máxima de DURACION_ESTUDIOS ──
    _ver_missing = archivo_subida["VERSION"].isna() & archivo_subida["COD_CAR"].notna() & archivo_subida["JOR"].notna()
    if _ver_missing.any() and not oferta_dim.empty:
        _dur_ver = oferta_dim[["CODIGO_UNICO"]].copy()
        _dur_ver["_VERSION"] = _dur_ver["CODIGO_UNICO"].str.extract(r"V(\d+)$", expand=False).astype(float)
        _dur_ver["_COD_CAR"] = _dur_ver["CODIGO_UNICO"].str.extract(r"C(\d+)J", expand=False).astype(float)
        _dur_ver["_JOR"] = _dur_ver["CODIGO_UNICO"].str.extract(r"J(\d+)V", expand=False).astype(float)
        _max_ver = _dur_ver.groupby(["_COD_CAR", "_JOR"])["_VERSION"].max()
        _ver_key = list(zip(
            archivo_subida.loc[_ver_missing, "COD_CAR"].astype(float),
            archivo_subida.loc[_ver_missing, "JOR"].astype(float),
        ))
        _ver_fallback = pd.Series([_max_ver.get(k, pd.NA) for k in _ver_key], index=archivo_subida.loc[_ver_missing].index)
        _ver_fallback = pd.to_numeric(_ver_fallback, errors="coerce").astype("Int64")
        archivo_subida.loc[_ver_missing, "VERSION"] = _ver_fallback
        _filled_ver = _ver_missing & archivo_subida["VERSION"].notna()
        print(f"    ↳ VERSION fallback DURACION_ESTUDIOS max: {int(_filled_ver.sum())} filas")

    mod_from_offer = pd.to_numeric(oferta_mod, errors="coerce")
    mod_from_jor = pd.to_numeric(_modalidad_from_jor(archivo_subida["JOR"]), errors="coerce")
    archivo_subida["MODALIDAD"] = mod_from_offer.combine_first(mod_from_jor).combine_first(mod_pre_final).astype("Int64")
    archivo_subida["DURACION_ESTUDIOS_REF"] = pd.to_numeric(oferta_duracion, errors="coerce").astype("Int64")

    cod_sed_source = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    cod_sed_method = pd.Series("SIN_METODO_FINAL", index=archivo_subida.index, dtype="object")
    cod_sed_audit = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    parsed_cod_sed_mask = parsed_cod_sed.notna()
    cod_sed_source.loc[parsed_cod_sed_mask] = "CODIGO_CARRERA_SIES_FINAL"
    cod_sed_method.loc[parsed_cod_sed_mask] = "PARSE_COMPONENTE_SIES"
    cod_sed_audit.loc[parsed_cod_sed_mask] = "CONSISTENTE_SIES_FINAL"
    legacy_cod_sed_mask = (~parsed_cod_sed_mask) & cod_sed_pre_final.notna()
    cod_sed_source.loc[legacy_cod_sed_mask] = archivo_subida.loc[legacy_cod_sed_mask, "COD_SED_STATUS"].replace(
        {
            "MAPEADO_GOB_SEDE": "DA_SEDE_GOBERNANZA",
            "SOURCE_EXACT": "FUENTE_EXACTA",
            "SIN_MAPEO_GOB_SEDE": "DA_SEDE_SIN_MAPEO",
        }
    )
    cod_sed_method.loc[legacy_cod_sed_mask] = archivo_subida.loc[legacy_cod_sed_mask, "COD_SED_STATUS"].replace(
        {
            "MAPEADO_GOB_SEDE": "MAPEO_GOB_SEDE",
            "SOURCE_EXACT": "COPIA_FUENTE",
            "SIN_MAPEO_GOB_SEDE": "SIN_METODO_FINAL",
        }
    )
    cod_sed_audit.loc[legacy_cod_sed_mask] = "LEGACY_PRE_FINAL"
    archivo_subida["COD_SED_FUENTE_FINAL"] = cod_sed_source
    archivo_subida["COD_SED_METODO_FINAL"] = cod_sed_method
    archivo_subida["COD_SED_AUDIT_STATUS"] = cod_sed_audit

    cod_car_source = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    cod_car_method = pd.Series("SIN_METODO_FINAL", index=archivo_subida.index, dtype="object")
    cod_car_audit = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    # Fuente 1: oferta/parsed SIES (cascada principal)
    parsed_cod_car_mask = parsed_cod_car.notna()
    oferta_cod_car_mask = oferta_cod_car_num.notna()
    sies_primary_mask = oferta_cod_car_mask | parsed_cod_car_mask
    cod_car_source.loc[sies_primary_mask] = "CODIGO_CARRERA_SIES_FINAL"
    cod_car_method.loc[sies_primary_mask] = "PARSE_COMPONENTE_SIES"
    cod_car_audit.loc[sies_primary_mask] = "CONSISTENTE_SIES_FINAL"
    cod_car_method.loc[oferta_cod_car_mask] = "OFERTA_LOOKUP_CODIGO_UNICO"
    # Fuente 2: fallback SIES potenciales compartidos
    _fb_shared_mask = (~sies_primary_mask) & _cod_car_still_missing & archivo_subida["COD_CAR"].notna()
    # Restar los que se llenaron por nombre (más adelante se marca)
    # _fb_shared_mask captura las filas que NO tenían COD_CAR antes pero ahora sí
    # y no vienen del parsed ni oferta → vienen del fallback compartido o nombre
    cod_car_source.loc[_fb_shared_mask] = "CODIGOS_SIES_POTENCIALES"
    cod_car_method.loc[_fb_shared_mask] = "COMPONENTE_C_COMPARTIDO"
    cod_car_audit.loc[_fb_shared_mask] = "INFERIDO_AMBIGUOS_MISMO_COD_CAR"
    # Fuente 3: fallback nombre carrera
    _fb_nombre_mask = (~sies_primary_mask) & _cod_car_still_missing2 & archivo_subida["COD_CAR"].notna()
    cod_car_source.loc[_fb_nombre_mask] = "DURACION_ESTUDIOS_NOMBRE"
    cod_car_method.loc[_fb_nombre_mask] = "MAPEO_NOMBRE_CARRERA"
    cod_car_audit.loc[_fb_nombre_mask] = "INFERIDO_NOMBRE_CARRERA_DURACION"
    # Fuente 4: legacy
    legacy_cod_car_mask = (~sies_primary_mask) & (~_fb_shared_mask) & (~_fb_nombre_mask) & cod_car_pre_final.notna()
    cod_car_source.loc[legacy_cod_car_mask] = "CODCARPR_FUENTE"
    cod_car_method.loc[legacy_cod_car_mask] = "COPIA_FUENTE"
    cod_car_audit.loc[legacy_cod_car_mask] = "LEGACY_PRE_FINAL"
    archivo_subida["COD_CAR_FUENTE_FINAL"] = cod_car_source
    archivo_subida["COD_CAR_METODO_FINAL"] = cod_car_method
    archivo_subida["COD_CAR_AUDIT_STATUS"] = cod_car_audit

    jor_source = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    jor_method = pd.Series("SIN_METODO_FINAL", index=archivo_subida.index, dtype="object")
    jor_audit = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    offer_jor_mask = jor_from_offer.notna()
    parsed_jor_mask = parsed_jor.notna()
    legacy_jor_mask = (~offer_jor_mask) & (~parsed_jor_mask) & jor_pre_final.notna()
    jor_matches_parsed = archivo_subida["JOR"].astype("Float64").eq(parsed_jor).fillna(False)
    jor_source.loc[offer_jor_mask] = "OFERTA_ACADEMICA"
    jor_method.loc[offer_jor_mask] = "LOOKUP_CODIGO_UNICO"
    jor_audit.loc[offer_jor_mask & parsed_jor_mask & jor_matches_parsed] = "CONSISTENTE_OFERTA_SIES"
    jor_audit.loc[offer_jor_mask & (~parsed_jor_mask)] = "CONSISTENTE_OFERTA"
    jor_audit.loc[offer_jor_mask & parsed_jor_mask & (~jor_matches_parsed)] = "DESALINEADO_OFERTA_SIES"
    jor_source.loc[(~offer_jor_mask) & parsed_jor_mask] = "CODIGO_CARRERA_SIES_FINAL"
    jor_method.loc[(~offer_jor_mask) & parsed_jor_mask] = "PARSE_COMPONENTE_SIES"
    jor_audit.loc[(~offer_jor_mask) & parsed_jor_mask] = "CONSISTENTE_SIES_FINAL"
    jor_source.loc[legacy_jor_mask] = "JORNADA_FUENTE_LEGACY"
    jor_method.loc[legacy_jor_mask] = "MAPEO_JORNADA_LEGACY"
    jor_audit.loc[legacy_jor_mask] = "LEGACY_PRE_FINAL"
    archivo_subida["JOR_FUENTE_FINAL"] = jor_source
    archivo_subida["JOR_METODO_FINAL"] = jor_method
    archivo_subida["JOR_AUDIT_STATUS"] = jor_audit

    mod_source = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    mod_method = pd.Series("SIN_METODO_FINAL", index=archivo_subida.index, dtype="object")
    mod_audit = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    offer_mod_mask = mod_from_offer.notna()
    derived_mod_mask = mod_from_jor.notna()
    legacy_mod_mask = (~offer_mod_mask) & (~derived_mod_mask) & mod_pre_final.notna()
    mod_matches_jor = archivo_subida["MODALIDAD"].astype("Float64").eq(mod_from_jor).fillna(False)
    mod_source.loc[offer_mod_mask] = "OFERTA_ACADEMICA"
    mod_method.loc[offer_mod_mask] = "LOOKUP_CODIGO_UNICO"
    mod_audit.loc[offer_mod_mask & derived_mod_mask & mod_matches_jor] = "CONSISTENTE_OFERTA_JOR"
    mod_audit.loc[offer_mod_mask & (~derived_mod_mask)] = "CONSISTENTE_OFERTA"
    mod_audit.loc[offer_mod_mask & derived_mod_mask & (~mod_matches_jor)] = "DESALINEADO_OFERTA_JOR"
    mod_source.loc[(~offer_mod_mask) & derived_mod_mask] = "JOR_FINAL"
    mod_method.loc[(~offer_mod_mask) & derived_mod_mask] = "DERIVACION_DESDE_JOR"
    mod_audit.loc[(~offer_mod_mask) & derived_mod_mask] = "DERIVADA_DESDE_JOR"
    mod_source.loc[legacy_mod_mask] = "JORNADA_FUENTE_LEGACY"
    mod_method.loc[legacy_mod_mask] = "MAPEO_JORNADA_LEGACY"
    mod_audit.loc[legacy_mod_mask] = "LEGACY_PRE_FINAL"
    archivo_subida["MODALIDAD_FUENTE_FINAL"] = mod_source
    archivo_subida["MODALIDAD_METODO_FINAL"] = mod_method
    archivo_subida["MODALIDAD_AUDIT_STATUS"] = mod_audit

    version_source = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    version_method = pd.Series("SIN_METODO_FINAL", index=archivo_subida.index, dtype="object")
    version_audit = pd.Series("SIN_FUENTE_FINAL", index=archivo_subida.index, dtype="object")
    parsed_version_mask = parsed_version.notna()
    version_source.loc[parsed_version_mask] = "CODIGO_CARRERA_SIES_FINAL"
    version_method.loc[parsed_version_mask] = "PARSE_COMPONENTE_SIES"
    version_audit.loc[parsed_version_mask] = "CONSISTENTE_SIES_FINAL"
    archivo_subida["VERSION_FUENTE_FINAL"] = version_source
    archivo_subida["VERSION_METODO_FINAL"] = version_method
    archivo_subida["VERSION_AUDIT_STATUS"] = version_audit

    # Sexo: homologa catálogos F/M/S -> H/M/NB.
    archivo_subida["SEXO"] = archivo_subida["SEXO"].map(_normalize_sexo_mu)
    for c in ["PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "NOMBRE"]:
        archivo_subida[c] = archivo_subida[c].map(_normalize_text)

    # Año y semestre: normalización explícita (3 -> 2), y completitud mínima.
    anio_act = pd.to_numeric(archivo_subida["ANIO_ING_ACT"], errors="coerce")
    anio_act = anio_act.where(anio_act.between(1990, 2026), pd.NA).fillna(2026).astype("Int64")
    archivo_subida["ANIO_ING_ACT"] = anio_act

    sem_act = pd.to_numeric(archivo_subida["SEM_ING_ACT"], errors="coerce").replace({3: 2})
    sem_act = sem_act.where(sem_act.isin([1, 2]), pd.NA).fillna(1).astype("Int64")
    archivo_subida["SEM_ING_ACT"] = sem_act

    anio_ori = pd.to_numeric(archivo_subida["ANIO_ING_ORI"], errors="coerce").fillna(anio_act)
    anio_ori = anio_ori.where((anio_ori == 1900) | anio_ori.between(1980, 2026), anio_act).astype("Int64")
    archivo_subida["ANIO_ING_ORI"] = anio_ori

    sem_ori = pd.to_numeric(archivo_subida["SEM_ING_ORI"], errors="coerce").replace({3: 2})
    sem_ori = sem_ori.where(sem_ori.isin([0, 1, 2]), pd.NA).fillna(sem_act)
    sem_ori = sem_ori.where(anio_ori != 1900, 0).astype("Int64")
    archivo_subida["SEM_ING_ORI"] = sem_ori

    # FOR_ING_ACT: solo aceptar catálogo 1..11; no usar 10 como fallback silencioso.
    for_ing = pd.to_numeric(archivo_subida["FOR_ING_ACT"], errors="coerce")
    for_ing = for_ing.where(for_ing.isin(sorted(valid_for_ing_act_codes)), pd.NA).astype("Int64")
    # BUG-004 fix: usar valor derivado por _resolve_for_ing_act_row; fallback a 1 solo si NA.
    _for_na_mask = for_ing.isna()
    archivo_subida["FOR_ING_ACT"] = for_ing.fillna(1).astype("Int64")
    # Trazabilidad: solo actualizar las filas que fueron defaulteadas a 1 por NA.
    archivo_subida.loc[_for_na_mask, "FOR_ING_ACT_FUENTE_VALOR"] = "1"
    archivo_subida.loc[_for_na_mask, "FOR_ING_ACT_FUENTE_CAMPO"] = "DEFAULT_FALLBACK_NA"
    archivo_subida.loc[_for_na_mask, "FOR_ING_ACT_FUENTE_NORM"] = "1"
    archivo_subida.loc[_for_na_mask, "FOR_ING_ACT_METODO"] = "DEFAULT_1_NA_FALLBACK"
    archivo_subida.loc[_for_na_mask, "FOR_ING_ACT_IMPUTADO"] = "SI"
    archivo_subida.loc[_for_na_mask, "FOR_ING_ACT_REQUIERE_REVISION"] = "SI"

    # --- DA-based overrides: detectar FOR=2 (continuidad) y FOR=3 (cambio interno) ---
    _nombre_carr = archivo_subida.get("NOMBRE_CARRERA_FUENTE", pd.Series("", index=archivo_subida.index)).fillna("").astype(str).str.upper()
    _da_sit = archivo_subida.get("DA_SITUACION", pd.Series("", index=archivo_subida.index)).fillna("").astype(str).str.strip().str.upper()
    _sit_interno = {"24 - CAMBIO DE CARRERA", "49 - CAMBIO DE JORNADA", "27 - CAMBIO PLAN OTRA JORNADA"}
    _is_continuidad = _nombre_carr.str.contains("CONTINUIDAD", na=False)
    _is_cambio_int = _da_sit.isin(_sit_interno)
    # Prioridad: continuidad (2) > cambio interno (3) > valor actual.
    _m3 = _is_cambio_int & ~_is_continuidad
    archivo_subida.loc[_m3, "FOR_ING_ACT"] = 3
    archivo_subida.loc[_m3, "FOR_ING_ACT_METODO"] = "DA_SITUACION_CAMBIO_INTERNO"
    archivo_subida.loc[_m3, "FOR_ING_ACT_FUENTE_CAMPO"] = "DA_SITUACION"
    archivo_subida.loc[_m3, "FOR_ING_ACT_FUENTE_VALOR"] = _da_sit[_m3]
    archivo_subida.loc[_m3, "FOR_ING_ACT_FUENTE_NORM"] = "3"
    archivo_subida.loc[_m3, "FOR_ING_ACT_IMPUTADO"] = "NO"
    archivo_subida.loc[_m3, "FOR_ING_ACT_REQUIERE_REVISION"] = "NO"
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT"] = 2
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT_METODO"] = "DA_NOMBRE_CONTINUIDAD"
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT_FUENTE_CAMPO"] = "NOMBRE_CARRERA_FUENTE"
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT_FUENTE_VALOR"] = _nombre_carr[_is_continuidad]
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT_FUENTE_NORM"] = "2"
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT_IMPUTADO"] = "NO"
    archivo_subida.loc[_is_continuidad, "FOR_ING_ACT_REQUIERE_REVISION"] = "NO"

    # --- DA-based override: FOR=11 (articulación) desde trace del motor standalone ---
    _trace_path = Path(__file__).resolve().parent / "control" / "for_ing_act_trace_long.tsv"
    if _trace_path.exists():
        _trace = pd.read_csv(_trace_path, sep="\t", usecols=["_RUT_NUM", "FOR_ING_ACT", "FOR_ING_ACT_RULE_DA"])
        _trace_11 = _trace[_trace["FOR_ING_ACT"] == 11][["_RUT_NUM"]].drop_duplicates()
        _rut_col = pd.to_numeric(archivo_subida["N_DOC"], errors="coerce").astype("Int64")
        _is_art = _rut_col.isin(_trace_11["_RUT_NUM"].dropna().astype("Int64"))
        # Solo marcar articulación si el programa actual es profesional (no técnico).
        _codcarpr = archivo_subida.get("CODCARPR_NORM", pd.Series("", index=archivo_subida.index)).fillna("").astype(str)
        _is_tecnico = _codcarpr.str.match(r"^T", na=False) | _nombre_carr.str.contains(r"TECNICO|TNS", na=False)
        _is_profesional = ~_is_tecnico
        _m11 = _is_art & _is_profesional
        archivo_subida.loc[_m11, "FOR_ING_ACT"] = 11
        archivo_subida.loc[_m11, "FOR_ING_ACT_METODO"] = "DA_TRACE_ARTICULACION_11"
        archivo_subida.loc[_m11, "FOR_ING_ACT_FUENTE_CAMPO"] = "TRACE_MOTOR_FOR_ING_ACT"
        archivo_subida.loc[_m11, "FOR_ING_ACT_FUENTE_VALOR"] = "11"
        archivo_subida.loc[_m11, "FOR_ING_ACT_FUENTE_NORM"] = "11"
        archivo_subida.loc[_m11, "FOR_ING_ACT_IMPUTADO"] = "NO"
        archivo_subida.loc[_m11, "FOR_ING_ACT_REQUIERE_REVISION"] = "NO"

    # FOR_ING_ACT == 1 → ORI == ACT (política: ingreso directo, origen coincide con actual).
    for_equal = archivo_subida["FOR_ING_ACT"].eq(1)
    archivo_subida.loc[for_equal, "ANIO_ING_ORI"] = archivo_subida.loc[for_equal, "ANIO_ING_ACT"]
    archivo_subida.loc[for_equal, "SEM_ING_ORI"] = archivo_subida.loc[for_equal, "SEM_ING_ACT"]
    # Trazabilidad ORI: diferenciar FOR=1 (copia) vs FOR!=1 (preservado).
    archivo_subida["ANIO_ING_ORI_FUENTE_FINAL"] = "PRESERVADO_VALOR_DERIVADO"
    archivo_subida["ANIO_ING_ORI_METODO_FINAL"] = "SIN_OVERRIDE"
    archivo_subida["ANIO_ING_ORI_AUDIT_STATUS"] = "VALOR_ORIGINAL_PRESERVADO"
    archivo_subida["SEM_ING_ORI_FUENTE_FINAL"] = "PRESERVADO_VALOR_DERIVADO"
    archivo_subida["SEM_ING_ORI_METODO_FINAL"] = "SIN_OVERRIDE"
    archivo_subida["SEM_ING_ORI_AUDIT_STATUS"] = "VALOR_ORIGINAL_PRESERVADO"
    archivo_subida.loc[for_equal, "ANIO_ING_ORI_FUENTE_FINAL"] = "POLITICA_FOR_ING_ACT_1"
    archivo_subida.loc[for_equal, "ANIO_ING_ORI_METODO_FINAL"] = "COPIA_DESDE_ANIO_ING_ACT"
    archivo_subida.loc[for_equal, "ANIO_ING_ORI_AUDIT_STATUS"] = "IGUAL_ACTUAL_POR_POLITICA_FOR_ING_ACT_1"
    archivo_subida.loc[for_equal, "SEM_ING_ORI_FUENTE_FINAL"] = "POLITICA_FOR_ING_ACT_1"
    archivo_subida.loc[for_equal, "SEM_ING_ORI_METODO_FINAL"] = "COPIA_DESDE_SEM_ING_ACT"
    archivo_subida.loc[for_equal, "SEM_ING_ORI_AUDIT_STATUS"] = "IGUAL_ACTUAL_POR_POLITICA_FOR_ING_ACT_1"

    asi_ins_ant = pd.to_numeric(archivo_subida["ASI_INS_ANT"], errors="coerce")
    asi_ins_ant = asi_ins_ant.where(asi_ins_ant.between(0, 99), pd.NA).fillna(0).astype("Int64")
    asi_apr_ant = pd.to_numeric(archivo_subida["ASI_APR_ANT"], errors="coerce")
    asi_apr_ant = asi_apr_ant.where(asi_apr_ant.between(0, 99), pd.NA)
    asi_apr_ant_cap_mask = asi_apr_ant > asi_ins_ant
    asi_apr_ant = asi_apr_ant.where(~asi_apr_ant_cap_mask, asi_ins_ant).astype("Int64")
    archivo_subida["ASI_INS_ANT"] = asi_ins_ant
    archivo_subida["ASI_APR_ANT"] = asi_apr_ant
    archivo_subida.loc[
        archivo_subida["ASI_INS_ANT_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1"),
        "ASI_INS_ANT_AUDIT_STATUS",
    ] = "CALCULADO_ANIO_REFERENCIA_HIST"
    archivo_subida.loc[
        archivo_subida["ASI_APR_ANT_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1"),
        "ASI_APR_ANT_AUDIT_STATUS",
    ] = "CALCULADO_APROB_ANIO_REFERENCIA_EXCL_EQUIV"
    archivo_subida.loc[asi_apr_ant_cap_mask.fillna(False), "ASI_APR_ANT_AUDIT_STATUS"] = "CAP_APR_A_INS_ANT"

    prom_pri = pd.to_numeric(archivo_subida["PROM_PRI_SEM"], errors="coerce")
    prom_pri = prom_pri.where((prom_pri == 0) | prom_pri.between(100, 700), pd.NA).fillna(0).astype("Int64")
    prom_seg = pd.to_numeric(archivo_subida["PROM_SEG_SEM"], errors="coerce")
    prom_seg = prom_seg.where((prom_seg == 0) | prom_seg.between(100, 700), pd.NA).fillna(0).astype("Int64")
    archivo_subida["PROM_PRI_SEM"] = prom_pri
    archivo_subida["PROM_SEG_SEM"] = prom_seg
    archivo_subida.loc[
        archivo_subida["PROM_PRI_SEM_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1") & prom_pri.eq(0),
        "PROM_PRI_SEM_AUDIT_STATUS",
    ] = "SIN_NOTAS_CALIFICABLES_SEM1_ANIO_REF"
    archivo_subida.loc[
        archivo_subida["PROM_PRI_SEM_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1") & prom_pri.gt(0),
        "PROM_PRI_SEM_AUDIT_STATUS",
    ] = "CALCULADO_NOTAS_MU_SEM1_ANIO_REF"
    archivo_subida.loc[
        archivo_subida["PROM_SEG_SEM_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1") & prom_seg.eq(0),
        "PROM_SEG_SEM_AUDIT_STATUS",
    ] = "SIN_NOTAS_CALIFICABLES_SEM2_ANIO_REF"
    archivo_subida.loc[
        archivo_subida["PROM_SEG_SEM_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1") & prom_seg.gt(0),
        "PROM_SEG_SEM_AUDIT_STATUS",
    ] = "CALCULADO_NOTAS_MU_SEM2_ANIO_REF"

    asi_ins_his = pd.to_numeric(archivo_subida["ASI_INS_HIS"], errors="coerce")
    asi_ins_his = asi_ins_his.where(asi_ins_his.between(0, 200), pd.NA).fillna(0).astype("Int64")
    asi_apr_his = pd.to_numeric(archivo_subida["ASI_APR_HIS"], errors="coerce")
    asi_apr_his = asi_apr_his.where(asi_apr_his.between(0, 200), pd.NA)
    asi_apr_his_cap_mask = asi_apr_his > asi_ins_his
    asi_apr_his = asi_apr_his.where(~asi_apr_his_cap_mask, asi_ins_his).astype("Int64")
    archivo_subida["ASI_INS_HIS"] = asi_ins_his
    archivo_subida["ASI_APR_HIS"] = asi_apr_his

    # REGLA BLOQUEANTE: si VIG=0, las 4 columnas deben ser 0
    vig_cero = archivo_subida["VIG"] == 0
    if vig_cero.any():
        archivo_subida.loc[vig_cero, "PROM_PRI_SEM"] = 0
        archivo_subida.loc[vig_cero, "PROM_SEG_SEM"] = 0
        archivo_subida.loc[vig_cero, "ASI_INS_HIS"] = 0
        archivo_subida.loc[vig_cero, "ASI_APR_HIS"] = 0

    hist_scope_status = archivo_subida["UZ_HIST_SCOPE_STATUS"].astype("object").copy()
    archivo_subida.loc[
        archivo_subida["ASI_INS_HIS_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1"),
        "ASI_INS_HIS_AUDIT_STATUS",
    ] = hist_scope_status.where(hist_scope_status.ne(""), "CALCULADO_HISTORICO_SIN_SCOPE")
    archivo_subida.loc[
        archivo_subida["ASI_APR_HIS_AUDIT_STATUS"].eq("CALCULADO_DESDE_HISTORICO_HOJA1"),
        "ASI_APR_HIS_AUDIT_STATUS",
    ] = hist_scope_status.where(hist_scope_status.ne(""), "CALCULADO_HISTORICO_SIN_SCOPE")
    archivo_subida.loc[asi_apr_his_cap_mask.fillna(False), "ASI_APR_HIS_AUDIT_STATUS"] = "CAP_APR_A_INS_HIS"

    # Campos numéricos obligatorios con fallback operativo explícito.
    nac_num = pd.to_numeric(archivo_subida["NAC"], errors="coerce")
    nac_status_final = archivo_subida["NAC_STATUS"].astype("object").copy()
    nac_default_mask = ~nac_num.between(1, 197)
    nac_status_final.loc[nac_default_mask & nac_status_final.eq("SIN_INSUMO")] = "DEFAULT_38_SIN_INSUMO"
    nac_status_final.loc[nac_default_mask & nac_status_final.eq("SOURCE_TEXT")] = "DEFAULT_38_SOURCE_TEXT_INVALIDO"
    nac_status_final.loc[nac_default_mask & nac_status_final.eq("SIN_MAPEO_GOB_NAC")] = "DEFAULT_38_SIN_MAPEO_GOB_NAC"
    nac_status_final.loc[nac_default_mask & nac_status_final.eq("REVISION_MANUAL_GOB_NAC")] = "DEFAULT_38_REVISION_MANUAL_GOB_NAC"
    nac_status_final.loc[nac_default_mask & nac_status_final.eq("")] = "DEFAULT_38_SIN_TRAZA"
    archivo_subida["NAC_STATUS"] = nac_status_final
    archivo_subida["NAC"] = nac_num.where(nac_num.between(1, 197), pd.NA).fillna(38).astype("Int64")
    pais_num = pd.to_numeric(archivo_subida["PAIS_EST_SEC"], errors="coerce")
    pais_status_final = archivo_subida["PAIS_EST_SEC_STATUS"].astype("object").copy()
    pais_default_mask = ~pais_num.between(1, 197)
    pais_status_final.loc[pais_default_mask & pais_status_final.eq("SIN_INSUMO")] = "DEFAULT_38_SIN_INSUMO"
    pais_status_final.loc[pais_default_mask & pais_status_final.eq("SOURCE_EMPTY")] = "DEFAULT_38_SOURCE_EMPTY"
    pais_status_final.loc[pais_default_mask & pais_status_final.eq("SOURCE_EXACT")] = "DEFAULT_38_SOURCE_INVALIDO"
    pais_status_final.loc[pais_default_mask & pais_status_final.eq("")] = "DEFAULT_38_SIN_TRAZA"
    archivo_subida["PAIS_EST_SEC_STATUS"] = pais_status_final
    archivo_subida["PAIS_EST_SEC"] = pais_num.where(pais_num.between(1, 197), pd.NA).fillna(38).astype("Int64")

    niv_aca_raw = pd.to_numeric(archivo_subida["NIV_ACA"], errors="coerce")
    niv_aca_status_final = archivo_subida["NIV_ACA_AUDIT_STATUS"].astype("object").copy()
    niv_default_mask = ~niv_aca_raw.ge(1)
    niv_aca = niv_aca_raw.where(niv_aca_raw >= 1, pd.NA).fillna(1)
    dur_ref = pd.to_numeric(archivo_subida["DURACION_ESTUDIOS_REF"], errors="coerce")
    niv_capped_dur_mask = niv_aca.gt(dur_ref).fillna(False)
    niv_aca = niv_aca.where(dur_ref.isna() | (niv_aca <= dur_ref), dur_ref)
    niv_capped_2026_mask = (archivo_subida["ANIO_ING_ORI"].eq(2026) & niv_aca.gt(2)).fillna(False)
    niv_aca = niv_aca.where(~archivo_subida["ANIO_ING_ORI"].eq(2026) | (niv_aca <= 2), 2)
    niv_aca_status_final.loc[niv_default_mask & niv_aca_status_final.eq("DEFAULT_1_SIN_FUENTE")] = "DEFAULT_1_SIN_FUENTE"
    niv_aca_status_final.loc[niv_default_mask & niv_aca_status_final.eq("DEFAULT_1_INPUT_INVALIDO")] = "DEFAULT_1_INPUT_INVALIDO"
    niv_aca_status_final.loc[niv_default_mask & niv_aca_status_final.eq("DEFAULT_1_DA_INVALIDO")] = "DEFAULT_1_DA_INVALIDO"
    niv_aca_status_final.loc[niv_default_mask & niv_aca_status_final.eq("DEFAULT_1_INPUT_DA_INVALIDOS")] = "DEFAULT_1_INPUT_DA_INVALIDOS"
    niv_aca_status_final.loc[niv_capped_dur_mask & ~niv_capped_2026_mask] = "ACOTADO_DURACION_ESTUDIOS"
    niv_aca_status_final.loc[(~niv_capped_dur_mask) & niv_capped_2026_mask] = "ACOTADO_COHORTE_2026"
    niv_aca_status_final.loc[niv_capped_dur_mask & niv_capped_2026_mask] = "ACOTADO_DURACION_Y_COHORTE_2026"
    archivo_subida["NIV_ACA"] = niv_aca.astype("Int64")
    archivo_subida["NIV_ACA_AUDIT_STATUS"] = niv_aca_status_final

    archivo_subida["SIT_FON_SOL"] = pd.Series(1, index=archivo_subida.index, dtype="Int64")
    archivo_subida["SUS_PRE"] = pd.Series(0, index=archivo_subida.index, dtype="Int64")
    archivo_subida["REINCORPORACION"] = pd.Series(0, index=archivo_subida.index, dtype="Int64")
    vig = pd.to_numeric(archivo_subida["VIG"], errors="coerce")
    archivo_subida["VIG"] = vig.where(vig.isin([0, 1, 2]), pd.NA).fillna(1).astype("Int64")

    # REGLA BLOQUEANTE INSTITUCIONAL: Forzar VIG=0 para TITULADO/ELIMINADO/SUSPENDIDO
    # Fundamento: En MU 2026, estos estados se clasifican como "sin matrícula"
    if "DA_ESTADOACADEMICO" in archivo_subida.columns:
        titulo_elim_susp = archivo_subida["DA_ESTADOACADEMICO"].fillna("").astype(str).str.strip().str.upper().isin(
            ["TITULADO", "ELIMINADO", "SUSPENDIDO"]
        )
        archivo_subida.loc[titulo_elim_susp, "VIG"] = 0

    # Regla bloqueante definitiva: tras resolver VIG final, forzar 4 columnas en cero para VIG=0.
    vig_cero_post_force = pd.to_numeric(archivo_subida["VIG"], errors="coerce").eq(0)
    if vig_cero_post_force.any():
        archivo_subida.loc[vig_cero_post_force, "PROM_PRI_SEM"] = 0
        archivo_subida.loc[vig_cero_post_force, "PROM_SEG_SEM"] = 0
        archivo_subida.loc[vig_cero_post_force, "ASI_INS_HIS"] = 0
        archivo_subida.loc[vig_cero_post_force, "ASI_APR_HIS"] = 0

    vig_final = pd.to_numeric(archivo_subida["VIG"], errors="coerce").astype("Int64")
    esperado = pd.to_numeric(archivo_subida["VIG_ESPERADO_DA"], errors="coerce").astype("Int64")
    flag_vig = pd.Series("SIN_REGLA_GOB_DA", index=archivo_subida.index, dtype="object")
    flag_vig.loc[archivo_subida["DA_ESTADOACADEMICO"].fillna("").astype(str).str.strip().eq("")] = "SIN_ESTADO_DA"
    flag_vig.loc[esperado.notna() & vig_final.eq(esperado)] = "OK"
    flag_vig.loc[esperado.notna() & vig_final.ne(esperado)] = "INCONSISTENTE"
    archivo_subida["FLAG_INCONSISTENCIA_VIG"] = flag_vig

    # Fechas: formato dd/mm/yyyy y fallback 01/01/1900 cuando no hay dato.
    archivo_subida["FECH_NAC"] = _to_ddmmyyyy(archivo_subida["FECH_NAC"], fallback="01/01/1900")
    archivo_subida.loc[
        archivo_subida["FECH_NAC"].eq("01/01/1900") & archivo_subida["FECH_NAC_STATUS"].eq("SIN_FUENTE"),
        "FECH_NAC_STATUS",
    ] = "FALLBACK_1900_SIN_FUENTE"
    archivo_subida.loc[
        archivo_subida["FECH_NAC"].eq("01/01/1900") & archivo_subida["FECH_NAC_STATUS"].eq("SOURCE_INPUT"),
        "FECH_NAC_STATUS",
    ] = "FALLBACK_1900_INPUT_INVALIDO"
    archivo_subida.loc[
        archivo_subida["FECH_NAC"].eq("01/01/1900") & archivo_subida["FECH_NAC_STATUS"].eq("FALLBACK_DATOS_ALUMNOS"),
        "FECH_NAC_STATUS",
    ] = "FALLBACK_1900_DATOS_ALUMNOS_INVALIDO"
    fecha_mat_dt = pd.to_datetime(archivo_subida["FECHA_MATRICULA"], errors="coerce", dayfirst=True)
    fecha_mat_status_final = archivo_subida["FECHA_MATRICULA_AUDIT_STATUS"].astype("object").copy()
    fecha_mat_source_final = archivo_subida["FECHA_MATRICULA_FUENTE_FINAL"].astype("object").copy()
    fecha_mat_method_final = archivo_subida["FECHA_MATRICULA_METODO_FINAL"].astype("object").copy()
    fecha_mat_future_mask = fecha_mat_dt > pd.Timestamp.today().normalize()
    fecha_mat_dt = fecha_mat_dt.where(~fecha_mat_future_mask, pd.Timestamp.today().normalize())
    fecha_mat_fmt = _to_ddmmyyyy(fecha_mat_dt, fallback="01/01/1900")
    # Manual MU 2026: FECHA_MATRICULA solo aplica a cohorte de origen 2026.
    fecha_aplica_mask = archivo_subida["ANIO_ING_ORI"].eq(2026)
    fecha_mat_fmt = fecha_mat_fmt.where(fecha_aplica_mask, "01/01/1900")
    fecha_mat_source_final.loc[~fecha_aplica_mask] = "POLITICA_COHORTE_ORIGEN_2026"
    fecha_mat_method_final.loc[~fecha_aplica_mask] = "NO_APLICA_FUERA_COHORTE_2026"
    fecha_mat_status_final.loc[~fecha_aplica_mask] = "POLITICA_1900_FUERA_COHORTE_2026"
    fecha_mat_status_final.loc[fecha_aplica_mask & fecha_mat_future_mask] = "AJUSTADA_FECHA_FUTURA_COHORTE_2026"
    fecha_mat_status_final.loc[fecha_aplica_mask & fecha_mat_fmt.eq("01/01/1900") & fecha_mat_status_final.eq("DEFAULT_1900_SIN_FUENTE")] = "FALLBACK_1900_COHORTE_2026_SIN_FUENTE"
    fecha_mat_status_final.loc[fecha_aplica_mask & fecha_mat_fmt.eq("01/01/1900") & fecha_mat_status_final.eq("DEFAULT_1900_INPUT_INVALIDA")] = "FALLBACK_1900_COHORTE_2026_INPUT_INVALIDA"
    fecha_mat_status_final.loc[fecha_aplica_mask & fecha_mat_fmt.eq("01/01/1900") & fecha_mat_status_final.eq("DEFAULT_1900_DA_INVALIDA")] = "FALLBACK_1900_COHORTE_2026_DA_INVALIDA"
    fecha_mat_status_final.loc[fecha_aplica_mask & fecha_mat_fmt.eq("01/01/1900") & fecha_mat_status_final.eq("DEFAULT_1900_INPUT_DA_INVALIDAS")] = "FALLBACK_1900_COHORTE_2026_INPUT_DA_INVALIDAS"
    archivo_subida["FECHA_MATRICULA"] = fecha_mat_fmt
    archivo_subida["FECHA_MATRICULA_FUENTE_FINAL"] = fecha_mat_source_final
    archivo_subida["FECHA_MATRICULA_METODO_FINAL"] = fecha_mat_method_final
    archivo_subida["FECHA_MATRICULA_AUDIT_STATUS"] = fecha_mat_status_final

    # Construcción de carga final (pregrado): excluir diplomados, no-match de datos alumnos
    # y deduplicar por clave de matrícula.
    estado_carga = pd.Series("OK_CARGA_PREGRADO", index=archivo_subida.index, dtype="object")
    # Excluir filas sin COD_CAR (incluye SIN_MATCH_SIES + ambiguos no resueltos).
    # Filas que obtuvieron COD_CAR por fallback (COMPONENTE_C_COMPARTIDO,
    # MAPEO_NOMBRE_CARRERA) ya NO se excluyen aquí; su completitud regulatoria
    # se valida en EXCLUIDO_CAMPOS_OBLIGATORIOS más adelante.
    sin_cod_car_final = archivo_subida["COD_CAR"].isna()
    excl_dipl = archivo_subida["ES_DIPLOMADO"].fillna(False)
    sin_match_da = archivo_subida["DA_MATCH_MODO"] == "SIN_MATCH"
    heuristica_sies_opaca = archivo_subida["SIES_RESOLUCION_HEURISTICA"].astype("object").eq("PRIMERA_OPCION")
    sin_for_ing_trazable = archivo_subida["FOR_ING_ACT"].isna()

    estado_carga.loc[sin_cod_car_final] = "EXCLUIDO_SIN_MATCH_SIES"
    estado_carga.loc[(estado_carga == "OK_CARGA_PREGRADO") & excl_dipl] = "EXCLUIDO_DIPLOMADO"
    estado_carga.loc[(estado_carga == "OK_CARGA_PREGRADO") & sin_match_da] = "EXCLUIDO_SIN_MATCH_DATOS_ALUMNOS"
    # Guardrail operativo: no publicar resoluciones SIES opacas mientras no exista
    # una regla trazable y auditada para reemplazar PRIMERA_OPCION.
    estado_carga.loc[(estado_carga == "OK_CARGA_PREGRADO") & heuristica_sies_opaca] = "EXCLUIDO_SIES_HEURISTICA_OPACA"
    estado_carga.loc[(estado_carga == "OK_CARGA_PREGRADO") & sin_for_ing_trazable] = "EXCLUIDO_SIN_FOR_ING_ACT_TRAZABLE"

    included_final_mask_pre_dedupe = estado_carga == "OK_CARGA_PREGRADO"
    titulado_aprobado_mask = archivo_subida["DA_SITUACION"].astype(str).str.startswith("31 - TITULADO APROBADO")
    # BUG-003 FIX: excluir ELIMINADO/SUSPENDIDO/TITULADO de la política de carga
    # que asumía VIG=1.  Estos estados ya fueron forzados a VIG=0 por regla
    # bloqueante institucional; la política NO debe revertirlos.
    forzar_vig0_mask = archivo_subida["DA_ESTADOACADEMICO"].fillna("").astype(str).str.strip().str.upper().isin(
        ["TITULADO", "ELIMINADO", "SUSPENDIDO"]
    )
    vig_policy_mask = included_final_mask_pre_dedupe & archivo_subida["VIG_FUENTE_FINAL"].eq("REGLA_DEFAULT_1_SIN_FUENTE") & ~forzar_vig0_mask
    archivo_subida.loc[vig_policy_mask & ~titulado_aprobado_mask, "VIG"] = 1
    archivo_subida.loc[vig_policy_mask & ~titulado_aprobado_mask, "VIG_FUENTE_FINAL"] = "POLITICA_CARGA_PREGRADO_INCLUIDA"
    archivo_subida.loc[vig_policy_mask & ~titulado_aprobado_mask, "VIG_METODO_FINAL"] = "REGLA_ESTUDIANTE_CON_MATRICULA_INFORMADA"
    archivo_subida.loc[vig_policy_mask & titulado_aprobado_mask, "VIG"] = 2
    archivo_subida.loc[vig_policy_mask & titulado_aprobado_mask, "VIG_FUENTE_FINAL"] = "POLITICA_CARGA_PREGRADO_INCLUIDA_DA_SITUACION"
    archivo_subida.loc[vig_policy_mask & titulado_aprobado_mask, "VIG_METODO_FINAL"] = "REGLA_EGRESADO_CON_MATRICULA_INFORMADA"
    archivo_subida.loc[vig_policy_mask & ~titulado_aprobado_mask, "VIG_AUDIT_STATUS"] = "POLITICA_CARGA_PREGRADO_VIG_1"
    archivo_subida.loc[vig_policy_mask & titulado_aprobado_mask, "VIG_AUDIT_STATUS"] = "POLITICA_CARGA_PREGRADO_VIG_2_TITULADO"

    # BUG-003 FIX (cont.): trazabilidad y refuerzo para forzados VIG=0
    archivo_subida.loc[forzar_vig0_mask, "VIG"] = 0
    archivo_subida.loc[forzar_vig0_mask, "VIG_FUENTE_FINAL"] = "REGLA_BLOQUEANTE_ESTADOACADEMICO"
    archivo_subida.loc[forzar_vig0_mask, "VIG_METODO_FINAL"] = "FORZADO_VIG0_ELIM_SUSP_TITULADO"
    archivo_subida.loc[forzar_vig0_mask, "VIG_AUDIT_STATUS"] = "VIG0_FORZADO_POR_ESTADO_DA"
    # Re-aplicar zero de 4 columnas para VIG=0 post todas las políticas
    vig_cero_final = pd.to_numeric(archivo_subida["VIG"], errors="coerce").eq(0)
    if vig_cero_final.any():
        archivo_subida.loc[vig_cero_final, "PROM_PRI_SEM"] = 0
        archivo_subida.loc[vig_cero_final, "PROM_SEG_SEM"] = 0
        archivo_subida.loc[vig_cero_final, "ASI_INS_HIS"] = 0
        archivo_subida.loc[vig_cero_final, "ASI_APR_HIS"] = 0
    # Recalcular FLAG_INCONSISTENCIA_VIG después de todas las políticas
    vig_final_post = pd.to_numeric(archivo_subida["VIG"], errors="coerce").astype("Int64")
    esperado_post = pd.to_numeric(archivo_subida["VIG_ESPERADO_DA"], errors="coerce").astype("Int64")
    flag_vig_post = pd.Series("SIN_REGLA_GOB_DA", index=archivo_subida.index, dtype="object")
    flag_vig_post.loc[archivo_subida["DA_ESTADOACADEMICO"].fillna("").astype(str).str.strip().eq("")] = "SIN_ESTADO_DA"
    flag_vig_post.loc[esperado_post.notna() & vig_final_post.eq(esperado_post)] = "OK"
    flag_vig_post.loc[esperado_post.notna() & vig_final_post.ne(esperado_post)] = "INCONSISTENTE"
    archivo_subida["FLAG_INCONSISTENCIA_VIG"] = flag_vig_post

    required_upload = [
        "TIPO_DOC",
        "N_DOC",
        "DV",
        "PRIMER_APELLIDO",
        "NOMBRE",
        "SEXO",
        "FECH_NAC",
        "NAC",
        "PAIS_EST_SEC",
        "COD_SED",
        "COD_CAR",
        "MODALIDAD",
        "JOR",
        "VERSION",
        "FOR_ING_ACT",
        "ANIO_ING_ACT",
        "SEM_ING_ACT",
        "ANIO_ING_ORI",
        "SEM_ING_ORI",
        "ASI_INS_ANT",
        "ASI_APR_ANT",
        "PROM_PRI_SEM",
        "PROM_SEG_SEM",
        "ASI_INS_HIS",
        "ASI_APR_HIS",
        "NIV_ACA",
        "SIT_FON_SOL",
        "SUS_PRE",
        "FECHA_MATRICULA",
        "REINCORPORACION",
        "VIG",
    ]
    required_ok = archivo_subida[required_upload].notna().all(axis=1)
    for required_col in required_upload:
        required_ok &= archivo_subida[required_col].astype(str).str.strip().ne("")
    estado_carga.loc[(estado_carga == "OK_CARGA_PREGRADO") & (~required_ok)] = "EXCLUIDO_CAMPOS_OBLIGATORIOS"

    candidatos = archivo_subida[estado_carga == "OK_CARGA_PREGRADO"].copy()
    candidatos["_FECHA_MAT_TMP"] = pd.to_datetime(candidatos["FECHA_MATRICULA"], errors="coerce", dayfirst=True)

    candidatos, estado_carga, auditoria_consolidacion = _consolidar_candidatos_por_codcli(
        candidatos, estado_carga,
    )

    matricula_unificada_32 = candidatos[MATRICULA_UNIFICADA_COLUMNS].copy()

    # ── Exclusiones por multi-carrera activa (gobernanza institucional) ──
    from scripts.aplicar_exclusiones_multi_carrera import aplicar_exclusiones as _aplicar_exc_mc
    archivo_subida, estado_carga, matricula_unificada_32, _audit_mc = _aplicar_exc_mc(
        archivo_subida, estado_carga, matricula_unificada_32,
        repo_dir=Path(__file__).resolve().parent,
    )

    # El artefacto contractual vigente mantiene FOR_ING_ACT fijo en 1 para carga oficial.
    if not matricula_unificada_32.empty:
        matricula_unificada_32["FOR_ING_ACT"] = pd.Series(1, index=matricula_unificada_32.index, dtype="Int64")

    archivo_subida["ESTADO_CARGA_PREGRADO"] = estado_carga
    archivo_subida["INCLUIR_EN_MATRICULA_32"] = (estado_carga == "OK_CARGA_PREGRADO").map({True: "SI", False: "NO"})

    resumen_carga_pregrado = (
        estado_carga.value_counts(dropna=False)
        .rename_axis("estado_carga")
        .reset_index(name="n")
        .sort_values("n", ascending=False)
    )
    excluidos_carga_pregrado = (
        archivo_subida[archivo_subida["ESTADO_CARGA_PREGRADO"] != "OK_CARGA_PREGRADO"][
            [
                "CODCLI",
                "N_DOC",
                "DV",
                "COD_CAR",
                "NOMBRE_CARRERA_FUENTE",
                "JORNADA_FUENTE",
                "SIES_MATCH_STATUS",
                "SIES_RESOLUCION_HEURISTICA",
                "FOR_ING_ACT_FUENTE_VALOR",
                "FOR_ING_ACT_METODO",
                "FOR_ING_ACT_IMPUTADO",
                "DA_MATCH_MODO",
                "ESTADO_CARGA_PREGRADO",
            ]
        ]
        .copy()
    )
    included_final_mask = archivo_subida["ESTADO_CARGA_PREGRADO"] == "OK_CARGA_PREGRADO"
    uz_cols = ["ASI_INS_ANT", "ASI_APR_ANT", "PROM_PRI_SEM", "PROM_SEG_SEM", "ASI_INS_HIS", "ASI_APR_HIS"]
    uz_all_zero_mask = matricula_unificada_32[uz_cols].eq(0).all(axis=1) if not matricula_unificada_32.empty else pd.Series(dtype=bool)
    for_ing_distribution_final = (
        matricula_unificada_32["FOR_ING_ACT"].astype(str).value_counts(dropna=False).to_dict()
        if not matricula_unificada_32.empty
        else {}
    )

    # ── Trazabilidad DA: aliases con sufijo _DA para auditoría ───────────────
    if "DA_ESTADOACADEMICO" in archivo_subida.columns:
        archivo_subida["ESTADOACADEMICO_DA"] = archivo_subida["DA_ESTADOACADEMICO"]
    if "DA_SITUACION" in archivo_subida.columns:
        archivo_subida["SITUACION_DA"] = archivo_subida["DA_SITUACION"]

    # ── Enriquecimiento duración vía CODIGO_UNICO desde DURACION_ESTUDIOS.tsv ─
    _duracion_tsv_path = Path(__file__).with_name("DURACION_ESTUDIOS.tsv")
    if not _duracion_tsv_path.exists():
        _duracion_tsv_path = Path.cwd() / "DURACION_ESTUDIOS.tsv"
    if _duracion_tsv_path.exists():
        _dur = pd.read_csv(_duracion_tsv_path, sep="\t", dtype=str)
        _dur_cols_map = {
            "NOMBRE_CARRERA": "NOMBRE_CARRERA_TSV",
            "DURACION_ESTUDIOS": "DURACION_ESTUDIOS_TSV",
            "DURACION_TITULACION": "DURACION_TITULACION_TSV",
            "DURACION_TOTAL": "DURACION_TOTAL_TSV",
        }
        _dur_key = "CODIGO_UNICO"
        if _dur_key in _dur.columns and all(c in _dur.columns for c in _dur_cols_map):
            _dur_dedup = _dur.drop_duplicates(subset=[_dur_key], keep="first")
            _dur_idx = _dur_dedup.set_index(_dur_key)
            _sies_key = archivo_subida[FINAL_SIES_CODE_COL].astype(str).str.strip()
            for src_col, dst_col in _dur_cols_map.items():
                archivo_subida[dst_col] = _sies_key.map(_dur_idx[src_col].to_dict())

            # Fallback: para filas con COD_CAR pero sin DURACION_TSV, mapear por CODIGO_CARRERA
            if "CODIGO_CARRERA" in _dur.columns:
                _dur["_COD_CAR_INT"] = pd.to_numeric(_dur["CODIGO_CARRERA"], errors="coerce")
                _dur_by_cod_car = _dur.drop_duplicates(subset=["_COD_CAR_INT"], keep="first")
                _dur_cc_idx = _dur_by_cod_car.set_index("_COD_CAR_INT")
                _needs_dur = archivo_subida["COD_CAR"].notna() & archivo_subida["NOMBRE_CARRERA_TSV"].isna()
                if _needs_dur.any():
                    _cod_car_key = archivo_subida.loc[_needs_dur, "COD_CAR"].astype(float)
                    for src_col, dst_col in _dur_cols_map.items():
                        _mapped = _cod_car_key.map(_dur_cc_idx[src_col].to_dict())
                        archivo_subida.loc[_needs_dur, dst_col] = archivo_subida.loc[
                            _needs_dur, dst_col
                        ].fillna(_mapped)
                    _filled_dur = _needs_dur & archivo_subida["NOMBRE_CARRERA_TSV"].notna()
                    print(f"    ↳ DURACION_TSV fallback por COD_CAR: {int(_filled_dur.sum())} filas")

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / MU_FUSION_OUTPUT_FILENAME
    csv_out_path = output_dir / MU_PREGRADO_CSV_FILENAME

    # ── Enriquecer CATALOGO_MANUAL y SIN_MATCH_DATOS_ALUMNOS con CODIGO_CARRERA ──
    _bridge_map = _build_bridge_codcarpr_to_codcar(df_bridge)
    if not df_manual.empty and _bridge_map:
        df_manual = df_manual.copy()
        df_manual["CODIGO_CARRERA"] = df_manual["CODCARPR"].map(
            lambda x: _bridge_map.get(str(x).strip().upper()) if pd.notna(x) else pd.NA
        )
    if not sin_match_datos_alumnos_df.empty and _bridge_map:
        sin_match_datos_alumnos_df = sin_match_datos_alumnos_df.copy()
        sin_match_datos_alumnos_df["CODIGO_CARRERA"] = sin_match_datos_alumnos_df["COD_CAR_FUENTE"].map(
            lambda x: _bridge_map.get(str(x).strip().upper()) if pd.notna(x) else pd.NA
        )

    # ── Construir hoja REVISION_MANUAL ──
    revision_manual = _build_revision_manual(archivo_subida, sin_match, sin_match_datos_alumnos_df)

    # ── Construir hoja RESUMEN_EJECUTIVO ──
    _codcli_fuente = src[_pick_first_column(src, ["CODCLI"])].astype(str).str.strip()
    _codcli_fuente_unicos = _codcli_fuente.nunique()
    _codcli_fuente_total = len(_codcli_fuente)
    _codcli_archivo = archivo_subida["CODCLI"].astype(str).str.strip()
    _codcli_archivo_unicos = _codcli_archivo.nunique()
    _codcli_mu32 = matricula_unificada_32["CODCLI"].astype(str).str.strip() if "CODCLI" in matricula_unificada_32.columns else pd.Series(dtype=str)
    _codcli_mu32_unicos = _codcli_mu32.nunique() if not _codcli_mu32.empty else 0
    # Desglose estado_carga
    _ec = archivo_subida["ESTADO_CARGA_PREGRADO"]
    _ok = int((_ec == "OK_CARGA_PREGRADO").sum())
    _exc_sin_sies = int((_ec == "EXCLUIDO_SIN_MATCH_SIES").sum())
    _exc_dipl = int((_ec == "EXCLUIDO_DIPLOMADO").sum())
    _exc_sin_da = int((_ec == "EXCLUIDO_SIN_MATCH_DATOS_ALUMNOS").sum())
    _exc_opaca = int((_ec == "EXCLUIDO_SIES_HEURISTICA_OPACA").sum())
    _exc_fig = int((_ec == "EXCLUIDO_SIN_FOR_ING_ACT_TRAZABLE").sum())
    _exc_campos = int((_ec == "EXCLUIDO_CAMPOS_OBLIGATORIOS").sum())
    _exc_dup_intra = int((_ec == "EXCLUIDO_DUPLICADO_INTRA_CODCLI").sum())
    _exc_dup_clave = int((_ec == "EXCLUIDO_DUPLICADO_CLAVE_CARGA").sum())
    _total_excluidos = len(archivo_subida) - _ok
    # Match DatosAlumnos
    _da_modo = archivo_subida["DA_MATCH_MODO"] if "DA_MATCH_MODO" in archivo_subida.columns else pd.Series(dtype=str)
    _match_codcli = int((_da_modo == "MATCH_CODCLI").sum()) if not _da_modo.empty else 0
    _match_rut = int((_da_modo == "MATCH_RUT").sum()) if not _da_modo.empty else 0
    _sin_match_da_n = int((_da_modo == "SIN_MATCH").sum()) if not _da_modo.empty else 0
    # SIES
    _sies_ok = int((archivo_subida["SIES_MATCH_STATUS"] == "MATCH_SIES").sum()) if "SIES_MATCH_STATUS" in archivo_subida.columns else 0
    _sies_ambiguo = int((archivo_subida["SIES_MATCH_STATUS"] == "AMBIGUO_SIES").sum()) if "SIES_MATCH_STATUS" in archivo_subida.columns else 0
    _sies_sin = int((archivo_subida["SIES_MATCH_STATUS"] == "SIN_MATCH_SIES").sum()) if "SIES_MATCH_STATUS" in archivo_subida.columns else 0
    _sies_pend = int((archivo_subida.get("SIES_RESOLUCION_HEURISTICA", pd.Series()) == "PENDIENTE_GOBERNANZA").sum())
    # Métricas COD_CAR por método
    _ccm = archivo_subida.get("COD_CAR_METODO_FINAL", pd.Series(dtype=str))
    _cod_car_total = int(archivo_subida["COD_CAR"].notna().sum())
    _cod_car_sies = int(_ccm.isin(["PARSE_COMPONENTE_SIES", "OFERTA_LOOKUP_CODIGO_UNICO"]).sum())
    _cod_car_shared = int((_ccm == "COMPONENTE_C_COMPARTIDO").sum())
    _cod_car_nombre = int((_ccm == "MAPEO_NOMBRE_CARRERA").sum())
    _cod_car_sin = int(archivo_subida["COD_CAR"].isna().sum())

    resumen_ejecutivo_rows = [
        {"seccion": "ENTRADA", "metrica": "CODCLI total filas fuente (Promedios)", "valor": _codcli_fuente_total, "pct": "100.0%"},
        {"seccion": "ENTRADA", "metrica": "CODCLI únicos en fuente", "valor": _codcli_fuente_unicos, "pct": f"{_codcli_fuente_unicos/_codcli_fuente_total*100:.1f}%"},
        {"seccion": "PROCESAMIENTO", "metrica": "CODCLI en ARCHIVO_LISTO_SUBIDA", "valor": len(archivo_subida), "pct": f"{len(archivo_subida)/_codcli_fuente_total*100:.1f}%"},
        {"seccion": "PROCESAMIENTO", "metrica": "CODCLI únicos en ARCHIVO_LISTO_SUBIDA", "valor": _codcli_archivo_unicos, "pct": ""},
        {"seccion": "MATCH_DATOS_ALUMNOS", "metrica": "Match por CODCLI", "valor": _match_codcli, "pct": f"{_match_codcli/len(archivo_subida)*100:.1f}%"},
        {"seccion": "MATCH_DATOS_ALUMNOS", "metrica": "Match por RUT (fallback)", "valor": _match_rut, "pct": f"{_match_rut/len(archivo_subida)*100:.1f}%"},
        {"seccion": "MATCH_DATOS_ALUMNOS", "metrica": "Sin match DatosAlumnos", "valor": _sin_match_da_n, "pct": f"{_sin_match_da_n/len(archivo_subida)*100:.1f}%"},
        {"seccion": "MATCH_SIES", "metrica": "MATCH_SIES (código único)", "valor": _sies_ok, "pct": f"{_sies_ok/len(archivo_subida)*100:.1f}%"},
        {"seccion": "MATCH_SIES", "metrica": "AMBIGUO_SIES (>1 código)", "valor": _sies_ambiguo, "pct": f"{_sies_ambiguo/len(archivo_subida)*100:.1f}%"},
        {"seccion": "MATCH_SIES", "metrica": "SIN_MATCH_SIES", "valor": _sies_sin, "pct": f"{_sies_sin/len(archivo_subida)*100:.1f}%"},
        {"seccion": "MATCH_SIES", "metrica": "PENDIENTE_GOBERNANZA (sin resolver)", "valor": _sies_pend, "pct": f"{_sies_pend/len(archivo_subida)*100:.1f}%"},
        {"seccion": "COD_CAR", "metrica": "COD_CAR asignado (total)", "valor": _cod_car_total, "pct": f"{_cod_car_total/len(archivo_subida)*100:.1f}%"},
        {"seccion": "COD_CAR", "metrica": "  → por SIES directo/oferta", "valor": _cod_car_sies, "pct": f"{_cod_car_sies/len(archivo_subida)*100:.1f}%"},
        {"seccion": "COD_CAR", "metrica": "  → por COMPONENTE_C_COMPARTIDO", "valor": _cod_car_shared, "pct": f"{_cod_car_shared/len(archivo_subida)*100:.1f}%"},
        {"seccion": "COD_CAR", "metrica": "  → por MAPEO_NOMBRE_CARRERA", "valor": _cod_car_nombre, "pct": f"{_cod_car_nombre/len(archivo_subida)*100:.1f}%"},
        {"seccion": "COD_CAR", "metrica": "Sin COD_CAR", "valor": _cod_car_sin, "pct": f"{_cod_car_sin/len(archivo_subida)*100:.1f}%"},
        {"seccion": "CARGA_FINAL", "metrica": "OK_CARGA_PREGRADO (→ MU32)", "valor": _ok, "pct": f"{_ok/len(archivo_subida)*100:.1f}%"},
        {"seccion": "CARGA_FINAL", "metrica": "CODCLI únicos en MATRICULA_UNIFICADA_32", "valor": _codcli_mu32_unicos, "pct": ""},
        {"seccion": "EXCLUIDOS", "metrica": "Total excluidos", "valor": _total_excluidos, "pct": f"{_total_excluidos/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_SIN_MATCH_SIES", "valor": _exc_sin_sies, "pct": f"{_exc_sin_sies/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_DIPLOMADO", "valor": _exc_dipl, "pct": f"{_exc_dipl/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_SIN_MATCH_DATOS_ALUMNOS", "valor": _exc_sin_da, "pct": f"{_exc_sin_da/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_SIES_HEURISTICA_OPACA", "valor": _exc_opaca, "pct": f"{_exc_opaca/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_SIN_FOR_ING_ACT_TRAZABLE", "valor": _exc_fig, "pct": f"{_exc_fig/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_CAMPOS_OBLIGATORIOS", "valor": _exc_campos, "pct": f"{_exc_campos/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_DUPLICADO_INTRA_CODCLI", "valor": _exc_dup_intra, "pct": f"{_exc_dup_intra/len(archivo_subida)*100:.1f}%"},
        {"seccion": "EXCLUIDOS", "metrica": "  → EXCLUIDO_DUPLICADO_CLAVE_CARGA", "valor": _exc_dup_clave, "pct": f"{_exc_dup_clave/len(archivo_subida)*100:.1f}%"},
        {"seccion": "VERIFICACION", "metrica": "Suma excluidos + OK = archivo_subida", "valor": _total_excluidos + _ok, "pct": "✅" if (_total_excluidos + _ok) == len(archivo_subida) else "❌ DESCUADRE"},
    ]
    resumen_ejecutivo = pd.DataFrame(resumen_ejecutivo_rows)

    # ── Máscara rojo para PENDIENTE_GOBERNANZA en ARCHIVO_LISTO_SUBIDA ──
    _red_mask = archivo_subida["SIES_RESOLUCION_HEURISTICA"] == "PENDIENTE_GOBERNANZA" if "SIES_RESOLUCION_HEURISTICA" in archivo_subida.columns else pd.Series(False, index=archivo_subida.index)

    sheets_export: dict[str, pd.DataFrame] = {}
    sheets_export["RESUMEN_EJECUTIVO"] = resumen_ejecutivo
    sheets_export["REVISION_MANUAL"] = revision_manual
    sheets_export["MATRICULA_UNIFICADA_32"] = matricula_unificada_32
    sheets_export["ARCHIVO_LISTO_SUBIDA"] = archivo_subida
    sheets_export["RESUMEN_MU"] = resumen
    sheets_export["RESUMEN_MANUAL"] = resumen_manual
    sheets_export["RESUMEN_SIES"] = resumen_sies
    sheets_export["RESUMEN_CARGA_PREGRADO"] = resumen_carga_pregrado
    sheets_export["EXCLUIDOS_CARGA_PREGR"] = excluidos_carga_pregrado
    sheets_export["SIES_AMBIGUOS_POR_RESOL"] = ambiguos
    sheets_export["SIN_MATCH_SIES"] = sin_match
    if not df_manual.empty:
        sheets_export["CATALOGO_MANUAL"] = df_manual
    if not df_bridge.empty:
        sheets_export["PUENTE_SIES"] = df_bridge
    if usar_gobernanza_v2:
        sheets_export["SIN_MATCH_DATOS_ALUMNOS"] = sin_match_datos_alumnos_df
    if not auditoria_consolidacion.empty:
        sheets_export["AUDITORIA_CONSOLIDACION"] = auditoria_consolidacion
    _write_excel_atomic(sheets_export, out_path, red_rows_sheet="ARCHIVO_LISTO_SUBIDA", red_rows_mask=_red_mask)
    _write_mu_csv_atomic(matricula_unificada_32, csv_out_path)

    if not auditoria_consolidacion.empty:
        audit_tsv_path = output_dir / "auditoria_consolidacion_codcli.tsv"
        auditoria_consolidacion.to_csv(audit_tsv_path, sep="\t", index=False)

    _report = {
        "output_file": str(out_path),
        "csv_output_file": str(csv_out_path),
        "sheet_used": selected_sheet,
        "gobernanza_mode": "v2_flagged" if usar_gobernanza_v2 else "legacy_default",
        "usar_gobernanza_v2": bool(usar_gobernanza_v2),
        "rows": len(archivo_subida),
        "rows_matricula_32_final": len(matricula_unificada_32),
        "rows_excluidas_carga_pregrado": int((archivo_subida["ESTADO_CARGA_PREGRADO"] != "OK_CARGA_PREGRADO").sum()),
        "rows_excluidas_sies_heuristica_opaca": int((archivo_subida["ESTADO_CARGA_PREGRADO"] == "EXCLUIDO_SIES_HEURISTICA_OPACA").sum()),
        "rows_excluidas_sin_for_ing_act_trazable": int((archivo_subida["ESTADO_CARGA_PREGRADO"] == "EXCLUIDO_SIN_FOR_ING_ACT_TRAZABLE").sum()),
        "rows_enriquecidas_datos_alumnos": rows_enriquecidas_datos_alumnos,
        "rows_sin_match_datos_alumnos": sin_match_datos_alumnos_rows,
        "cod_sed_resuelto_por_regla": cod_sed_resueltos_regla,
        "pais_est_sec_inferido_localidad": pais_est_sec_inferidos_localidad,
        "for_ing_act_distribution_final": for_ing_distribution_final,
        "rows_for_ing_act_imputado_final": int((included_final_mask & archivo_subida["FOR_ING_ACT_IMPUTADO"].eq("SI")).sum()),
        "rows_for_ing_act_revision_final": int((included_final_mask & archivo_subida["FOR_ING_ACT_REQUIERE_REVISION"].eq("SI")).sum()),
        "rows_uz_all_zero_final": int(uz_all_zero_mask.sum()) if not matricula_unificada_32.empty else 0,
        "historico_mu_anio_referencia": anio_ref_historico_mu,
        "periodo_filtro_anio": int(periodo_filtro_anio),
        "periodo_filtro_sem": int(periodo_filtro_sem),
        "anio_anterior_prom": int(anio_anterior_prom),
        "estado_inicial": estado_inicial.value_counts(dropna=False).to_dict(),
        "manual_match": resumen_manual.set_index("estado")["n"].to_dict(),
        "sies_diag": resumen_sies.set_index("estado")["n"].to_dict(),
        "catalogo_manual_rows": len(df_manual),
        "puente_sies_rows": len(df_bridge),
        "gob_nac_rows": len(gob_nac_df),
        "gob_pais_est_sec_rows": len(gob_pais_est_sec_df),
        "gob_sede_rows": len(gob_sede_df),
        "oferta_academica_rows": len(oferta_dim),
        "catalogo_manual_source": manual_source,
        "puente_sies_source": puente_source,
        "gob_nac_source": gob_nac_tsv_path or "no_file",
        "gob_pais_est_sec_source": gob_pais_est_sec_tsv_path or "no_file",
        "gob_sede_source": gob_sede_tsv_path or "no_file",
        "gob_for_ing_act_source": gob_for_ing_act_source,
        "oferta_academica_source": oferta_source,
    }
    if _filtro_bd_stats:
        _report["filtro_base_datos"] = _filtro_bd_stats
    if _stats_depur:
        _report["depuracion_rut_multi_codcli"] = _stats_depur
    # Persistir JSON del pipeline de matrícula para trazabilidad
    _mu_json_path = output_dir / "reporte_matricula.json"
    try:
        _mu_json_path.write_text(json.dumps(_report, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass  # no bloquear pipeline por fallo de escritura JSON
    return _report


# ==============================
# CAPA C: Validación + Export
# ==============================
def _check_schema(df: pd.DataFrame, expected: list[str], area: str, issues: list[Issue]) -> None:
    if df.columns.tolist() != expected:
        issues.append(Issue("ERROR", area, "Schema exacto incumplido"))


def validar_carreras(carr: pd.DataFrame) -> list[Issue]:
    issues: list[Issue] = []
    _check_schema(carr, CARRERAS_AC_COLUMNS, "carreras", issues)

    if carr["PLAN_ESTUDIOS"].isna().any():
        issues.append(Issue("ERROR", "carreras", "PLAN_ESTUDIOS vacío", int(carr["PLAN_ESTUDIOS"].isna().sum())))
    bad_tum = ~carr["TIPO_UNIDAD_MEDIDA"].astype(str).isin(["1", "2", "3"]) & carr["TIPO_UNIDAD_MEDIDA"].notna()
    if bad_tum.any():
        issues.append(Issue("ERROR", "carreras", "TIPO_UNIDAD_MEDIDA fuera de catálogo", int(bad_tum.sum())))
    tum3 = carr["TIPO_UNIDAD_MEDIDA"].astype(str) == "3"
    if tum3.any() and carr.loc[tum3, "OTRA_UNIDAD_MEDIDA"].isna().any():
        issues.append(
            Issue(
                "ERROR",
                "carreras",
                "OTRA_UNIDAD_MEDIDA faltante cuando TIPO_UNIDAD_MEDIDA=3",
                int(carr.loc[tum3, "OTRA_UNIDAD_MEDIDA"].isna().sum()),
            )
        )

    annual = carr[ANUAL_COLS].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    total = pd.to_numeric(carr["TOTAL_UNIDADES_MEDIDA"], errors="coerce")
    mismatch = total.notna() & (annual != total)
    if mismatch.any():
        issues.append(Issue("WARN", "carreras", "Suma de unidades anuales distinta al total", int(mismatch.sum())))

    vig_bad = ~carr["VIGENCIA"].astype(str).isin(["0", "1"]) & carr["VIGENCIA"].notna()
    if vig_bad.any():
        issues.append(Issue("ERROR", "carreras", "VIGENCIA fuera de catálogo 0/1", int(vig_bad.sum())))
    return issues


def validar_matricula_ac(mat: pd.DataFrame, carr: pd.DataFrame) -> list[Issue]:
    issues: list[Issue] = []
    _check_schema(mat, MATRICULA_AC_COLUMNS, "matricula_ac", issues)

    for c in ["TIPO_DOCUMENTO", "NUM_DOCUMENTO", "DV", "CODIGO_UNICO", "PLAN_ESTUDIOS"]:
        if mat[c].isna().any():
            issues.append(Issue("ERROR", "matricula_ac", f"{c} vacío", int(mat[c].isna().sum())))

    bad_td = ~mat["TIPO_DOCUMENTO"].astype(str).isin(["R", "P"]) & mat["TIPO_DOCUMENTO"].notna()
    if bad_td.any():
        issues.append(Issue("ERROR", "matricula_ac", "TIPO_DOCUMENTO fuera de catálogo R/P", int(bad_td.sum())))
    bad_sex = ~mat["SEXO"].astype(str).isin(["M", "H", "X"]) & mat["SEXO"].notna()
    if bad_sex.any():
        issues.append(Issue("WARN", "matricula_ac", "SEXO fuera de catálogo M/H/X", int(bad_sex.sum())))

    curso1 = mat["CURSO_1ER_SEM"].astype(str).str.strip().str.upper()
    curso2 = mat["CURSO_2DO_SEM"].astype(str).str.strip().str.upper()
    bad_curso = ~curso1.isin(["SI", "NO"]) & mat["CURSO_1ER_SEM"].notna()
    bad_curso2 = ~curso2.isin(["SI", "NO"]) & mat["CURSO_2DO_SEM"].notna()
    if bad_curso.any() or bad_curso2.any():
        issues.append(
            Issue(
                "ERROR",
                "matricula_ac",
                "CURSO_1ER_SEM/CURSO_2DO_SEM fuera de catálogo SI/NO",
                int(bad_curso.sum() + bad_curso2.sum()),
            )
        )

    sem1 = pd.to_numeric(mat["SEM_INGRESO_CARRERA_ACTUAL"], errors="coerce")
    sem2 = pd.to_numeric(mat["SEM_INGRESO_CARRERA_ORIGEN"], errors="coerce")
    bad_sem = (~sem1.isin([1, 2]) & sem1.notna()) | (~sem2.isin([1, 2]) & sem2.notna())
    if bad_sem.any():
        issues.append(Issue("WARN", "matricula_ac", "Semestres de ingreso fuera de 1/2", int(bad_sem.sum())))

    apr_gt_cur = pd.to_numeric(mat["UNIDADES_APROBADAS"], errors="coerce") > pd.to_numeric(
        mat["UNIDADES_CURSADAS"], errors="coerce"
    )
    if apr_gt_cur.any():
        issues.append(Issue("WARN", "matricula_ac", "UNIDADES_APROBADAS > UNIDADES_CURSADAS", int(apr_gt_cur.sum())))

    aprt_gt_curt = pd.to_numeric(mat["UNID_APROBADAS_TOTAL"], errors="coerce") > pd.to_numeric(
        mat["UNID_CURSADAS_TOTAL"], errors="coerce"
    )
    if aprt_gt_curt.any():
        issues.append(Issue("WARN", "matricula_ac", "UNID_APROBADAS_TOTAL > UNID_CURSADAS_TOTAL", int(aprt_gt_curt.sum())))

    keys_c = set(map(tuple, carr[["CODIGO_UNICO", "PLAN_ESTUDIOS"]].dropna().drop_duplicates().to_records(index=False)))
    keys_m = set(map(tuple, mat[["CODIGO_UNICO", "PLAN_ESTUDIOS"]].dropna().drop_duplicates().to_records(index=False)))
    if keys_m - keys_c:
        issues.append(Issue("ERROR", "matricula_ac", "PLAN_ESTUDIOS no referenciado en carreras", len(keys_m - keys_c)))
    return issues


def validar_matricula_unificada(mu: pd.DataFrame) -> list[Issue]:
    issues: list[Issue] = []
    _check_schema(mu, MATRICULA_UNIFICADA_COLUMNS, "matricula_unificada", issues)

    # Manual 2026 (Anexo 7, Cuadro N°1): VIG permite 0/1/2.
    vig_text = mu["VIG"].astype(str).str.strip()
    vig_num = pd.to_numeric(mu["VIG"], errors="coerce")
    vig_bad = mu["VIG"].notna() & ~(vig_text.isin(["0", "1", "2"]) | vig_num.isin([0, 1, 2]))
    if vig_bad.any():
        issues.append(Issue("ERROR", "matricula_unificada", "VIG fuera de catálogo 0/1/2", int(vig_bad.sum())))

    reinc_bad = ~_is_binary_valid(mu["REINCORPORACION"]) & mu["REINCORPORACION"].notna()
    if reinc_bad.any():
        issues.append(Issue("ERROR", "matricula_unificada", "REINCORPORACION fuera de 0/1", int(reinc_bad.sum())))
    if mu["REINCORPORACION"].isna().any():
        issues.append(
            Issue(
                "BLOCKER",
                "matricula_unificada",
                "REINCORPORACION obligatorio vacío",
                int(mu["REINCORPORACION"].isna().sum()),
            )
        )

    if mu["PAIS_EST_SEC"].isna().any():
        issues.append(
            Issue("BLOCKER", "matricula_unificada", "PAIS_EST_SEC obligatorio vacío", int(mu["PAIS_EST_SEC"].isna().sum()))
        )

    # Manual 2026 (Anexo 7, Cuadro N°1): campos obligatorios y rangos explícitos.
    required_manual = [
        "ASI_INS_ANT",
        "ASI_APR_ANT",
        "PROM_PRI_SEM",
        "PROM_SEG_SEM",
        "ASI_INS_HIS",
        "ASI_APR_HIS",
        "NIV_ACA",
        "SIT_FON_SOL",
        "SUS_PRE",
    ]
    for col in required_manual:
        if mu[col].isna().any():
            issues.append(Issue("BLOCKER", "matricula_unificada", f"{col} obligatorio vacío", int(mu[col].isna().sum())))

    asi_ins_ant = pd.to_numeric(mu["ASI_INS_ANT"], errors="coerce")
    asi_apr_ant = pd.to_numeric(mu["ASI_APR_ANT"], errors="coerce")
    prom_pri = pd.to_numeric(mu["PROM_PRI_SEM"], errors="coerce")
    prom_seg = pd.to_numeric(mu["PROM_SEG_SEM"], errors="coerce")
    asi_ins_his = pd.to_numeric(mu["ASI_INS_HIS"], errors="coerce")
    asi_apr_his = pd.to_numeric(mu["ASI_APR_HIS"], errors="coerce")
    niv_aca = pd.to_numeric(mu["NIV_ACA"], errors="coerce")
    sit_fon = pd.to_numeric(mu["SIT_FON_SOL"], errors="coerce")
    sus_pre = pd.to_numeric(mu["SUS_PRE"], errors="coerce")

    bad_asi_ins_ant = mu["ASI_INS_ANT"].notna() & ~asi_ins_ant.between(0, 99)
    if bad_asi_ins_ant.any():
        issues.append(Issue("ERROR", "matricula_unificada", "ASI_INS_ANT fuera de rango 0..99", int(bad_asi_ins_ant.sum())))

    bad_asi_apr_ant = mu["ASI_APR_ANT"].notna() & ~asi_apr_ant.between(0, 99)
    if bad_asi_apr_ant.any():
        issues.append(Issue("ERROR", "matricula_unificada", "ASI_APR_ANT fuera de rango 0..99", int(bad_asi_apr_ant.sum())))

    bad_prom_pri = mu["PROM_PRI_SEM"].notna() & ~((prom_pri == 0) | prom_pri.between(100, 700))
    if bad_prom_pri.any():
        issues.append(
            Issue("ERROR", "matricula_unificada", "PROM_PRI_SEM fuera de rango permitido (0 o 100..700)", int(bad_prom_pri.sum()))
        )

    bad_prom_seg = mu["PROM_SEG_SEM"].notna() & ~((prom_seg == 0) | prom_seg.between(100, 700))
    if bad_prom_seg.any():
        issues.append(
            Issue("ERROR", "matricula_unificada", "PROM_SEG_SEM fuera de rango permitido (0 o 100..700)", int(bad_prom_seg.sum()))
        )

    bad_asi_ins_his = mu["ASI_INS_HIS"].notna() & ~asi_ins_his.between(0, 200)
    if bad_asi_ins_his.any():
        issues.append(Issue("ERROR", "matricula_unificada", "ASI_INS_HIS fuera de rango 0..200", int(bad_asi_ins_his.sum())))

    bad_asi_apr_his = mu["ASI_APR_HIS"].notna() & ~asi_apr_his.between(0, 200)
    if bad_asi_apr_his.any():
        issues.append(Issue("ERROR", "matricula_unificada", "ASI_APR_HIS fuera de rango 0..200", int(bad_asi_apr_his.sum())))

    bad_niv_aca = mu["NIV_ACA"].notna() & ~(niv_aca >= 1)
    if bad_niv_aca.any():
        issues.append(Issue("ERROR", "matricula_unificada", "NIV_ACA debe ser >= 1", int(bad_niv_aca.sum())))

    bad_sit_fon = mu["SIT_FON_SOL"].notna() & ~sit_fon.isin([0, 1, 2])
    if bad_sit_fon.any():
        issues.append(Issue("ERROR", "matricula_unificada", "SIT_FON_SOL fuera de catálogo 0/1/2", int(bad_sit_fon.sum())))

    bad_sus_pre = mu["SUS_PRE"].notna() & ~sus_pre.between(0, 99)
    if bad_sus_pre.any():
        issues.append(Issue("ERROR", "matricula_unificada", "SUS_PRE fuera de rango 0..99", int(bad_sus_pre.sum())))

    apr_ant_gt_ins_ant = asi_apr_ant.notna() & asi_ins_ant.notna() & (asi_apr_ant > asi_ins_ant)
    if apr_ant_gt_ins_ant.any():
        issues.append(
            Issue(
                "ERROR",
                "matricula_unificada",
                "ASI_APR_ANT no puede ser mayor que ASI_INS_ANT",
                int(apr_ant_gt_ins_ant.sum()),
            )
        )

    apr_his_gt_ins_his = asi_apr_his.notna() & asi_ins_his.notna() & (asi_apr_his > asi_ins_his)
    if apr_his_gt_ins_his.any():
        issues.append(
            Issue(
                "ERROR",
                "matricula_unificada",
                "ASI_APR_HIS no puede ser mayor que ASI_INS_HIS",
                int(apr_his_gt_ins_his.sum()),
            )
        )

    fechas = pd.to_datetime(mu["FECHA_MATRICULA"], errors="coerce")
    futuras = fechas.notna() & (fechas > pd.Timestamp.today().normalize())
    if futuras.any():
        issues.append(Issue("ERROR", "matricula_unificada", "FECHA_MATRICULA posterior a fecha de carga", int(futuras.sum())))

    mod_missing = mu["COD_CAR"].notna() & mu["MODALIDAD"].isna()
    if mod_missing.any():
        issues.append(
            Issue(
                "WARN",
                "matricula_unificada",
                "COD_CAR con MODALIDAD no resuelta (equivalencia dudosa)",
                int(mod_missing.sum()),
            )
        )
    
    # Contar ambigüedades pendientes
    ambiguas = mu["_SIES_AMBIGUO"].sum() if "_SIES_AMBIGUO" in mu.columns else 0
    if ambiguas > 0:
        issues.append(Issue(
            "WARN", 
            "matricula_unificada", 
            f"Registros con ambigüedad SIES sin resolver (Fase 2 requerida)",
            int(ambiguas)
        ))
    
    return issues


def exportar_control_y_pes(
    df: pd.DataFrame, control_path: Path, pes_path: Path, issues: list[Issue], area: str
) -> None:
    df.to_csv(control_path, index=False, encoding="utf-8")
    if "CODIGO_IES_NUM" not in df.columns:
        issues.append(Issue("ERROR", area, "No existe CODIGO_IES_NUM para generar pes_ready"))
        return
    pes = df.drop(columns=["CODIGO_IES_NUM"])
    pes.to_csv(pes_path, index=False, header=False, encoding="utf-8")


def _profile_column(series: pd.Series) -> dict[str, float]:
    s = series.astype(object)
    null_pct = float(s.isna().mean())
    zero_pct = float((pd.to_numeric(s, errors="coerce") == 0).fillna(False).mean())
    default_like_pct = float(s.fillna("").astype(str).str.strip().isin(["", "0", "NO", "N/A", "NA"]).mean())
    return {
        "null_pct": round(null_pct, 4),
        "zero_pct": round(zero_pct, 4),
        "default_like_pct": round(default_like_pct, 4),
    }


def _is_binary_valid(series: pd.Series) -> pd.Series:
    s = series.astype(object)
    as_text = s.astype(str).str.strip()
    as_num = pd.to_numeric(s, errors="coerce")
    valid = as_text.isin(["0", "1"]) | as_num.isin([0, 1])
    return valid | s.isna()


def resolver_ambiguedad_sies(codcarpr: str, jornada: str, version: str = "V1") -> tuple:
    """Resuelve código SIES usando matriz de desambiguación.
    
    Args:
        codcarpr: Código carrera (IINF, ICRE, ICIB, etc.)
        jornada: Jornada (D, V, O)
        version: Versión plan (V1, V2, V3, V4) - default V1
    
    Returns:
        (codigo_sies, confianza, notas, es_ambiguo)
        - codigo_sies: Código SIES resuelto o None
        - confianza: Porcentaje (100%, 95%, etc.)
        - notas: Razón del mapeo
        - es_ambiguo: True si no está en matriz (requiere revisión)
    """
    if not codcarpr or not jornada:
        return (None, "0%", "Parámetros incompletos", True)
    
    key = (str(codcarpr).strip().upper(), str(jornada).strip().upper(), str(version).strip().upper())
    
    if key in MATRIZ_DESAMBIGUACION:
        sies, conf, notas = MATRIZ_DESAMBIGUACION[key]
        return (sies, conf, notas, False)
    else:
        return (None, "0%", f"No encontrado en matriz: ({codcarpr}, {jornada}, {version})", True)


def _build_oferta_index(oferta_dim: pd.DataFrame) -> dict:
    """Construye índice CODIGO_UNICO → atributos desde oferta_dim."""
    idx = {}
    if oferta_dim.empty:
        return idx
    for _, row in oferta_dim.iterrows():
        cu = str(row.get("CODIGO_UNICO", "")).strip().upper()
        if not cu:
            continue
        idx[cu] = {
            "TIPO_PLAN_CARRERA": row.get("TIPO_PLAN_CARRERA"),
            "JORNADA": row.get("JORNADA"),
            "DURACION_ESTUDIOS": row.get("DURACION_ESTUDIOS"),
            "MODALIDAD": row.get("MODALIDAD"),
            "CODIGO_CARRERA": row.get("CODIGO_CARRERA"),
            "NIVEL_CARRERA": row.get("NIVEL_CARRERA"),
        }
    return idx


def _load_cuadro_homologacion(input_file: Path) -> dict:
    """Carga CUADRO HOMOLOGACIÓN → dict (CODCARPR, JORNADA_DA) → CODIGO_SIES."""
    homol: dict[tuple[str, str], str] = {}
    try:
        xls = pd.ExcelFile(input_file)
        target = None
        for sheet in xls.sheet_names:
            if "HOMOLOG" in sheet.upper():
                target = sheet
                break
        if target is None:
            return homol
        hm = pd.read_excel(input_file, sheet_name=target)
        for _, row in hm.iterrows():
            codcarpr = str(row.get("CODCARPR", "")).strip().upper()
            jornada_da = str(row.get("JORNADA_DA", "")).strip().upper()
            codigo_sies = str(row.get("CODIGO_SIES", "")).strip().upper()
            if codcarpr and jornada_da and codigo_sies:
                homol[(codcarpr, jornada_da)] = codigo_sies
    except Exception as e:
        print(f"⚠️  No se pudo cargar CUADRO HOMOLOGACIÓN: {e}")
    return homol


def _resolver_ambiguedades_sies_heuristica(
    ambiguos_df: pd.DataFrame,
    oferta_idx: dict | None = None,
    homol_dict: dict | None = None,
) -> pd.DataFrame:
    """Resuelve ambigüedades SIES usando una cascada trazable y auditada.

    Lógica:
    - Entrada: CODCARPR_NORM, JORNADA_FUENTE, CODIGOS_SIES_POTENCIALES
    - Resolución por cascada: TIPO_PLAN_CARRERA → JORNADA → HOMOLOGACIÓN
    - Sin uso de PRIMERA_OPCION ni VERSION como input primario

    Args:
        ambiguos_df: DataFrame con registros SIES_MATCH_STATUS == "AMBIGUO_SIES"
        oferta_idx: dict CODIGO_UNICO → {TIPO_PLAN_CARRERA, JORNADA, ...}
        homol_dict: dict (CODCARPR, JORNADA_DA_LETRA) → CODIGO_SIES

    Returns:
        DataFrame con ambigüedades resueltas o marcadas como PENDIENTE_GOBERNANZA
    """
    if oferta_idx is None:
        oferta_idx = {}
    if homol_dict is None:
        homol_dict = {}

    if ambiguos_df.empty:
        ambiguos_df["SIES_RESOLUCION_HEURISTICA"] = pd.NA
        ambiguos_df["SIES_CONFIANZA_POST"] = pd.NA
        return ambiguos_df

    result = ambiguos_df.copy()

    if "SIES_RESOLUCION_HEURISTICA" not in result.columns:
        result["SIES_RESOLUCION_HEURISTICA"] = pd.NA
    if "SIES_CONFIANZA_POST" not in result.columns:
        result["SIES_CONFIANZA_POST"] = pd.NA

    # Mapeo de jornada fuente (letra o texto) a numérico SIES y a letra normalizada
    _JOR_TO_SIES = {"D": 1, "V": 2, "O": 4, "1": 1, "2": 2, "4": 4,
                     "DIURNA": 1, "VESPERTINA": 2, "A DISTANCIA": 4, "DISTANCIA": 4, "ONLINE": 4}
    _JOR_TO_LETTER = {"D": "D", "V": "V", "O": "O", "1": "D", "2": "V", "4": "O",
                       "DIURNA": "D", "VESPERTINA": "V", "A DISTANCIA": "O", "DISTANCIA": "O", "ONLINE": "O"}

    for idx, row in result.iterrows():
        codcarpr = str(row.get("CODCARPR_NORM", "")).strip().upper()
        # Preferir JORNADA_FUENTE (letra original) sobre JOR (numérico post-pipeline)
        jornada_raw = str(row.get("JORNADA_FUENTE", "") or row.get("JOR", "")).strip().upper()
        jornada_sies = _JOR_TO_SIES.get(jornada_raw)
        jornada_letra = _JOR_TO_LETTER.get(jornada_raw, jornada_raw)
        candidatos_raw = row.get("CODIGOS_SIES_POTENCIALES", "")

        # Parsear candidatos desde CODIGOS_SIES_POTENCIALES o fallback a _1.._5
        if pd.notna(candidatos_raw) and str(candidatos_raw).strip():
            candidatos = [c.strip() for c in str(candidatos_raw).split(" | ")]
        else:
            candidatos = [row.get(f"CODIGO_CARRERA_SIES_{i}") for i in range(1, 6)]
            candidatos = [str(c).strip() for c in candidatos if pd.notna(c)]

        if not candidatos:
            result.at[idx, "SIES_RESOLUCION_HEURISTICA"] = "PENDIENTE_GOBERNANZA"
            result.at[idx, "SIES_CONFIANZA_POST"] = "0%"
            continue

        # Cargar atributos canónicos desde oferta; comparar como int para robustez
        atributos = []
        for codigo in candidatos:
            oferta = oferta_idx.get(codigo, {})
            tp_val = oferta.get("TIPO_PLAN_CARRERA")
            jor_val = oferta.get("JORNADA")
            atributos.append({
                "codigo": codigo,
                "tp": int(tp_val) if pd.notna(tp_val) else None,
                "jor": int(jor_val) if pd.notna(jor_val) else None,
                "dur": oferta.get("DURACION_ESTUDIOS"),
                "niv": oferta.get("NIVEL_CARRERA"),
                "mod": oferta.get("MODALIDAD"),
            })

        # Paso 1: Filtrar por TIPO_PLAN_CARRERA esperado
        tp_esperado = 3 if codcarpr[:2] in {"CI", "CO", "CA", "CN"} else 1
        filtrados_tp = [a for a in atributos if a["tp"] == tp_esperado]
        if len(filtrados_tp) == 1:
            result.at[idx, FINAL_SIES_CODE_COL] = filtrados_tp[0]["codigo"]
            result.at[idx, "SIES_MATCH_STATUS"] = "MATCH_SIES"
            result.at[idx, "SIES_RESOLUCION_HEURISTICA"] = "REGLA_TIPO_PLAN"
            result.at[idx, "SIES_CONFIANZA_POST"] = "95%"
            continue
        # Si tp no redujo a 0, usar los filtrados; si redujo a 0, conservar todos
        residuales = filtrados_tp if filtrados_tp else atributos

        # Paso 2: Filtrar por JORNADA esperada (numérica SIES)
        if jornada_sies is not None:
            filtrados_jor = [a for a in residuales if a["jor"] == jornada_sies]
        else:
            filtrados_jor = []
        if len(filtrados_jor) == 1:
            result.at[idx, FINAL_SIES_CODE_COL] = filtrados_jor[0]["codigo"]
            result.at[idx, "SIES_MATCH_STATUS"] = "MATCH_SIES"
            result.at[idx, "SIES_RESOLUCION_HEURISTICA"] = "REGLA_TIPO_PLAN_JORNADA"
            result.at[idx, "SIES_CONFIANZA_POST"] = "95%"
            continue
        residuales = filtrados_jor if filtrados_jor else residuales

        # Paso 3: Consultar homologación (CODCARPR, JORNADA_DA_LETRA) → CODIGO_SIES
        homologado = homol_dict.get((codcarpr, jornada_letra))
        if homologado:
            # Verificar contra residuales primero, luego contra todos los candidatos
            if any(a["codigo"] == homologado for a in residuales):
                result.at[idx, FINAL_SIES_CODE_COL] = homologado
                result.at[idx, "SIES_MATCH_STATUS"] = "MATCH_SIES"
                result.at[idx, "SIES_RESOLUCION_HEURISTICA"] = "REGLA_HOMOLOGACION"
                result.at[idx, "SIES_CONFIANZA_POST"] = "99%"
                continue
            if any(a["codigo"] == homologado for a in atributos):
                result.at[idx, FINAL_SIES_CODE_COL] = homologado
                result.at[idx, "SIES_MATCH_STATUS"] = "MATCH_SIES"
                result.at[idx, "SIES_RESOLUCION_HEURISTICA"] = "REGLA_HOMOLOGACION"
                result.at[idx, "SIES_CONFIANZA_POST"] = "95%"
                continue

        # Caso no resuelto — nunca PRIMERA_OPCION
        result.at[idx, "SIES_RESOLUCION_HEURISTICA"] = "PENDIENTE_GOBERNANZA"
        result.at[idx, "SIES_CONFIANZA_POST"] = "0%"

    return result


def generar_procedencia_y_calidad(
    output_dir: Path, mu: pd.DataFrame, ca: pd.DataFrame, ma: pd.DataFrame, issues: list[Issue]
) -> dict[str, object]:
    provenance_rules = {
        "matricula_unificada": {
            "TIPO_DOC": "source_exact",
            "N_DOC": "source_exact",
            "DV": "source_exact",
            "COD_CAR": "manual_mapping",
            "MODALIDAD": "manual_mapping",
            "JOR": "manual_mapping",
            "ANIO_ING_ACT": "source_exact",
            "SEM_ING_ACT": "source_exact",
            "ANIO_ING_ORI": "source_exact",
            "SEM_ING_ORI": "source_exact",
            "VIG": "source_exact",
            "PAIS_EST_SEC": "missing_blocker",
            "REINCORPORACION": "missing_blocker",
        },
        "carreras_ac": {c: "template_preserved" for c in CARRERAS_AC_COLUMNS},
        "matricula_ac": {c: "template_preserved" for c in MATRICULA_AC_COLUMNS},
    }

    for c in [
        "CURSO_1ER_SEM",
        "CURSO_2DO_SEM",
        "UNIDADES_CURSADAS",
        "UNIDADES_APROBADAS",
        "UNID_CURSADAS_TOTAL",
        "UNID_APROBADAS_TOTAL",
    ]:
        provenance_rules["matricula_ac"][c] = "derived_rule"

    frames = {"matricula_unificada": mu, "carreras_ac": ca, "matricula_ac": ma}
    out_rows = []
    summary: dict[str, object] = {}

    for name, df in frames.items():
        rows = []
        for col in df.columns:
            prov = provenance_rules.get(name, {}).get(col, "source_normalized")
            profile = _profile_column(df[col])
            row = {"archivo": name, "columna": col, "provenance": prov, **profile}
            if prov == "missing_blocker" and profile["null_pct"] > 0.2:
                issues.append(Issue("BLOCKER", name, f"{col} con missing_blocker masivo", int(profile["null_pct"] * len(df))))
            if col in {
                "CURSO_1ER_SEM",
                "CURSO_2DO_SEM",
                "UNIDADES_CURSADAS",
                "UNIDADES_APROBADAS",
                "UNID_CURSADAS_TOTAL",
                "UNID_APROBADAS_TOTAL",
            } and profile["default_like_pct"] > 0.98 and profile["null_pct"] > 0.95:
                issues.append(
                    Issue(
                        "WARN",
                        name,
                        f"{col} con default_like_pct alto (revisar fuente histórica)",
                        int(profile["default_like_pct"] * len(df)),
                    )
                )
            rows.append(row)

        t = pd.DataFrame(rows)
        summary[name] = {
            "avg_null_pct": round(float(t["null_pct"].mean()), 4),
            "avg_default_like_pct": round(float(t["default_like_pct"].mean()), 4),
            "provenance_pct": {k: round(float((t["provenance"] == k).mean()), 4) for k in t["provenance"].unique()},
        }
        out_rows.extend(rows)

    prov_df = pd.DataFrame(out_rows)
    prov_df.to_csv(output_dir / "reporte_procedencia.csv", index=False, encoding="utf-8")
    (output_dir / "reporte_calidad_semantica.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return summary


def ejecutar_pipeline(input_file: Path, output_dir: Path) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    issues: list[Issue] = []

    carreras_raw, mat_raw, hist_raw, equiv = cargar_fuentes(input_file)

    mat_i = preparar_matricula_intermedia(mat_raw)
    bridge, diag_amb = construir_puente_equiv(equiv)
    hist_map, review_nomap = mapear_historico_con_equiv(hist_raw, bridge)
    resumen = construir_resumen_historico(hist_map)

    carreras_ctrl = construir_carreras_control(carreras_raw)
    matac_ctrl = construir_matricula_ac_control(mat_i, resumen)
    mu_ctrl = construir_matricula_unificada_control(matac_ctrl, equiv)

    issues.extend(validar_carreras(carreras_ctrl))
    issues.extend(validar_matricula_ac(matac_ctrl, carreras_ctrl))
    issues.extend(validar_matricula_unificada(mu_ctrl))

    mu_ctrl.to_csv(output_dir / "matricula_unificada_2026_control.csv", index=False, encoding="utf-8")
    mu_ctrl.to_excel(output_dir / "matricula_unificada_2026_oficial.xlsx", index=False)
    exportar_control_y_pes(
        carreras_ctrl,
        output_dir / "carreras_avance_curricular_2025_control.csv",
        output_dir / "carreras_avance_curricular_2025_pes_ready.csv",
        issues,
        "carreras",
    )
    exportar_control_y_pes(
        matac_ctrl,
        output_dir / "matricula_avance_curricular_2025_control.csv",
        output_dir / "matricula_avance_curricular_2025_pes_ready.csv",
        issues,
        "matricula_ac",
    )

    diag_amb.to_csv(output_dir / "sies_ambiguedad_diagnostico.csv", index=False, encoding="utf-8")
    review_nomap.to_csv(output_dir / "sies_codcarr_sin_mapeo.csv", index=False, encoding="utf-8")

    calidad_semantica = generar_procedencia_y_calidad(output_dir, mu_ctrl, carreras_ctrl, matac_ctrl, issues)

    report = {
        "rows": {
            "carreras_control": len(carreras_ctrl),
            "matricula_ac_control": len(matac_ctrl),
            "matricula_unificada_control": len(mu_ctrl),
            "historico_total": len(hist_raw),
            "historico_mapeado": int(hist_map["CODIGO_UNICO"].notna().sum()),
            "historico_sin_mapeo": int(hist_map["CODIGO_UNICO"].isna().sum()),
        },
        "issues": [i.__dict__ for i in issues],
        "ambiguedad": {
            "codcarr_total": int(len(diag_amb)),
            "codcarr_ambiguos": int(diag_amb["es_ambiguo"].sum()) if not diag_amb.empty else 0,
        },
        "calidad_semantica": calidad_semantica,
        "apto_oficial": {
            "matricula_unificada": not any(
                i.severity in {"BLOCKER", "ERROR"} and i.area in {"matricula_unificada"} for i in issues
            ),
            "avance_curricular": not any(
                i.severity in {"BLOCKER", "ERROR"} and i.area in {"carreras", "matricula_ac", "carreras_ac"}
                for i in issues
            ),
        },
    }
    (output_dir / "reporte_validacion.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return report


def generar_comparacion_versiones(output_dir: Path) -> None:
    md = """# Comparación evolutiva de pipeline

| Dimensión | Versión actual previa | Idea histórica (referencia) | Decisión híbrida aplicada |
|---|---|---|---|
| Granularidad AC | colapsaba por RUT | mantener múltiples programas | deduplicación por clave de negocio, no por RUT |
| Equivalencias | `drop_duplicates(CODCARR)` | no forzar ambiguos | mapeo por `(CODCARR,JORNADA)` y fallback solo si CODCARR único |
| Validación | schema + catálogos | auditoría por bloques | capas A/B/C + issues con severidad |
| Exportación PES | quitaba primera columna por índice | contrato explícito | quitar `CODIGO_IES_NUM` por nombre |
| Defaults sensibles | rellenos por defecto | evitar inventar | `PAIS_EST_SEC` y `REINCORPORACION` quedan BLOCKER si faltan |
| Ambigüedad SIES | diagnóstico simple | revisión consolidada | diagnóstico + archivo de `CODCARR` sin mapeo |
"""
    (output_dir / "comparacion_versiones.md").write_text(md, encoding="utf-8")


def generar_diccionario_columnas(output_dir: Path) -> None:
    rows = [
        ("MODALIDAD", "respaldada por manual"),
        ("VIG", "respaldada por manual"),
        ("COD_CAR", "respaldada por código"),
        ("ANIO_ING_ACT", "respaldada por código"),
        ("SEM_ING_ACT", "respaldada por código"),
        ("ANIO_ING_ORI", "respaldada por código"),
        ("SEM_ING_ORI", "respaldada por código"),
    ]
    known = {r[0] for r in rows}
    rows.extend((c, "inferencia operativa / pendiente de validación documental") for c in MATRICULA_UNIFICADA_COLUMNS if c not in known)
    lines = ["# Diccionario de columnas y clasificación", "", "| Columna | Clasificación |", "|---|---|"]
    lines.extend(f"| {c} | {k} |" for c, k in rows)
    (output_dir / "diccionario_columnas.md").write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Pipeline híbrido SIES/PES")
    p.add_argument(
        "--input",
        default=_default_input_arg(),
        help="Ruta del Excel de entrada (por defecto detecta PROMEDIOSDEALUMNOS_7804.xlsx si existe).",
    )
    p.add_argument("--output-dir", default="resultados", help="Carpeta de salida")
    p.add_argument(
        "--proceso",
        default="avance",
        choices=["avance", "matricula", "ambos"],
        help="Qué pipeline ejecutar: avance curricular, matrícula unificada (legacy-like) o ambos",
    )
    p.add_argument(
        "--sheet",
        default=None,
        help="Nombre de hoja para el proceso matrícula (por defecto usa la primera)",
    )
    p.add_argument(
        "--catalogo-manual-tsv",
        default=None,
        help="Ruta a TSV de catálogo manual (opcional). Si no se informa, usa CATALOGO_MANUAL_TSV embebido.",
    )
    p.add_argument(
        "--puente-sies-tsv",
        default=None,
        help=(
            "DEPRECADO: override puente SIES. "
            "No se consume directamente en el pipeline; usar scripts/compile_puente_sies_compilado.py."
        ),
    )
    p.add_argument(
        "--oferta-academica-xlsx",
        default=None,
        help=(
            "Ruta al XLSX de oferta académica (opcional). "
            "Se usa para validar y ajustar MODALIDAD/JOR y limitar NIV_ACA por DURACION_ESTUDIOS."
        ),
    )
    p.add_argument(
        "--gob-nac-tsv",
        default=None,
        help="Ruta a TSV de gobernanza NAC (opcional).",
    )
    p.add_argument(
        "--gob-pais-est-sec-tsv",
        default=None,
        help="Ruta a TSV de gobernanza PAIS_EST_SEC (opcional).",
    )
    p.add_argument(
        "--gob-sede-tsv",
        default=None,
        help="Ruta a TSV de gobernanza COD_SED (opcional).",
    )
    p.add_argument(
        "--excluir-diplomados",
        choices=["true", "false"],
        default="true" if DEFAULT_EXCLUIR_DIPLOMADOS else "false",
        help="Para proceso matrícula: excluir diplomados en asignación SIES.",
    )
    p.add_argument(
        "--filtro-base-datos-sheet",
        default=None,
        help=(
            "Nombre de hoja del Excel de entrada que contiene los RUT (N_DOC) a incluir. "
            "Solo se procesarán las filas de la hoja fuente cuyo RUT aparezca en esta hoja. "
            "Ejemplo: --filtro-base-datos-sheet base_datos"
        ),
    )
    p.add_argument(
        "--usar-gobernanza-v2",
        choices=["true", "false"],
        default="false",
        help=(
            "Activa el nuevo flujo de matrícula (enriquecimiento por CODCLI con hoja DatosAlumnos). "
            "Por defecto se mantiene el flujo legacy para rollback inmediato."
        ),
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"No se encontró archivo de entrada: {input_path}. Usa --input para indicar uno válido.")

    out = Path(args.output_dir).expanduser().resolve()
    reports: dict[str, object] = {}

    if args.proceso in {"avance", "ambos"}:
        report_avance = ejecutar_pipeline(input_path, out)
        generar_comparacion_versiones(out)
        generar_diccionario_columnas(out)
        reports["avance"] = report_avance

    if args.proceso in {"matricula", "ambos"}:
        catalogo_manual_tsv_path = _resolve_optional_path(args.catalogo_manual_tsv, DEFAULT_CATALOGO_MANUAL_CANDIDATES)
        puente_sies_tsv_path = args.puente_sies_tsv
        gob_nac_tsv_path = _resolve_optional_path(args.gob_nac_tsv, DEFAULT_GOB_NAC_CANDIDATES)
        gob_pais_est_sec_tsv_path = _resolve_optional_path(args.gob_pais_est_sec_tsv, DEFAULT_GOB_PAIS_EST_SEC_CANDIDATES)
        gob_sede_tsv_path = _resolve_optional_path(args.gob_sede_tsv, DEFAULT_GOB_SEDE_CANDIDATES)
        report_mu = ejecutar_pipeline_matricula_unificada_legacy_like(
            input_path,
            out,
            sheet_name=args.sheet,
            catalogo_manual_tsv_path=catalogo_manual_tsv_path,
            puente_sies_tsv_path=puente_sies_tsv_path,
            oferta_academica_xlsx_path=args.oferta_academica_xlsx,
            gob_nac_tsv_path=gob_nac_tsv_path,
            gob_pais_est_sec_tsv_path=gob_pais_est_sec_tsv_path,
            gob_sede_tsv_path=gob_sede_tsv_path,
            excluir_diplomados=(args.excluir_diplomados == "true"),
            usar_gobernanza_v2=(args.usar_gobernanza_v2 == "true"),
            filtro_base_datos_sheet=args.filtro_base_datos_sheet,
        )
        reports["matricula"] = report_mu

    if args.proceso == "avance":
        print(json.dumps(reports["avance"], indent=2, ensure_ascii=False))
    elif args.proceso == "matricula":
        print(json.dumps(reports["matricula"], indent=2, ensure_ascii=False))
    else:
        print(json.dumps(reports, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
