from __future__ import annotations

<<<<<<< ours
import argparse
import json
from dataclasses import dataclass
from datetime import date
=======
# ──────────────────────────────────────────────────────────────
# ❶ · Librerías y utilidades generales
# ──────────────────────────────────────────────────────────────
import re, warnings
>>>>>>> theirs
from pathlib import Path
<<<<<<< ours
import pandas as pd, numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

warnings.filterwarnings("ignore", category=pd.errors.DtypeWarning)

# Columnas requeridas por el instructivo SIES 2025
DIST_UNIDADES = [
    "UNIDADES_1ER_ANIO", "UNIDADES_2DO_ANIO", "UNIDADES_3ER_ANIO",
    "UNIDADES_4TO_ANIO", "UNIDADES_5TO_ANIO", "UNIDADES_6TO_ANIO",
    "UNIDADES_7MO_ANIO",
]

REQ_CARRERAS = {
    "CODIGO_UNICO", "PLAN_ESTUDIOS", "TOTAL_UNIDADES_MEDIDA", "TIPO_UNIDAD_MEDIDA", "VIGENCIA",
    *DIST_UNIDADES,
}
REQ_MATRICULA = {
    "NUM_DOCUMENTO", "DV", "CODIGO_UNICO", "PLAN_ESTUDIOS",
    "CURSO_1ER_SEM", "CURSO_2DO_SEM", "UNIDADES_CURSADAS", "UNIDADES_APROBADAS",
    "UNID_CURSADAS_TOTAL", "UNID_APROBADAS_TOTAL", "VIGENCIA",
}
REQ_HIST = {"RUT", "DIG", "ANO", "PERIODO", "CODRAMO", "DESCRIPCION_ESTADO"}


def _validate_columns(df: pd.DataFrame, required: set[str], nombre: str):
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(f"La hoja {nombre} no cumple el instructivo SIES: faltan columnas {sorted(missing)}")


# ──────────────────────────────────────────────────────────────
# ❷ · Carga / descarga de archivos en Google Colab
# ──────────────────────────────────────────────────────────────
def _in_colab() -> bool:
    """Detecta si el entorno de ejecución es Google Colab."""
    try:
        import google.colab  # type: ignore
        return True
    except ModuleNotFoundError:
        return False

def _ruta_content(fname: str | Path) -> Path:
    """
    🔧 Corrección aplicada:
    Convierte cualquier nombre de archivo a una ruta absoluta dentro de /content
    cuando el código se ejecuta en Colab, previniendo errores de I/O.
    """
    p = Path(fname)
    if _in_colab():
        return Path("/content") / p.name
    return p.resolve()

def cargar_excel() -> Path:
    """
    • En Colab muestra el widget `files.upload()`  
    • En local (Jupyter/Lab) solicita la ruta por input().
    Devuelve siempre la ruta absoluta dentro del directorio de trabajo.
    """
    if _in_colab():
        from google.colab import files  # type: ignore
        uploads = files.upload()
        while not uploads:
            print("⚠ No file selected. Please upload an Excel file.")
            uploads = files.upload()
        ruta_subida = Path(next(iter(uploads)))
        destino = _ruta_content(ruta_subida)
        ruta_subida.replace(destino)
        print(f"✔ Archivo guardado en: {destino}")
        return destino
    else:
        ruta_local = Path(input("Ruta de archivo .xlsx: ").strip())
        if not ruta_local.exists():
            raise FileNotFoundError(ruta_local)
        return ruta_local.resolve()

def descargar_excel(path: str | Path) -> None:
    """
    Descarga el archivo al equipo en Colab; en local no hace nada.
    """
    if _in_colab():
        from google.colab import files  # type: ignore
        files.download(str(path))

# ──────────────────────────────────────────────────────────────
# ❸ · Normalización de RUT + formateo de Excel
# ──────────────────────────────────────────────────────────────
def normalizar_rut(num, dv) -> str:
    return re.sub(r"[^0-9]", "", str(num)) + str(dv).upper().strip()

def formatear_excel(path: str | Path) -> None:
    """
    • Activa filtro en fila 1  
    • Centra texto horizontal/vertical  
    • Ajusta ancho de columnas (12 – 45 px)
    """
    path = Path(path)
    wb = load_workbook(path)
    alin = Alignment(horizontal="center", vertical="center")
    for ws in wb.worksheets:
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            w = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(45, w + 2))
        for row in ws.iter_rows():
            for c in row:
                c.alignment = alin
    wb.save(path)

# ╔═════════════════════════════════════════════════════════════╗
#   BLOQUE 2 · FONDO — lógica Avance Curricular SIES 2025       #
# ╚═════════════════════════════════════════════════════════════╝

# ──────────────────────────────────────────────────────────────
# ❹ · Carga y validación de hojas obligatorias + Equivalencia
# ──────────────────────────────────────────────────────────────
def _find_sheet(book: dict, required: set[str]):
    return next((df for df in book.values() if required.issubset(df.columns)), None)


def cargar_datos(ruta: Path):
    book = pd.read_excel(ruta, sheet_name=None)

    df_carreras  = _find_sheet(book, REQ_CARRERAS)
    df_matricula = _find_sheet(book, REQ_MATRICULA)
    hojas_hist   = [df for df in book.values() if REQ_HIST.intersection(df.columns) == REQ_HIST]
    df_equiv     = book.get("Equivalencia")

    if df_carreras is None or df_matricula is None or not hojas_hist or df_equiv is None:
        raise ValueError("Faltan hojas Carreras, Matrícula, Históricos o Equivalencia.")

    _validate_columns(df_carreras, REQ_CARRERAS, "Carreras")
    _validate_columns(df_matricula, REQ_MATRICULA, "Matrícula")
    for i, hoja in enumerate(hojas_hist, start=1):
        _validate_columns(hoja, REQ_HIST, f"Histórico #{i}")

    # Validación de coherencia de unidades por año
    dist_sum = df_carreras[DIST_UNIDADES].fillna(0).sum(axis=1)
    incoherentes = df_carreras.loc[dist_sum != pd.to_numeric(df_carreras["TOTAL_UNIDADES_MEDIDA"], errors="coerce")]
    if not incoherentes.empty:
        raise ValueError("TOTAL_UNIDADES_MEDIDA no coincide con la suma por año en algunas filas de Carreras.")

    # Validación OTRA_UNIDAD_MEDIDA cuando TIPO_UNIDAD_MEDIDA == 3
    mask_tipo3 = df_carreras["TIPO_UNIDAD_MEDIDA"].eq(3)
    if mask_tipo3.any() and df_carreras.loc[mask_tipo3, "OTRA_UNIDAD_MEDIDA"].isna().any():
        raise ValueError("Para TIPO_UNIDAD_MEDIDA=3 se debe informar OTRA_UNIDAD_MEDIDA.")

    df_hist = pd.concat(hojas_hist, ignore_index=True)
    return df_carreras, df_matricula, df_hist, df_equiv

# ──────────────────────────────────────────────────────────────
# ❺ · Limpieza, normalización y selección de carrera vigente
# ──────────────────────────────────────────────────────────────
def preparar_datos(df_carreras, df_matricula, df_hist, df_equiv):
    # A · Normalizar RUT
    df_hist['RUT_NORM']      = [normalizar_rut(r, d) for r, d in zip(df_hist['RUT'], df_hist['DIG'])]
    df_matricula['RUT_NORM'] = [normalizar_rut(n, d) for n, d in zip(df_matricula['NUM_DOCUMENTO'], df_matricula['DV'])]

    # B · Deduplicación
    df_carreras  = df_carreras.drop_duplicates(['CODIGO_UNICO', 'PLAN_ESTUDIOS'])
    df_matricula = df_matricula.drop_duplicates(['RUT_NORM', 'CODIGO_UNICO', 'PLAN_ESTUDIOS'])

    # C · Columnas opcionales con valores por defecto
    if "ESTADO_ACADEMICO" not in df_matricula.columns:
        df_matricula["ESTADO_ACADEMICO"] = "VIGENTE"
    if "VIGENCIA" not in df_matricula.columns:
        df_matricula["VIGENCIA"] = 1
    col_plan = "PLAN_DE_ESTUDIO" if "PLAN_DE_ESTUDIO" in df_matricula.columns else None

    # D · Cálculo de PESO_CARRERA
    prioridad = {"VIGENTE": 3, "SUSPENDIDO": 2, "ELIMINADO": 1}
    df_matricula["ANIO_PLAN"] = (
        df_matricula[col_plan].astype(str).str.extract(r"(\d{4})")[0].astype(float)
        if col_plan else 0
    )
    df_matricula["PESO_CARRERA"] = (
        df_matricula["ESTADO_ACADEMICO"].map(prioridad).fillna(0) * 1e6 +
        df_matricula["VIGENCIA"] * 1e3 +
        df_matricula["ANIO_PLAN"].fillna(0)
    )

    # E · Seleccionar carrera vigente por RUT
    idx_vig = (df_matricula.sort_values("PESO_CARRERA", ascending=False)
               .groupby("RUT_NORM").head(1).index)
    df_mat_ok = df_matricula.loc[idx_vig].copy()

    # F · Mapear CODCARR → CODIGO_UNICO usando Equivalencia
    if "CODCARR" in df_hist.columns:
        mapa_eq = (df_equiv[["CODCARR", "CODIGO_UNICO"]]
                   .dropna(subset=["CODCARR", "CODIGO_UNICO"])
                   .drop_duplicates("CODCARR")
                   .set_index("CODCARR")["CODIGO_UNICO"])
        df_hist["CODIGO_UNICO_EQ"] = df_hist["CODCARR"].map(mapa_eq)
        df_hist = df_hist.merge(
            df_mat_ok[["RUT_NORM", "CODIGO_UNICO"]],
            left_on=["RUT_NORM", "CODIGO_UNICO_EQ"],
            right_on=["RUT_NORM", "CODIGO_UNICO"],
            how="inner"
        ).drop(columns=["CODIGO_UNICO_EQ"])

    # G · Asegurar CODIGO_UNICO y PLAN_ESTUDIOS en histórico
    df_hist = df_hist.merge(
        df_mat_ok[["RUT_NORM", "CODIGO_UNICO", "PLAN_ESTUDIOS"]],
        on="RUT_NORM", how="left", suffixes=("", "_MAT")
    )
    for col in ["CODIGO_UNICO", "PLAN_ESTUDIOS"]:
        alt = f"{col}_MAT"
        if alt in df_hist.columns:
            df_hist[col] = df_hist[col].combine_first(df_hist[alt])
            df_hist.drop(columns=[alt], inplace=True)

    # H · Filtrar histórico por la carrera vigente seleccionada
    df_hist_ok = df_hist.merge(
        df_mat_ok[["RUT_NORM", "CODIGO_UNICO", "PLAN_ESTUDIOS"]],
        on=["RUT_NORM", "CODIGO_UNICO", "PLAN_ESTUDIOS"],
        how="inner"
    )
    return df_carreras, df_mat_ok, df_hist_ok

# ──────────────────────────────────────────────────────────────
# ❻ · Hoja Intentos — histórico alineado a la carrera vigente
# ──────────────────────────────────────────────────────────────
def construir_hoja_intentos(df_hist, df_matricula):
    return df_hist.copy()

# ──────────────────────────────────────────────────────────────
# ❼ · Construcción Resumen SIES 2024 por estudiante
# ──────────────────────────────────────────────────────────────
def construir_resumen_sies(df_intentos, df_carreras, log_errores=None, *, ano_focal=2024, periodo_focal=None):
    """
    Calcula el resumen SIES para un año focal dado. Si se especifica
    ``periodo_focal`` (1–3), el desglose del año focal se limita a ese
    período; en caso contrario se usa el año completo.
    """

    res, pat24, pat_hist = [], r"APROB", r"(?:APROB|CONVALID|RECONOC|EQUIV|HOMOLOG)"
    tot_plan = (df_carreras[["CODIGO_UNICO", "PLAN_ESTUDIOS", "TOTAL_UNIDADES_MEDIDA"]]
                .assign(TOTAL_UNIDADES_MEDIDA=lambda d: pd.to_numeric(d["TOTAL_UNIDADES_MEDIDA"], errors="coerce"))
                .set_index(["CODIGO_UNICO", "PLAN_ESTUDIOS"])["TOTAL_UNIDADES_MEDIDA"]
                .to_dict())

    for (rut, cu, plan), sub in df_intentos.groupby(["RUT_NORM", "CODIGO_UNICO", "PLAN_ESTUDIOS"]):
        total = tot_plan.get((cu, plan), np.nan)
        if pd.isna(total) or total <= 0:
            if log_errores is not None:
                log_errores.append({"tipo": "TOTAL_UNIDADES_MEDIDA vacío", "CODIGO_UNICO": cu, "PLAN_ESTUDIOS": plan})
            total = max(sub["CODRAMO"].nunique(), 1)

        sub24 = sub[sub["ANO"] == ano_focal]
        if periodo_focal:
            sub24 = sub24[sub24["PERIODO"] == periodo_focal]

        e24   = sub24["DESCRIPCION_ESTADO"].astype(str).str.upper()
        eh    = sub["DESCRIPCION_ESTADO"].astype(str).str.upper()

        cur24 = sub24["CODRAMO"].drop_duplicates().size
        apr24 = sub24[e24.str.contains(pat24, regex=True, na=False)]["CODRAMO"].drop_duplicates().size
        curt  = sub["CODRAMO"].drop_duplicates().size
        aprt  = min(
            (sub[eh.str.contains(pat_hist, regex=True, na=False)]
             .sort_values(["CODRAMO", "ANO", "PERIODO"])
             .drop_duplicates("CODRAMO").shape[0]),
            int(round(1.25 * total))
        )

        desglose = {
            f"UNIDADES_{n}_ANIO": sub24[sub24["PERIODO"] == n]["CODRAMO"].drop_duplicates().size
            for n in range(1, 8)
        }
        rec = sub.sort_values(["ANO", "PERIODO"]).iloc[-1]
        res.append({
            "RUT": rut, "CODIGO_UNICO": cu, "PLAN_ESTUDIOS": plan,
            "CURSO_1ER_SEM": "SI" if (sub24["PERIODO"] == 1).any() else "NO",
            "CURSO_2DO_SEM": "SI" if (sub24["PERIODO"] == 2).any() else "NO",
            "UNIDADES_CURSADAS": cur24, "UNIDADES_APROBADAS": apr24,
            "UNID_CURSADAS_TOTAL": curt, "UNID_APROBADAS_TOTAL": aprt,
            "ESTADO_ACADEMICO": rec.get("ESTADO_ACADEMICO", "SIN_INFO"),
            **desglose
=======
from typing import Iterable

import pandas as pd

# ==============================
# Contratos oficiales (Capa C)
# ==============================
MATRICULA_UNIFICADA_COLUMNS = [
    "TIPO_DOC", "N_DOC", "DV", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "NOMBRE", "SEXO", "FECH_NAC",
    "NAC", "PAIS_EST_SEC", "COD_SED", "COD_CAR", "MODALIDAD", "JOR", "VERSION", "FOR_ING_ACT",
    "ANIO_ING_ACT", "SEM_ING_ACT", "ANIO_ING_ORI", "SEM_ING_ORI", "ASI_INS_ANT", "ASI_APR_ANT",
    "PROM_PRI_SEM", "PROM_SEG_SEM", "ASI_INS_HIS", "ASI_APR_HIS", "NIV_ACA", "SIT_FON_SOL", "SUS_PRE",
    "FECHA_MATRICULA", "REINCORPORACION", "VIG",
]

CARRERAS_AC_COLUMNS = [
    "CODIGO_IES_NUM", "CODIGO_UNICO", "PLAN_ESTUDIOS", "NOMBRE_SEDE", "NOMBRE_CARRERA", "JORNADA", "VERSION",
    "DURACION_ESTUDIOS", "DURACION_TITULACION", "DURACION_TOTAL", "NIVEL_CARRERA", "TIPO_UNIDAD_MEDIDA",
    "OTRA_UNIDAD_MEDIDA", "TOTAL_UNIDADES_MEDIDA", "UNIDADES_1ER_ANIO", "UNIDADES_2DO_ANIO", "UNIDADES_3ER_ANIO",
    "UNIDADES_4TO_ANIO", "UNIDADES_5TO_ANIO", "UNIDADES_6TO_ANIO", "UNIDADES_7MO_ANIO", "VIGENCIA",
]

MATRICULA_AC_COLUMNS = [
    "CODIGO_IES_NUM", "TIPO_DOCUMENTO", "NUM_DOCUMENTO", "DV", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "NOMBRES",
    "SEXO", "FECHA_NACIMIENTO", "CODIGO_UNICO", "PLAN_ESTUDIOS", "ANIO_INGRESO_CARRERA_ACTUAL",
    "SEM_INGRESO_CARRERA_ACTUAL", "ANIO_INGRESO_CARRERA_ORIGEN", "SEM_INGRESO_CARRERA_ORIGEN",
    "CURSO_1ER_SEM", "CURSO_2DO_SEM", "UNIDADES_CURSADAS", "UNIDADES_APROBADAS", "UNID_CURSADAS_TOTAL",
    "UNID_APROBADAS_TOTAL", "VIGENCIA",
]

ANUAL_COLS = [
    "UNIDADES_1ER_ANIO", "UNIDADES_2DO_ANIO", "UNIDADES_3ER_ANIO", "UNIDADES_4TO_ANIO",
    "UNIDADES_5TO_ANIO", "UNIDADES_6TO_ANIO", "UNIDADES_7MO_ANIO",
]


@dataclass
class Issue:
    severity: str
    area: str
    message: str
    count: int = 0


def _normalize_doc(num: object, dv: object) -> str:
    return "".join(ch for ch in str(num) if ch.isdigit()) + str(dv).strip().upper()


def _pick_sheet(book: dict[str, pd.DataFrame], required: Iterable[str]) -> pd.DataFrame:
    req = set(required)
    for df in book.values():
        if req.issubset(df.columns):
            return df.copy()
    raise ValueError(f"No se encontró hoja con columnas: {sorted(req)}")


# ==============================
# CAPA A: Ingesta flexible
# ==============================
def cargar_fuentes(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    book = pd.read_excel(path, sheet_name=None)
    carreras = _pick_sheet(book, {"CODIGO_UNICO", "PLAN_ESTUDIOS"})
    matricula = _pick_sheet(book, {"NUM_DOCUMENTO", "CODIGO_UNICO", "PLAN_ESTUDIOS", "DV"})
    hist = pd.concat([df for df in book.values() if {"ANO", "PERIODO", "RUT", "DIG", "CODCARR"}.issubset(df.columns)], ignore_index=True)
    if hist.empty:
        raise ValueError("No se encontró histórico con ANO/PERIODO/RUT/DIG/CODCARR")
    equiv = book.get("Equivalencia")
    if equiv is None:
        raise ValueError("Falta hoja Equivalencia")
    return carreras, matricula, hist.copy(), equiv.copy()


def preparar_matricula_intermedia(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(out["NUM_DOCUMENTO"], out["DV"])]
    # Regla: no colapsar por RUT, solo deduplicación exacta de clave de negocio
    key = ["NUM_DOCUMENTO", "DV", "CODIGO_UNICO", "PLAN_ESTUDIOS"]
    out = out.drop_duplicates(key)
    return out


def construir_puente_equiv(df_equiv: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    cols = [c for c in ["CODCARR", "CODIGO_UNICO", "JORNADA", "VERSION"] if c in df_equiv.columns]
    bridge = df_equiv[cols].dropna(subset=["CODCARR", "CODIGO_UNICO"]).copy()
    bridge["CODCARR"] = bridge["CODCARR"].astype(str)

    g = bridge.groupby("CODCARR").agg(codigos_unicos=("CODIGO_UNICO", "nunique"), jornadas=("JORNADA", "nunique") if "JORNADA" in bridge.columns else ("CODIGO_UNICO", "size"), versiones=("VERSION", "nunique") if "VERSION" in bridge.columns else ("CODIGO_UNICO", "size")).reset_index()
    g["es_ambiguo"] = g["codigos_unicos"] > 1
    return bridge, g


def mapear_historico_con_equiv(df_hist: pd.DataFrame, bridge: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    hist = df_hist.copy()
    hist["RUT_NORM"] = [_normalize_doc(n, d) for n, d in zip(hist["RUT"], hist["DIG"])]
    hist["CODCARR"] = hist["CODCARR"].astype(str)

    base_cols = ["CODCARR", "CODIGO_UNICO"]
    key_jornada = {"CODCARR", "JORNADA"}.issubset(hist.columns) and {"CODCARR", "JORNADA"}.issubset(bridge.columns)

    if key_jornada:
        hist["JORNADA"] = hist["JORNADA"].astype(str)
        bridge = bridge.copy()
        bridge["JORNADA"] = bridge["JORNADA"].astype(str)
        pair = bridge.drop_duplicates(["CODCARR", "JORNADA", "CODIGO_UNICO"])
        counts = pair.groupby(["CODCARR", "JORNADA"])["CODIGO_UNICO"].nunique().reset_index(name="n")
        unique_pair = pair.merge(counts[counts["n"] == 1], on=["CODCARR", "JORNADA"], how="inner")[["CODCARR", "JORNADA", "CODIGO_UNICO"]]
        hist = hist.merge(unique_pair, on=["CODCARR", "JORNADA"], how="left")

    unresolved = hist[hist["CODIGO_UNICO"].isna()].copy()
    if not unresolved.empty:
        single = bridge.groupby("CODCARR")["CODIGO_UNICO"].nunique().reset_index(name="n")
        single = single[single["n"] == 1].merge(bridge[["CODCARR", "CODIGO_UNICO"]].drop_duplicates("CODCARR"), on="CODCARR", how="left")
        unresolved = unresolved.drop(columns=[c for c in ["CODIGO_UNICO"] if c in unresolved.columns]).merge(single[["CODCARR", "CODIGO_UNICO"]], on="CODCARR", how="left")
        resolved = hist[hist["CODIGO_UNICO"].notna()].copy()
        hist = pd.concat([resolved, unresolved], ignore_index=True)

    review = hist[hist["CODIGO_UNICO"].isna()][["CODCARR"]].value_counts().reset_index(name="filas_sin_map")
    return hist, review


# ==============================
# CAPA B: Modelo intermedio
# ==============================
def construir_resumen_historico(hist_mapeado: pd.DataFrame) -> pd.DataFrame:
    valid = hist_mapeado[hist_mapeado["CODIGO_UNICO"].notna()].copy()
    if valid.empty:
        return pd.DataFrame(columns=["RUT_NORM", "CODIGO_UNICO", "CURSO_1ER_SEM", "CURSO_2DO_SEM", "UNIDADES_CURSADAS", "UNIDADES_APROBADAS", "UNID_CURSADAS_TOTAL", "UNID_APROBADAS_TOTAL"])

    rows = []
    for (rut, cod), sub in valid.groupby(["RUT_NORM", "CODIGO_UNICO"]):
        s24 = sub[sub["ANO"] == 2024]
        e24 = s24.get("DESCRIPCION_ESTADO", "").astype(str).str.upper()
        eh = sub.get("DESCRIPCION_ESTADO", "").astype(str).str.upper()
        rows.append({
            "RUT_NORM": rut,
            "CODIGO_UNICO": cod,
            "CURSO_1ER_SEM": "SI" if (s24["PERIODO"] == 1).any() else "NO",
            "CURSO_2DO_SEM": "SI" if (s24["PERIODO"] == 2).any() else "NO",
            "UNIDADES_CURSADAS": s24.get("CODRAMO", pd.Series(dtype=object)).nunique(),
            "UNIDADES_APROBADAS": s24.loc[e24.str.contains("APROB", na=False), "CODRAMO"].nunique(),
            "UNID_CURSADAS_TOTAL": sub.get("CODRAMO", pd.Series(dtype=object)).nunique(),
            "UNID_APROBADAS_TOTAL": sub.loc[eh.str.contains(r"APROB|CONVALID|RECONOC|EQUIV|HOMOLOG", regex=True, na=False), "CODRAMO"].nunique(),
>>>>>>> theirs
        })
    return pd.DataFrame(rows)


def construir_carreras_control(df_carreras: pd.DataFrame) -> pd.DataFrame:
    out = df_carreras.copy().drop_duplicates(["CODIGO_UNICO", "PLAN_ESTUDIOS"])
    for col in CARRERAS_AC_COLUMNS:
        if col not in out.columns:
            out[col] = pd.NA
    return out[CARRERAS_AC_COLUMNS].copy()


def construir_matricula_ac_control(df_matricula: pd.DataFrame, resumen: pd.DataFrame) -> pd.DataFrame:
    out = df_matricula.copy()
    if not resumen.empty:
        out = out.merge(resumen, on=["RUT_NORM", "CODIGO_UNICO"], how="left", suffixes=("", "_CALC"))
    for col in ["CURSO_1ER_SEM", "CURSO_2DO_SEM", "UNIDADES_CURSADAS", "UNIDADES_APROBADAS", "UNID_CURSADAS_TOTAL", "UNID_APROBADAS_TOTAL"]:
        if col not in out.columns and f"{col}_CALC" in out.columns:
            out[col] = out[f"{col}_CALC"]
    for col in MATRICULA_AC_COLUMNS:
        if col not in out.columns:
            out[col] = pd.NA
    return out[MATRICULA_AC_COLUMNS].copy()


def construir_matricula_unificada_control(mat_ac: pd.DataFrame, df_equiv: pd.DataFrame) -> pd.DataFrame:
    jmap = {}
    if {"CODIGO_UNICO", "JORNADA"}.issubset(df_equiv.columns):
        for _, r in df_equiv[["CODIGO_UNICO", "JORNADA"]].dropna().drop_duplicates().iterrows():
            j = str(r["JORNADA"]).strip().lower()
            if "diurn" in j:
                jmap[r["CODIGO_UNICO"]] = ("PRESENCIAL", "1")
            elif "vespert" in j:
                jmap[r["CODIGO_UNICO"]] = ("PRESENCIAL", "2")
            elif "semi" in j:
                jmap[r["CODIGO_UNICO"]] = ("SEMIPRESENCIAL", "3")
            elif "dist" in j:
                jmap[r["CODIGO_UNICO"]] = ("DISTANCIA", "4")
            else:
                jmap[r["CODIGO_UNICO"]] = (pd.NA, pd.NA)

    out = pd.DataFrame(index=mat_ac.index)
    out["TIPO_DOC"] = mat_ac.get("TIPO_DOCUMENTO")
    out["N_DOC"] = mat_ac.get("NUM_DOCUMENTO")
    out["DV"] = mat_ac.get("DV")
    out["PRIMER_APELLIDO"] = mat_ac.get("PRIMER_APELLIDO")
    out["SEGUNDO_APELLIDO"] = mat_ac.get("SEGUNDO_APELLIDO")
    out["NOMBRE"] = mat_ac.get("NOMBRES")
    out["SEXO"] = mat_ac.get("SEXO")
    out["FECH_NAC"] = mat_ac.get("FECHA_NACIMIENTO")
    out["NAC"] = pd.NA
    out["PAIS_EST_SEC"] = pd.NA
    out["COD_SED"] = pd.NA
    out["COD_CAR"] = mat_ac.get("CODIGO_UNICO")
    modality = mat_ac.get("CODIGO_UNICO").map(lambda c: jmap.get(c, (pd.NA, pd.NA)))
    out[["MODALIDAD", "JOR"]] = pd.DataFrame(modality.tolist(), index=out.index)
    out["VERSION"] = pd.NA
    out["FOR_ING_ACT"] = pd.NA
    out["ANIO_ING_ACT"] = mat_ac.get("ANIO_INGRESO_CARRERA_ACTUAL")
    out["SEM_ING_ACT"] = mat_ac.get("SEM_INGRESO_CARRERA_ACTUAL")
    out["ANIO_ING_ORI"] = mat_ac.get("ANIO_INGRESO_CARRERA_ORIGEN")
    out["SEM_ING_ORI"] = mat_ac.get("SEM_INGRESO_CARRERA_ORIGEN")
    out["ASI_INS_ANT"] = pd.NA
    out["ASI_APR_ANT"] = pd.NA
    out["PROM_PRI_SEM"] = pd.NA
    out["PROM_SEG_SEM"] = pd.NA
    out["ASI_INS_HIS"] = pd.NA
    out["ASI_APR_HIS"] = pd.NA
    out["NIV_ACA"] = pd.NA
    out["SIT_FON_SOL"] = pd.NA
    out["SUS_PRE"] = pd.NA
    out["FECHA_MATRICULA"] = pd.NA
    out["REINCORPORACION"] = pd.NA
    out["VIG"] = mat_ac.get("VIGENCIA")
    return out[MATRICULA_UNIFICADA_COLUMNS].copy()


# ==============================
# CAPA C: Validación + Export
# ==============================
def _check_schema(df: pd.DataFrame, expected: list[str], area: str, issues: list[Issue]) -> None:
    if df.columns.tolist() != expected:
        issues.append(Issue("ERROR", area, "Schema exacto incumplido"))


def validar_carreras(carr: pd.DataFrame) -> list[Issue]:
    issues: list[Issue] = []
    _check_schema(carr, CARRERAS_AC_COLUMNS, "carreras", issues)

    if carr["PLAN_ESTUDIOS"].isna().any():
        issues.append(Issue("ERROR", "carreras", "PLAN_ESTUDIOS vacío", int(carr["PLAN_ESTUDIOS"].isna().sum())))
    bad_tum = ~carr["TIPO_UNIDAD_MEDIDA"].astype(str).isin(["1", "2", "3"]) & carr["TIPO_UNIDAD_MEDIDA"].notna()
    if bad_tum.any():
        issues.append(Issue("ERROR", "carreras", "TIPO_UNIDAD_MEDIDA fuera de catálogo", int(bad_tum.sum())))
    tum3 = carr["TIPO_UNIDAD_MEDIDA"].astype(str) == "3"
    if tum3.any() and carr.loc[tum3, "OTRA_UNIDAD_MEDIDA"].isna().any():
        issues.append(Issue("ERROR", "carreras", "OTRA_UNIDAD_MEDIDA faltante cuando TIPO_UNIDAD_MEDIDA=3", int(carr.loc[tum3, "OTRA_UNIDAD_MEDIDA"].isna().sum())))

    annual = carr[ANUAL_COLS].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    total = pd.to_numeric(carr["TOTAL_UNIDADES_MEDIDA"], errors="coerce")
    mismatch = total.notna() & (annual != total)
    if mismatch.any():
        issues.append(Issue("WARN", "carreras", "Suma de unidades anuales distinta al total", int(mismatch.sum())))

    vig_bad = ~carr["VIGENCIA"].astype(str).isin(["0", "1"]) & carr["VIGENCIA"].notna()
    if vig_bad.any():
        issues.append(Issue("ERROR", "carreras", "VIGENCIA fuera de catálogo 0/1", int(vig_bad.sum())))
    return issues


def validar_matricula_ac(mat: pd.DataFrame, carr: pd.DataFrame) -> list[Issue]:
    issues: list[Issue] = []
    _check_schema(mat, MATRICULA_AC_COLUMNS, "matricula_ac", issues)

    for c in ["TIPO_DOCUMENTO", "NUM_DOCUMENTO", "DV", "CODIGO_UNICO", "PLAN_ESTUDIOS"]:
        if mat[c].isna().any():
            issues.append(Issue("ERROR", "matricula_ac", f"{c} vacío", int(mat[c].isna().sum())))

    bad_td = ~mat["TIPO_DOCUMENTO"].astype(str).isin(["R", "P"]) & mat["TIPO_DOCUMENTO"].notna()
    if bad_td.any():
        issues.append(Issue("ERROR", "matricula_ac", "TIPO_DOCUMENTO fuera de catálogo R/P", int(bad_td.sum())))
    bad_sex = ~mat["SEXO"].astype(str).isin(["M", "H", "X"]) & mat["SEXO"].notna()
    if bad_sex.any():
        issues.append(Issue("WARN", "matricula_ac", "SEXO fuera de catálogo M/H/X", int(bad_sex.sum())))

    bad_curso = ~mat["CURSO_1ER_SEM"].astype(str).isin(["SI", "NO"]) & mat["CURSO_1ER_SEM"].notna()
    bad_curso2 = ~mat["CURSO_2DO_SEM"].astype(str).isin(["SI", "NO"]) & mat["CURSO_2DO_SEM"].notna()
    if bad_curso.any() or bad_curso2.any():
        issues.append(Issue("ERROR", "matricula_ac", "CURSO_1ER_SEM/CURSO_2DO_SEM fuera de catálogo SI/NO", int(bad_curso.sum() + bad_curso2.sum())))

    sem1 = pd.to_numeric(mat["SEM_INGRESO_CARRERA_ACTUAL"], errors="coerce")
    sem2 = pd.to_numeric(mat["SEM_INGRESO_CARRERA_ORIGEN"], errors="coerce")
    bad_sem = (~sem1.isin([1, 2]) & sem1.notna()) | (~sem2.isin([1, 2]) & sem2.notna())
    if bad_sem.any():
        issues.append(Issue("WARN", "matricula_ac", "Semestres de ingreso fuera de 1/2", int(bad_sem.sum())))

    apr_gt_cur = pd.to_numeric(mat["UNIDADES_APROBADAS"], errors="coerce") > pd.to_numeric(mat["UNIDADES_CURSADAS"], errors="coerce")
    if apr_gt_cur.any():
        issues.append(Issue("WARN", "matricula_ac", "UNIDADES_APROBADAS > UNIDADES_CURSADAS", int(apr_gt_cur.sum())))

    aprt_gt_curt = pd.to_numeric(mat["UNID_APROBADAS_TOTAL"], errors="coerce") > pd.to_numeric(mat["UNID_CURSADAS_TOTAL"], errors="coerce")
    if aprt_gt_curt.any():
        issues.append(Issue("WARN", "matricula_ac", "UNID_APROBADAS_TOTAL > UNID_CURSADAS_TOTAL", int(aprt_gt_curt.sum())))

    keys_c = set(map(tuple, carr[["CODIGO_UNICO", "PLAN_ESTUDIOS"]].dropna().drop_duplicates().to_records(index=False)))
    keys_m = set(map(tuple, mat[["CODIGO_UNICO", "PLAN_ESTUDIOS"]].dropna().drop_duplicates().to_records(index=False)))
    if keys_m - keys_c:
        issues.append(Issue("ERROR", "matricula_ac", "PLAN_ESTUDIOS no referenciado en carreras", len(keys_m - keys_c)))
    return issues


def validar_matricula_unificada(mu: pd.DataFrame) -> list[Issue]:
    issues: list[Issue] = []
    _check_schema(mu, MATRICULA_UNIFICADA_COLUMNS, "matricula_unificada", issues)

    vig_bad = ~mu["VIG"].astype(str).isin(["0", "1"]) & mu["VIG"].notna()
    if vig_bad.any():
        issues.append(Issue("ERROR", "matricula_unificada", "VIG fuera de catálogo 0/1", int(vig_bad.sum())))

    reinc_bad = ~mu["REINCORPORACION"].astype(str).isin(["0", "1"]) & mu["REINCORPORACION"].notna()
    if reinc_bad.any():
        issues.append(Issue("ERROR", "matricula_unificada", "REINCORPORACION fuera de 0/1", int(reinc_bad.sum())))
    if mu["REINCORPORACION"].isna().any():
        issues.append(Issue("BLOCKER", "matricula_unificada", "REINCORPORACION obligatorio vacío", int(mu["REINCORPORACION"].isna().sum())))

    if mu["PAIS_EST_SEC"].isna().any():
        issues.append(Issue("BLOCKER", "matricula_unificada", "PAIS_EST_SEC obligatorio vacío", int(mu["PAIS_EST_SEC"].isna().sum())))

    fechas = pd.to_datetime(mu["FECHA_MATRICULA"], errors="coerce")
    futuras = fechas.notna() & (fechas > pd.Timestamp.today().normalize())
    if futuras.any():
        issues.append(Issue("ERROR", "matricula_unificada", "FECHA_MATRICULA posterior a fecha de carga", int(futuras.sum())))

    mod_missing = mu["COD_CAR"].notna() & mu["MODALIDAD"].isna()
    if mod_missing.any():
        issues.append(Issue("WARN", "matricula_unificada", "COD_CAR con MODALIDAD no resuelta (equivalencia dudosa)", int(mod_missing.sum())))
    return issues


def exportar_control_y_pes(df: pd.DataFrame, control_path: Path, pes_path: Path, issues: list[Issue], area: str) -> None:
    df.to_csv(control_path, index=False, encoding="utf-8")
    if "CODIGO_IES_NUM" not in df.columns:
        issues.append(Issue("ERROR", area, "No existe CODIGO_IES_NUM para generar pes_ready"))
        return
    pes = df.drop(columns=["CODIGO_IES_NUM"])
    pes.to_csv(pes_path, index=False, header=False, encoding="utf-8")




def _profile_column(series: pd.Series) -> dict[str, float]:
    n = len(series) if len(series) else 1
    s = series.astype(object)
    null_pct = float(s.isna().mean())
    zero_pct = float((pd.to_numeric(s, errors="coerce") == 0).fillna(False).mean())
    default_like_pct = float(s.fillna("").astype(str).str.strip().isin(["", "0", "NO", "N/A", "NA"]).mean())
    return {"null_pct": round(null_pct, 4), "zero_pct": round(zero_pct, 4), "default_like_pct": round(default_like_pct, 4)}


def generar_procedencia_y_calidad(output_dir: Path, mu: pd.DataFrame, ca: pd.DataFrame, ma: pd.DataFrame, issues: list[Issue]) -> dict[str, object]:
    provenance_rules = {
        "matricula_unificada": {
            "TIPO_DOC": "source_exact", "N_DOC": "source_exact", "DV": "source_exact", "COD_CAR": "manual_mapping",
            "MODALIDAD": "manual_mapping", "JOR": "manual_mapping", "ANIO_ING_ACT": "source_exact", "SEM_ING_ACT": "source_exact",
            "ANIO_ING_ORI": "source_exact", "SEM_ING_ORI": "source_exact", "VIG": "source_exact",
            "PAIS_EST_SEC": "missing_blocker", "REINCORPORACION": "missing_blocker"
        },
        "carreras_ac": {c: "template_preserved" for c in CARRERAS_AC_COLUMNS},
        "matricula_ac": {c: "template_preserved" for c in MATRICULA_AC_COLUMNS},
    }
    # campos calculados en matrícula AC
    for c in ["CURSO_1ER_SEM", "CURSO_2DO_SEM", "UNIDADES_CURSADAS", "UNIDADES_APROBADAS", "UNID_CURSADAS_TOTAL", "UNID_APROBADAS_TOTAL"]:
        provenance_rules["matricula_ac"][c] = "derived_rule"

    frames = {"matricula_unificada": mu, "carreras_ac": ca, "matricula_ac": ma}
    out_rows=[]
    summary={}
    for name,df in frames.items():
        rows=[]
        for col in df.columns:
            prov = provenance_rules.get(name, {}).get(col, "source_normalized")
            profile=_profile_column(df[col])
            row={"archivo":name,"columna":col,"provenance":prov,**profile}
            if prov=="missing_blocker" and profile["null_pct"]>0.2:
                issues.append(Issue("BLOCKER", name, f"{col} con missing_blocker masivo", int(profile['null_pct']*len(df))))
            if col in {"CURSO_1ER_SEM","CURSO_2DO_SEM","UNIDADES_CURSADAS","UNIDADES_APROBADAS","UNID_CURSADAS_TOTAL","UNID_APROBADAS_TOTAL"} and profile["default_like_pct"]>0.8:
                sev = "BLOCKER" if col in {"UNIDADES_CURSADAS","UNID_CURSADAS_TOTAL"} else "ERROR"
                issues.append(Issue(sev, name, f"{col} con default_like_pct alto", int(profile['default_like_pct']*len(df))))
            rows.append(row)
        t=pd.DataFrame(rows)
        summary[name]={
            "avg_null_pct": round(float(t["null_pct"].mean()),4),
            "avg_default_like_pct": round(float(t["default_like_pct"].mean()),4),
            "provenance_pct": {k: round(float((t["provenance"]==k).mean()),4) for k in t["provenance"].unique()},
        }
        out_rows.extend(rows)
    prov_df=pd.DataFrame(out_rows)
    prov_df.to_csv(output_dir/"reporte_procedencia.csv",index=False,encoding="utf-8")
    (output_dir/"reporte_calidad_semantica.json").write_text(json.dumps(summary,indent=2,ensure_ascii=False),encoding="utf-8")
    return summary

def ejecutar_pipeline(input_file: Path, output_dir: Path) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    issues: list[Issue] = []

    carreras_raw, mat_raw, hist_raw, equiv = cargar_fuentes(input_file)

    mat_i = preparar_matricula_intermedia(mat_raw)
    bridge, diag_amb = construir_puente_equiv(equiv)
    hist_map, review_nomap = mapear_historico_con_equiv(hist_raw, bridge)
    resumen = construir_resumen_historico(hist_map)

    carreras_ctrl = construir_carreras_control(carreras_raw)
    matac_ctrl = construir_matricula_ac_control(mat_i, resumen)
    mu_ctrl = construir_matricula_unificada_control(matac_ctrl, equiv)

    issues.extend(validar_carreras(carreras_ctrl))
    issues.extend(validar_matricula_ac(matac_ctrl, carreras_ctrl))
    issues.extend(validar_matricula_unificada(mu_ctrl))

    # Entregables
    mu_ctrl.to_csv(output_dir / "matricula_unificada_2026_control.csv", index=False, encoding="utf-8")
    mu_ctrl.to_excel(output_dir / "matricula_unificada_2026_oficial.xlsx", index=False)
    exportar_control_y_pes(carreras_ctrl, output_dir / "carreras_avance_curricular_2025_control.csv", output_dir / "carreras_avance_curricular_2025_pes_ready.csv", issues, "carreras")
    exportar_control_y_pes(matac_ctrl, output_dir / "matricula_avance_curricular_2025_control.csv", output_dir / "matricula_avance_curricular_2025_pes_ready.csv", issues, "matricula_ac")

    diag_amb.to_csv(output_dir / "sies_ambiguedad_diagnostico.csv", index=False, encoding="utf-8")
    review_nomap.to_csv(output_dir / "sies_codcarr_sin_mapeo.csv", index=False, encoding="utf-8")

    calidad_semantica = generar_procedencia_y_calidad(output_dir, mu_ctrl, carreras_ctrl, matac_ctrl, issues)

    report = {
        "rows": {
            "carreras_control": len(carreras_ctrl),
            "matricula_ac_control": len(matac_ctrl),
            "matricula_unificada_control": len(mu_ctrl),
            "historico_total": len(hist_raw),
            "historico_mapeado": int(hist_map["CODIGO_UNICO"].notna().sum()),
            "historico_sin_mapeo": int(hist_map["CODIGO_UNICO"].isna().sum()),
        },
        "issues": [i.__dict__ for i in issues],
        "ambiguedad": {
            "codcarr_total": int(len(diag_amb)),
            "codcarr_ambiguos": int(diag_amb["es_ambiguo"].sum()) if not diag_amb.empty else 0,
        },
        "calidad_semantica": calidad_semantica,
        "apto_oficial": {
            "matricula_unificada": not any(i.severity in {"BLOCKER","ERROR"} and i.area in {"matricula_unificada"} for i in issues),
            "avance_curricular": not any(i.severity in {"BLOCKER","ERROR"} and i.area in {"carreras", "matricula_ac", "carreras_ac"} for i in issues),
        },
    }
    (output_dir / "reporte_validacion.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return report


def generar_comparacion_versiones(output_dir: Path) -> None:
    md = """# Comparación evolutiva de pipeline

| Dimensión | Versión actual previa | Idea histórica (referencia) | Decisión híbrida aplicada |
|---|---|---|---|
| Granularidad AC | colapsaba por RUT | mantener múltiples programas | deduplicación por clave de negocio, no por RUT |
| Equivalencias | `drop_duplicates(CODCARR)` | no forzar ambiguos | mapeo por `(CODCARR,JORNADA)` y fallback solo si CODCARR único |
| Validación | schema + catálogos | auditoría por bloques | capas A/B/C + issues con severidad |
| Exportación PES | quitaba primera columna por índice | contrato explícito | quitar `CODIGO_IES_NUM` por nombre |
| Defaults sensibles | rellenos por defecto | evitar inventar | `PAIS_EST_SEC` y `REINCORPORACION` quedan BLOCKER si faltan |
| Ambigüedad SIES | diagnóstico simple | revisión consolidada | diagnóstico + archivo de `CODCARR` sin mapeo |
"""
    (output_dir / "comparacion_versiones.md").write_text(md, encoding="utf-8")


def generar_diccionario_columnas(output_dir: Path) -> None:
    rows = [
        ("MODALIDAD", "respaldada por manual"),
        ("VIG", "respaldada por manual"),
        ("COD_CAR", "respaldada por código"),
        ("ANIO_ING_ACT", "respaldada por código"),
        ("SEM_ING_ACT", "respaldada por código"),
        ("ANIO_ING_ORI", "respaldada por código"),
        ("SEM_ING_ORI", "respaldada por código"),
    ]
    known = {r[0] for r in rows}
    rows.extend((c, "inferencia operativa / pendiente de validación documental") for c in MATRICULA_UNIFICADA_COLUMNS if c not in known)
    lines = ["# Diccionario de columnas y clasificación", "", "| Columna | Clasificación |", "|---|---|"]
    lines.extend(f"| {c} | {k} |" for c, k in rows)
    (output_dir / "diccionario_columnas.md").write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Pipeline híbrido SIES/PES")
    p.add_argument("--input", default="subir prueba.xlsx")
    p.add_argument("--output-dir", default="resultados")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    out = Path(args.output_dir).resolve()
    report = ejecutar_pipeline(Path(args.input).resolve(), out)
    generar_comparacion_versiones(out)
    generar_diccionario_columnas(out)
    print(json.dumps(report, indent=2, ensure_ascii=False))

<<<<<<< ours
    columnas = [
        "RUT", "CODIGO_UNICO", "PLAN_ESTUDIOS",
        "CURSO_1ER_SEM", "CURSO_2DO_SEM",
        "UNIDADES_CURSADAS", "UNIDADES_APROBADAS",
        "UNID_CURSADAS_TOTAL", "UNID_APROBADAS_TOTAL",
        "ESTADO_ACADEMICO"
    ] + [f"UNIDADES_{n}_ANIO" for n in range(1, 8)]
    return pd.DataFrame(res, columns=columnas)

# ──────────────────────────────────────────────────────────────
# ❽ · Hoja “A. C. Carrera 2024” (estudiantes)
# ──────────────────────────────────────────────────────────────
def construir_hoja_ac_carrera2024(df_matricula, df_resumen):
    hoja = df_matricula.merge(
        df_resumen,
        left_on=["RUT_NORM", "CODIGO_UNICO", "PLAN_ESTUDIOS"],
        right_on=["RUT",       "CODIGO_UNICO", "PLAN_ESTUDIOS"],
        how="left",
        suffixes=("_MAT", "_RES")
    )

    # Aseguramos NUM_DOCUMENTO
    if "NUM_DOCUMENTO" not in hoja.columns:
        hoja["NUM_DOCUMENTO"] = df_matricula["NUM_DOCUMENTO"].values

    campos = [
        "NUM_DOCUMENTO",
        "CURSO_1ER_SEM", "CURSO_2DO_SEM",
        "UNIDADES_CURSADAS", "UNIDADES_APROBADAS",
        "UNID_CURSADAS_TOTAL", "UNID_APROBADAS_TOTAL",
        "ESTADO_ACADEMICO"
    ] + [f"UNIDADES_{n}_ANIO" for n in range(1, 8)]

    for c in campos:
        src = f"{c}_RES"
        base = hoja[c]  if c in hoja.columns else pd.Series(np.nan, index=hoja.index)
        srcs = hoja[src] if src in hoja.columns else pd.Series(np.nan, index=hoja.index)
        # 🔧 Corrección: sustituimos combine_first por where para evitar FutureWarning
        hoja[c] = base.where(~base.isna(), srcs)
        hoja.drop(columns=[src], errors="ignore", inplace=True)

    hoja.drop(columns=[col for col in hoja.columns if col.endswith("_MAT")] + ["RUT"], errors="ignore", inplace=True)

    final_cols = [col for col in df_matricula.columns if col != "VIGENCIA"] + campos + ["VIGENCIA"]
    for col in final_cols:
        if col not in hoja.columns:
            hoja[col] = np.nan

    hoja = hoja.loc[:, ~hoja.columns.duplicated()].reindex(columns=final_cols)
    hoja["VIGENCIA"] = 1
    return hoja

# ──────────────────────────────────────────────────────────────
# ❾ · Hoja “A. C. Matrícula 2024” (planes)
# ──────────────────────────────────────────────────────────────
def construir_hoja_ac_matricula2024(df_carreras, df_resumen):
    hoja = df_carreras.drop_duplicates(['CODIGO_UNICO', 'PLAN_ESTUDIOS']).copy()

    dist = [f"UNIDADES_{t}" for t in
            ["1ER_ANIO", "2DO_ANIO", "3ER_ANIO", "4TO_ANIO", "5TO_ANIO", "6TO_ANIO", "7MO_ANIO"]]

    # Si faltan distribuciones en la entrada, las completamos con el máximo observado en resumen
    agreg = (
        df_resumen
        .groupby(["CODIGO_UNICO", "PLAN_ESTUDIOS"])[[f"UNIDADES_{n}_ANIO" for n in range(1, 8)]]
        .max()
        .rename(columns={f"UNIDADES_{n}_ANIO": dist[n-1] for n in range(1, 8)})
        .reset_index()
    )
    hoja = hoja.merge(agreg, on=["CODIGO_UNICO", "PLAN_ESTUDIOS"], how="left", suffixes=("", "_RES"))

    for col in dist:
        base = hoja[col] if col in hoja.columns else pd.Series(np.nan, index=hoja.index)
        src = hoja.get(f"{col}_RES", pd.Series(np.nan, index=hoja.index))
        hoja[col] = base.where(~base.isna(), src).fillna(0)
        hoja.drop(columns=[f"{col}_RES"], errors="ignore", inplace=True)

    if "TOTAL_UNIDADES_MEDIDA" not in hoja.columns or hoja["TOTAL_UNIDADES_MEDIDA"].isna().any():
        hoja["TOTAL_UNIDADES_MEDIDA"] = hoja[dist].sum(axis=1)

    if "VIGENCIA" not in hoja.columns:
        hoja["VIGENCIA"] = 1
    else:
        hoja["VIGENCIA"].fillna(1, inplace=True)

    return hoja.loc[:, ~hoja.columns.duplicated()]

# ──────────────────────────────────────────────────────────────
# ❿ · Exportación y formateo final
# ──────────────────────────────────────────────────────────────
def exportar_archivos(hoja_carrera, hoja_matricula):
    dest_car = _ruta_content("A. C. Carrera 2024.xlsx")
    dest_mat = _ruta_content("A. C. Matrícula 2024.xlsx")

    with pd.ExcelWriter(dest_car, engine="openpyxl") as w1:
        hoja_carrera.to_excel(w1, index=False, sheet_name="A. C. Carrera 2024")
    with pd.ExcelWriter(dest_mat, engine="openpyxl") as w2:
        hoja_matricula.to_excel(w2, index=False, sheet_name="A. C. Matrícula 2024")

    formatear_excel(dest_car)
    formatear_excel(dest_mat)

# ──────────────────────────────────────────────────────────────
# ⓫ · Ejecución principal con contadores adicionales
# ──────────────────────────────────────────────────────────────
def main():
    try:
        ano_focal = int(input("🔢 Año matrícula focal (ej. 2024): ").strip())
    except ValueError:
        raise ValueError("El año de matrícula debe ser un número entero.")

    elegir_periodo = input("¿Harás un análisis por período específico? [s/n]: ").strip().lower()
    periodo_focal = None
    if elegir_periodo.startswith("s"):
        try:
            periodo_focal = int(input("🔢 Período matrícula focal (1-3): ").strip())
            if periodo_focal not in (1, 2, 3):
                raise ValueError
        except ValueError:
            raise ValueError("El período focal debe ser 1, 2 o 3.")

    ruta = cargar_excel()
    df_car, df_mat, df_hist, df_equiv = cargar_datos(ruta)
    df_car, df_mat, df_hist = preparar_datos(df_car, df_mat, df_hist, df_equiv)

    print(f"✔ Matrícula leída:               {len(df_mat)} registros")
    print(f"✔ Histórico filtrado:            {len(df_hist)} registros")
    if periodo_focal:
        print(f"✔ Año focal:                     {ano_focal} (período {periodo_focal})")
    else:
        print(f"✔ Año focal:                     {ano_focal} (año completo)")

    df_int = construir_hoja_intentos(df_hist, df_mat)
    print(f"✔ Intentos a procesar:           {len(df_int)} filas")

    errores = []
    df_res  = construir_resumen_sies(df_int, df_car, errores,
                                     ano_focal=ano_focal,
                                     periodo_focal=periodo_focal)
    print(f"✔ Estudiantes procesados:        {len(df_res)}")
    print(f"✔ Cursaron 1er semestre (SI):    {df_res['CURSO_1ER_SEM'].eq('SI').sum()}")
    print(f"✔ Cursaron 2do semestre (SI):    {df_res['CURSO_2DO_SEM'].eq('SI').sum()}")

    hoja_car = construir_hoja_ac_carrera2024(df_mat, df_res)
    hoja_mat = construir_hoja_ac_matricula2024(df_car, df_res)

    exportar_archivos(hoja_car, hoja_mat)

    if errores:
        print("\n⚠ Inconsistencias registradas:")
        for e in errores:
            print("  •", e)

    descargar_excel(_ruta_content("A. C. Carrera 2024.xlsx"))
    descargar_excel(_ruta_content("A. C. Matrícula 2024.xlsx"))
=======
>>>>>>> theirs

if __name__ == "__main__":
    main()
