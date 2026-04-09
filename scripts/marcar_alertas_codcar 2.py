from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    excel_path = Path("resultados/archivo_listo_para_sies.xlsx")
    backup_path = Path("resultados/archivo_listo_para_sies_BACKUP.xlsx")

    if not excel_path.exists():
        raise SystemExit(f"❌ No existe: {excel_path}")

    backup_path.parent.mkdir(parents=True, exist_ok=True)
    if not backup_path.exists():
        backup_path.write_bytes(excel_path.read_bytes())
        print(f"✅ Backup creado: {backup_path}")
    else:
        print(f"ℹ️ Backup ya existe: {backup_path}")

    wb = load_workbook(excel_path)
    if "ARCHIVO_LISTO_SUBIDA" not in wb.sheetnames:
        raise SystemExit("❌ No existe hoja ARCHIVO_LISTO_SUBIDA en el Excel.")

    ws = wb["ARCHIVO_LISTO_SUBIDA"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required = ["SIES_MATCH_STATUS", "SIES_MATCH_DIAG", "MANUAL_MATCH_STATUS", "COD_CAR_AUDIT_STATUS"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise SystemExit(f"❌ Faltan columnas para marcar ambigüedad COD_CAR: {missing}")

    opt = ["CODCLI", "N_DOC", "DV", "COD_CAR", "CODCARPR_NORM", "SOURCE_KEY_3", "MATCH_KEY_3", "KEY_3_NO_JORNADA"]
    opt_present = [c for c in opt if c in idx]

    def get(r, col):
        v = ws.cell(row=r, column=idx[col]).value
        return "" if v is None else str(v).strip()

    amb_sies_status = {"AMBIGUO_SIES", "SIN_MATCH_SIES"}
    amb_sies_diag = {
        "MATCH_SIES_AMBIGUO",
        "SIN_CODCARPR_EN_PUENTE_SIES",
        "PROBABLE_PROBLEMA_JORNADA_SIES",
        "PROBABLE_PROBLEMA_NOMBRE_SIES",
    }

    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Re-crear hoja ALERTAS_REVISION
    if "ALERTAS_REVISION" in wb.sheetnames:
        del wb["ALERTAS_REVISION"]
    ws2 = wb.create_sheet("ALERTAS_REVISION")
    ws2.append(["ROW", "MOTIVO"] + opt_present)

    def requiere_revision(r):
        sies_status = get(r, "SIES_MATCH_STATUS").upper()
        sies_diag = get(r, "SIES_MATCH_DIAG").upper()
        manual_status = get(r, "MANUAL_MATCH_STATUS").upper()
        codcar_audit = get(r, "COD_CAR_AUDIT_STATUS").upper()

        if sies_status in amb_sies_status:
            return True, f"SIES_MATCH_STATUS={sies_status}"
        if sies_diag in amb_sies_diag:
            return True, f"SIES_MATCH_DIAG={sies_diag}"
        if manual_status == "SIN_MATCH_MANUAL":
            return True, "MANUAL_MATCH_STATUS=SIN_MATCH_MANUAL"
        if codcar_audit == "SIN_FUENTE_FINAL":
            return True, "COD_CAR_AUDIT_STATUS=SIN_FUENTE_FINAL"
        return False, ""

    marked = 0
    for r in range(2, ws.max_row + 1):
        ok, motivo = requiere_revision(r)
        if ok:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = red
            marked += 1
            ws2.append([r, motivo] + [get(r, c) for c in opt_present])

    ws2.append([])
    ws2.append(["RESUMEN", f"FILAS_MARCADAS_EN_ROJO={marked}"])

    wb.save(excel_path)
    print(f"✅ Listo: {marked} filas marcadas en rojo. Hoja ALERTAS_REVISION creada.")
    print(f"📄 Excel actualizado: {excel_path}")

if __name__ == "__main__":
    main()
