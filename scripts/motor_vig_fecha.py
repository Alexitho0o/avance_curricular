#!/usr/bin/env python3
"""
Motor de derivación VIG + FECHA_MATRICULA — MU 2026
Campos: AF VIG (Vigencia), AD FECHA_MATRICULA

Lee DatosAlumnos + base_datos (filtro) + catálogo gobernanza,
aplica reglas normativas (config JSON) y genera:
  - vig_fecha_trace_long.tsv       (trazabilidad por registro)
  - AUDIT_VIG_FECHA.xlsx           (auditoría con formato visual)
  - vig_fecha_governance_report.md (reporte + dictamen)
"""
import json, os, sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np

# ── paths ────────────────────────────────────────────────────────────────
BASE     = Path(__file__).resolve().parent.parent          # avance_curricular/
EXCEL_IN = Path(os.environ.get(
    "EXCEL_INPUT",
    "/Users/alexi/Downloads/PROMEDIOSDEALUMNOS_7804.xlsx"))
CFG_PATH = BASE / "control" / "config_vig_fecha.json"
GOB_EA   = BASE / "gobernanza_catalogos" / "gob_datosalumnos_estadoacademico_situacion.tsv"
OUT_DIR  = BASE / "resultados"
CTRL_DIR = BASE / "control"

TS = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
FECHA_CORTE = pd.Timestamp("2026-04-30")
FECHA_RANGO_MIN = pd.Timestamp("2020-01-01")
FALLBACK_1900 = "01/01/1900"

# ── load config ──────────────────────────────────────────────────────────
with open(CFG_PATH) as f:
    CFG = json.load(f)

VIG_CATALOGO = set(CFG["campos"]["VIG"]["catalogo"])
VIG_MAPEO = CFG["campos"]["VIG"]["mapeo_estadoacademico"]
VIG_CONTAINS = CFG["campos"]["VIG"]["mapeo_contains"]
VIG_FALLBACK = CFG["campos"]["VIG"]["fallback"]

# ── load gobernanza catálogo ─────────────────────────────────────────────
def _load_gob_catalogo():
    """Carga catálogo ESTADOACADEMICO → VIG_ESPERADO."""
    if GOB_EA.exists():
        df = pd.read_csv(GOB_EA, sep="\t")
        return dict(zip(df["ESTADOACADEMICO"].str.upper().str.strip(),
                        df["VIG_ESPERADO"].astype(int)))
    return {}

GOB_VIG_MAP = _load_gob_catalogo()


# ═══════════════════════════════════════════════════════════════════════════
# 1. CARGA
# ═══════════════════════════════════════════════════════════════════════════

def load_data():
    """Carga DatosAlumnos filtrado por base_datos RUTs."""
    xls = pd.ExcelFile(EXCEL_IN)
    da  = pd.read_excel(xls, sheet_name="DatosAlumnos")
    bd  = pd.read_excel(xls, sheet_name="base_datos")

    ruts_bd = set(bd["N_DOC"].dropna().astype(int))

    da["_RUT_NUM"] = (da["RUT"].astype(str)
                      .str.extract(r"(\d+)", expand=False)
                      .astype(float).astype("Int64"))

    da_f = da[da["_RUT_NUM"].isin(ruts_bd)].copy()

    # Cast numéricos
    for col in ["ANOINGRESO", "PERIODOINGRESO", "ANOMATRICULA", "PERIODOMATRICULA"]:
        if col in da_f.columns:
            da_f[col] = pd.to_numeric(da_f[col], errors="coerce").astype("Int64")

    return da_f


# ═══════════════════════════════════════════════════════════════════════════
# 2. DERIVACIÓN VIG  (Campo AF)
# ═══════════════════════════════════════════════════════════════════════════

def derive_vig(da: pd.DataFrame) -> pd.DataFrame:
    """
    Regla: ESTADOACADEMICO → VIG (0/1/2)
    Cascada:
      1. Mapeo exacto por ESTADOACADEMICO
      2. Mapeo parcial (contains) si no hay match exacto
      3. Fallback VIG=1 si valor desconocido o nulo
    """
    da["VIG"] = pd.NA
    da["VIG_FUENTE"] = ""
    da["VIG_REGLA"] = ""

    for idx, row in da.iterrows():
        ea = str(row.get("ESTADOACADEMICO", "")).strip().upper() if pd.notna(row.get("ESTADOACADEMICO")) else ""

        # Caso 1: ESTADOACADEMICO nulo/vacío → fallback
        if not ea:
            da.at[idx, "VIG"] = VIG_FALLBACK
            da.at[idx, "VIG_FUENTE"] = "DEFAULT"
            da.at[idx, "VIG_REGLA"] = "R_VIG_10"
            continue

        # Caso 2: Mapeo exacto
        if ea in VIG_MAPEO:
            da.at[idx, "VIG"] = VIG_MAPEO[ea]
            da.at[idx, "VIG_FUENTE"] = "DA_ESTADOACADEMICO"
            # Determinar regla
            rule_map = {
                "VIGENTE": "R_VIG_01", "ELIMINADO": "R_VIG_02",
                "SUSPENDIDO": "R_VIG_03", "EGRESADO": "R_VIG_04",
                "TITULADO": "R_VIG_05",
            }
            da.at[idx, "VIG_REGLA"] = rule_map.get(ea, "R_VIG_01")
            continue

        # Caso 3: Mapeo parcial (contains)
        matched = False
        for pattern, vig_val in VIG_CONTAINS.items():
            if pattern.upper() in ea:
                da.at[idx, "VIG"] = vig_val
                da.at[idx, "VIG_FUENTE"] = "DA_ESTADOACADEMICO_CONTAINS"
                # Determinar regla por pattern
                contains_rule = {
                    "ALUMNO REGULAR": "R_VIG_06", "REINCORPORACION": "R_VIG_07",
                    "PROCESO DE TITULO": "R_VIG_08", "TITULACION": "R_VIG_09",
                    "ELIMINADO": "R_VIG_02", "SUSPENDIDO": "R_VIG_03",
                    "VIGENTE": "R_VIG_01", "EGRESADO": "R_VIG_04",
                }
                da.at[idx, "VIG_REGLA"] = contains_rule.get(pattern.upper(), "R_VIG_11")
                matched = True
                break

        if not matched:
            # Caso 4: ESTADOACADEMICO no reconocido → fallback
            da.at[idx, "VIG"] = VIG_FALLBACK
            da.at[idx, "VIG_FUENTE"] = "DEFAULT_NO_RECONOCIDO"
            da.at[idx, "VIG_REGLA"] = "R_VIG_11"

    da["VIG"] = da["VIG"].astype(int)
    return da


# ═══════════════════════════════════════════════════════════════════════════
# 3. DERIVACIÓN FECHA_MATRICULA  (Campo AD)
# ═══════════════════════════════════════════════════════════════════════════

def derive_fecha_matricula(da: pd.DataFrame) -> pd.DataFrame:
    """
    Cascada:
      1. FECHAMATRICULA parseable → formatear dd/mm/yyyy
      2. Fallback → 01/01/1900
    """
    da["FECHA_MATRICULA"] = ""
    da["FECHA_MATRICULA_DT"] = pd.NaT
    da["FECHA_MATRICULA_FUENTE"] = ""
    da["FECHA_MATRICULA_REGLA"] = ""

    for idx, row in da.iterrows():
        raw_fm = row.get("FECHAMATRICULA")

        # Caso 1: FECHAMATRICULA parseable
        dt = pd.NaT
        if pd.notna(raw_fm):
            dt = pd.to_datetime(raw_fm, errors="coerce", dayfirst=True)

        if pd.notna(dt):
            da.at[idx, "FECHA_MATRICULA"] = dt.strftime("%d/%m/%Y")
            da.at[idx, "FECHA_MATRICULA_DT"] = dt
            da.at[idx, "FECHA_MATRICULA_FUENTE"] = "DA_FECHAMATRICULA"
            da.at[idx, "FECHA_MATRICULA_REGLA"] = "R_FM_01"
            continue

        # Caso 2: Fallback 01/01/1900
        da.at[idx, "FECHA_MATRICULA"] = FALLBACK_1900
        da.at[idx, "FECHA_MATRICULA_DT"] = pd.Timestamp("1900-01-01")
        da.at[idx, "FECHA_MATRICULA_FUENTE"] = "DEFAULT"
        da.at[idx, "FECHA_MATRICULA_REGLA"] = "R_FM_03"

    return da


# ═══════════════════════════════════════════════════════════════════════════
# 4. VALIDACIONES
# ═══════════════════════════════════════════════════════════════════════════

def run_validations(da: pd.DataFrame) -> list[dict]:
    """Ejecuta validaciones normativas y devuelve hallazgos."""
    findings = []

    # V_CATALOGO_VIG: VIG debe ser 0, 1 o 2
    fuera = da[~da["VIG"].isin(VIG_CATALOGO)]
    if len(fuera) > 0:
        findings.append({
            "id": "V_CATALOGO_VIG", "severidad": "BLOQUEANTE",
            "msg": f"{len(fuera)} registros con VIG fuera de catálogo {VIG_CATALOGO}",
            "n": len(fuera), "codclis": fuera["CODCLI"].tolist()[:10]
        })

    # V_COHERENCIA_VIG_ESTADO: VIG coherente con catálogo gobernanza
    incoherentes = []
    for idx, row in da.iterrows():
        ea = str(row.get("ESTADOACADEMICO", "")).strip().upper() if pd.notna(row.get("ESTADOACADEMICO")) else ""
        if ea in GOB_VIG_MAP:
            esperado = GOB_VIG_MAP[ea]
            actual = int(row["VIG"])
            if actual != esperado:
                incoherentes.append({
                    "CODCLI": row["CODCLI"],
                    "ESTADOACADEMICO": ea,
                    "VIG_ACTUAL": actual,
                    "VIG_ESPERADO": esperado,
                })
    if incoherentes:
        findings.append({
            "id": "V_COHERENCIA_VIG_ESTADO", "severidad": "BLOQUEANTE",
            "msg": f"{len(incoherentes)} registros con VIG incoherente con catálogo gobernanza",
            "n": len(incoherentes),
            "codclis": [r["CODCLI"] for r in incoherentes[:10]],
            "detalle": incoherentes[:5],
        })

    # V_FORMATO_FECHA: FECHA_MATRICULA en formato dd/mm/yyyy
    import re
    pat = re.compile(r"^\d{2}/\d{2}/\d{4}$")
    bad_fmt = da[~da["FECHA_MATRICULA"].apply(lambda x: bool(pat.match(str(x))))]
    if len(bad_fmt) > 0:
        findings.append({
            "id": "V_FORMATO_FECHA", "severidad": "BLOQUEANTE",
            "msg": f"{len(bad_fmt)} registros con FECHA_MATRICULA en formato inválido",
            "n": len(bad_fmt), "codclis": bad_fmt["CODCLI"].tolist()[:10]
        })

    # V_FECHA_NO_FUTURA: FECHA_MATRICULA no debe superar fecha de corte
    dt_col = da["FECHA_MATRICULA_DT"]
    futuras = da[dt_col > FECHA_CORTE]
    if len(futuras) > 0:
        findings.append({
            "id": "V_FECHA_NO_FUTURA", "severidad": "WARNING",
            "msg": f"{len(futuras)} registros con FECHA_MATRICULA posterior a {FECHA_CORTE.date()}",
            "n": len(futuras), "codclis": futuras["CODCLI"].tolist()[:10]
        })

    # V_FECHA_RANGO: FECHA_MATRICULA entre 2020-01-01 y corte (excluyendo 1900)
    reales = da[dt_col > pd.Timestamp("1901-01-01")]
    fuera_rango = reales[(dt_col < FECHA_RANGO_MIN)]
    if len(fuera_rango) > 0:
        findings.append({
            "id": "V_FECHA_RANGO", "severidad": "WARNING",
            "msg": f"{len(fuera_rango)} registros con FECHA_MATRICULA anterior a {FECHA_RANGO_MIN.date()}",
            "n": len(fuera_rango), "codclis": fuera_rango["CODCLI"].tolist()[:10]
        })

    # V_NULOS
    for campo in ["VIG", "FECHA_MATRICULA"]:
        nulos = da[da[campo].isna() | (da[campo].astype(str).str.strip() == "")]
        if len(nulos) > 0:
            findings.append({
                "id": f"V_NULO_{campo}", "severidad": "BLOQUEANTE",
                "msg": f"{len(nulos)} registros con {campo} nulo/vacío",
                "n": len(nulos)
            })

    # V_VIG0_MARCADO: Info sobre filas ELIMINADO/SUSPENDIDO
    vig0 = da[da["VIG"] == 0]
    if len(vig0) > 0:
        findings.append({
            "id": "V_VIG0_MARCADO_ROJO", "severidad": "INFO",
            "msg": f"{len(vig0)} registros con VIG=0 (ELIMINADO/SUSPENDIDO) → marcar en rojo en auditoría",
            "n": len(vig0),
            "detalle_ea": vig0["ESTADOACADEMICO"].value_counts().to_dict(),
        })

    return findings


# ═══════════════════════════════════════════════════════════════════════════
# 5. ARTEFACTOS
# ═══════════════════════════════════════════════════════════════════════════

TRACE_COLS = [
    "CODCLI", "_RUT_NUM",
    "ESTADOACADEMICO", "SITUACION",
    "FECHAMATRICULA",
    "VIG", "VIG_FUENTE", "VIG_REGLA",
    "FECHA_MATRICULA", "FECHA_MATRICULA_FUENTE", "FECHA_MATRICULA_REGLA",
]


def write_trace_tsv(da: pd.DataFrame, path: Path):
    cols = [c for c in TRACE_COLS if c in da.columns]
    da[cols].to_csv(path, sep="\t", index=False)


def write_audit_xlsx(da: pd.DataFrame, findings: list, path: Path):
    """Genera Excel de auditoría con formato visual (rojo para VIG=0)."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1: TRAZABILIDAD
        cols = [c for c in TRACE_COLS if c in da.columns]
        da[cols].to_excel(writer, sheet_name="TRAZABILIDAD", index=False)

        # Sheet 2: RESUMEN_VIG
        vig_dist = da.groupby("VIG").agg(
            N=("CODCLI", "count"),
            REGLAS=("VIG_REGLA", lambda x: ", ".join(sorted(x.unique()))),
        ).reset_index()
        vig_dist.to_excel(writer, sheet_name="RESUMEN_VIG", index=False)

        # Sheet 3: RESUMEN_FECHA
        da["_FM_YEAR"] = da["FECHA_MATRICULA_DT"].dt.year
        fm_dist = da.groupby("_FM_YEAR").agg(
            N=("CODCLI", "count"),
        ).reset_index()
        fm_dist.to_excel(writer, sheet_name="RESUMEN_FECHA", index=False)

        # Sheet 4: CRUCE_VIG_x_ESTADO
        cruce = pd.crosstab(da["ESTADOACADEMICO"], da["VIG"], margins=True)
        cruce.to_excel(writer, sheet_name="CRUCE_VIG_x_ESTADO")

        # Sheet 5: HALLAZGOS
        if findings:
            hall_df = pd.DataFrame([
                {"ID": f["id"], "SEVERIDAD": f["severidad"], "MENSAJE": f["msg"], "N": f["n"]}
                for f in findings
            ])
        else:
            hall_df = pd.DataFrame({"ID": ["NINGUNO"], "SEVERIDAD": ["OK"], "MENSAJE": ["Sin hallazgos"], "N": [0]})
        hall_df.to_excel(writer, sheet_name="HALLAZGOS", index=False)

    # Aplicar formato rojo a filas VIG=0
    _apply_red_formatting(path, da)


def _apply_red_formatting(path: Path, da: pd.DataFrame):
    """Marca en rojo las filas con VIG=0 en la hoja TRAZABILIDAD."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font

        wb = load_workbook(path)
        ws = wb["TRAZABILIDAD"]

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")

        # Encontrar columna VIG
        vig_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "VIG":
                vig_col_idx = col_idx
                break

        if vig_col_idx is None:
            return

        # Marcar filas con VIG=0
        for row_idx in range(2, ws.max_row + 1):
            vig_val = ws.cell(row=row_idx, column=vig_col_idx).value
            if vig_val is not None and int(vig_val) == 0:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = red_fill
                    cell.font = red_font

        wb.save(path)
    except ImportError:
        pass  # openpyxl no disponible


def write_governance_report(da: pd.DataFrame, findings: list, path: Path):
    """Genera reporte Markdown de gobernanza."""
    n = len(da)
    bloq = [f for f in findings if f["severidad"] == "BLOQUEANTE"]
    warn = [f for f in findings if f["severidad"] == "WARNING"]
    info = [f for f in findings if f["severidad"] == "INFO"]

    if bloq:
        dictamen = "🔴 BLOQUEADO"
    elif warn:
        dictamen = "⚠️ LISTO CON OBSERVACIONES"
    else:
        dictamen = "✅ LISTO"

    lines = [
        f"# Reporte de Gobernanza — VIG + FECHA_MATRICULA",
        f"",
        f"**Motor**: motor_vig_fecha.py v1.0",
        f"**Fecha**: {TS}",
        f"**Registros**: {n}",
        f"**Dictamen**: {dictamen}",
        f"",
        f"---",
        f"",
        f"## 1. Distribución VIG",
        f"",
        f"| VIG | Significado | N | % |",
        f"|-----|------------|---|---|",
    ]

    for vig_val in sorted(da["VIG"].unique()):
        count = (da["VIG"] == vig_val).sum()
        desc = CFG["campos"]["VIG"]["descripcion_catalogo"].get(str(vig_val), "?")
        lines.append(f"| {vig_val} | {desc} | {count} | {count/n*100:.1f}% |")

    lines += [
        f"",
        f"## 2. Reglas VIG aplicadas",
        f"",
        f"| Regla | N | % |",
        f"|-------|---|---|",
    ]
    for regla, count in da["VIG_REGLA"].value_counts().items():
        lines.append(f"| {regla} | {count} | {count/n*100:.1f}% |")

    lines += [
        f"",
        f"## 3. Cruce ESTADOACADEMICO × VIG",
        f"",
    ]
    cruce = pd.crosstab(da["ESTADOACADEMICO"], da["VIG"])
    # Manual markdown table (avoid tabulate dependency)
    cruce_lines = ["| ESTADOACADEMICO | " + " | ".join(str(c) for c in cruce.columns) + " |"]
    cruce_lines.append("|" + "|".join(["---"] * (len(cruce.columns) + 1)) + "|")
    for ea, row_data in cruce.iterrows():
        cruce_lines.append(f"| {ea} | " + " | ".join(str(v) for v in row_data) + " |")
    lines.append("\n".join(cruce_lines))

    lines += [
        f"",
        f"## 4. Distribución FECHA_MATRICULA",
        f"",
        f"| Año | N | % |",
        f"|-----|---|---|",
    ]
    year_dist = da["FECHA_MATRICULA_DT"].dt.year.value_counts().sort_index()
    for yr, count in year_dist.items():
        lines.append(f"| {yr} | {count} | {count/n*100:.1f}% |")

    lines += [
        f"",
        f"## 5. Reglas FECHA_MATRICULA aplicadas",
        f"",
        f"| Regla | N | % |",
        f"|-------|---|---|",
    ]
    for regla, count in da["FECHA_MATRICULA_REGLA"].value_counts().items():
        lines.append(f"| {regla} | {count} | {count/n*100:.1f}% |")

    lines += [
        f"",
        f"## 6. Hallazgos",
        f"",
    ]
    if not findings:
        lines.append("Sin hallazgos.")
    else:
        for f in findings:
            sev = f["severidad"]
            lines.append(f"- **[{sev}]** {f['id']}: {f['msg']}")

    lines += [
        f"",
        f"---",
        f"",
        f"## Dictamen Final",
        f"",
        f"**{dictamen}**",
        f"",
        f"- BLOQUEANTES: {len(bloq)}",
        f"- WARNINGS: {len(warn)}",
        f"- INFO: {len(info)}",
    ]

    path.write_text("\n".join(lines), encoding="utf-8")
    return dictamen


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print(f"{'═'*60}")
    print(f" Motor VIG + FECHA_MATRICULA — MU 2026")
    print(f" {TS}")
    print(f"{'═'*60}")

    # 1. Carga
    da = load_data()
    print(f"\n✅ Cargados {len(da)} registros filtrados")
    print(f"   ESTADOACADEMICO distribución:")
    for ea, count in da["ESTADOACADEMICO"].value_counts().items():
        print(f"     {ea}: {count}")

    # 2. Derivar VIG
    da = derive_vig(da)
    print(f"\n✅ VIG derivado:")
    for vig_val, count in da["VIG"].value_counts().sort_index().items():
        print(f"     VIG={vig_val}: {count}")

    # 3. Derivar FECHA_MATRICULA
    da = derive_fecha_matricula(da)
    n_1900 = (da["FECHA_MATRICULA"] == FALLBACK_1900).sum()
    n_real = len(da) - n_1900
    print(f"\n✅ FECHA_MATRICULA derivado:")
    print(f"     Fecha real: {n_real}")
    print(f"     Fallback 1900: {n_1900}")

    # 4. Validaciones
    findings = run_validations(da)
    print(f"\n{'─'*40}")
    print(f" Validaciones: {len(findings)} hallazgos")
    for f in findings:
        print(f"   [{f['severidad']}] {f['id']}: {f['msg']}")

    # 5. Artefactos
    CTRL_DIR.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    trace_path = CTRL_DIR / "vig_fecha_trace_long.tsv"
    write_trace_tsv(da, trace_path)
    print(f"\n📄 {trace_path}")

    audit_path = OUT_DIR / "AUDIT_VIG_FECHA.xlsx"
    write_audit_xlsx(da, findings, audit_path)
    print(f"📄 {audit_path}")

    report_path = CTRL_DIR / "vig_fecha_governance_report.md"
    dictamen = write_governance_report(da, findings, report_path)
    print(f"📄 {report_path}")

    # Golden cases
    golden = _generate_golden_cases(da)
    golden_path = CTRL_DIR / "vig_fecha_golden_cases.json"
    with open(golden_path, "w") as f:
        json.dump(golden, f, indent=2, ensure_ascii=False, default=str)
    print(f"📄 {golden_path}")

    print(f"\n{'═'*60}")
    print(f" ═══ DICTAMEN: {dictamen} ═══")
    print(f"{'═'*60}")

    bloq = [f for f in findings if f["severidad"] == "BLOQUEANTE"]
    return 1 if bloq else 0


def _generate_golden_cases(da: pd.DataFrame) -> list[dict]:
    """Genera golden cases representativos."""
    cases = []

    # GC01: VIGENTE típico
    vig = da[da["ESTADOACADEMICO"] == "VIGENTE"].head(1)
    if len(vig) > 0:
        r = vig.iloc[0]
        cases.append({
            "id": "GC01", "desc": "VIGENTE → VIG=1",
            "CODCLI": r["CODCLI"], "ESTADOACADEMICO": r["ESTADOACADEMICO"],
            "VIG_ESPERADO": 1, "VIG_REAL": int(r["VIG"]),
            "FECHA_MATRICULA": r["FECHA_MATRICULA"],
        })

    # GC02: ELIMINADO
    elim = da[da["ESTADOACADEMICO"] == "ELIMINADO"].head(1)
    if len(elim) > 0:
        r = elim.iloc[0]
        cases.append({
            "id": "GC02", "desc": "ELIMINADO → VIG=0",
            "CODCLI": r["CODCLI"], "ESTADOACADEMICO": r["ESTADOACADEMICO"],
            "VIG_ESPERADO": 0, "VIG_REAL": int(r["VIG"]),
            "FECHA_MATRICULA": r["FECHA_MATRICULA"],
        })

    # GC03: SUSPENDIDO
    susp = da[da["ESTADOACADEMICO"] == "SUSPENDIDO"].head(1)
    if len(susp) > 0:
        r = susp.iloc[0]
        cases.append({
            "id": "GC03", "desc": "SUSPENDIDO → VIG=0",
            "CODCLI": r["CODCLI"], "ESTADOACADEMICO": r["ESTADOACADEMICO"],
            "VIG_ESPERADO": 0, "VIG_REAL": int(r["VIG"]),
            "FECHA_MATRICULA": r["FECHA_MATRICULA"],
        })

    # GC04: FECHA más antigua (2025)
    da_sorted = da.sort_values("FECHA_MATRICULA_DT")
    old = da_sorted[da_sorted["FECHA_MATRICULA_DT"] > pd.Timestamp("1901-01-01")].head(1)
    if len(old) > 0:
        r = old.iloc[0]
        cases.append({
            "id": "GC04", "desc": "Fecha matrícula más antigua",
            "CODCLI": r["CODCLI"], "FECHA_MATRICULA": r["FECHA_MATRICULA"],
            "FECHA_MATRICULA_REGLA": r["FECHA_MATRICULA_REGLA"],
        })

    # GC05: FECHA más reciente
    recent = da_sorted.tail(1)
    if len(recent) > 0:
        r = recent.iloc[0]
        cases.append({
            "id": "GC05", "desc": "Fecha matrícula más reciente",
            "CODCLI": r["CODCLI"], "FECHA_MATRICULA": r["FECHA_MATRICULA"],
            "FECHA_MATRICULA_REGLA": r["FECHA_MATRICULA_REGLA"],
        })

    # GC06: ELIMINADO con fecha 2026
    elim_2026 = da[(da["ESTADOACADEMICO"] == "ELIMINADO") &
                   (da["FECHA_MATRICULA_DT"].dt.year == 2026)].head(1)
    if len(elim_2026) > 0:
        r = elim_2026.iloc[0]
        cases.append({
            "id": "GC06", "desc": "ELIMINADO con fecha 2026 → VIG=0 + fecha real",
            "CODCLI": r["CODCLI"], "ESTADOACADEMICO": r["ESTADOACADEMICO"],
            "VIG_ESPERADO": 0, "VIG_REAL": int(r["VIG"]),
            "FECHA_MATRICULA": r["FECHA_MATRICULA"],
        })

    return cases


if __name__ == "__main__":
    sys.exit(main())
